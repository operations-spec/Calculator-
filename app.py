from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_from_directory, abort, make_response
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from functools import wraps
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired, Email
from waitress import serve
import os
import json
from datetime import datetime, timedelta, timezone
from io import BytesIO
import uuid
import hashlib
import secrets

try:
    from weasyprint import HTML
except Exception:
    HTML = None
import re
import random
import time
import threading
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from bson.objectid import ObjectId
import logging
import traceback
import resend
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from markupsafe import Markup, escape
import base64

# Import MongoDB users module
try:
    from mongo_users import (
        find_user_by_id as mu_find_user_by_id,
        find_user_by_email_or_username as mu_find_user_by_email_or_username,
        create_user as mu_create_user,
        verify_password as mu_verify_password,
        update_user as mu_update_user,
        users_col
    )
    MONGO_AVAILABLE = True
except (ImportError, RuntimeError) as e:
    print(f"MongoDB module not available: {e}")
    MONGO_AVAILABLE = False
    users_col = None

# Load environment variables
load_dotenv()

# Debug environment variables
print("\n=== Environment Variables ===")
print(f"MONGO_URI: {'Set' if os.getenv('MONGO_URI') else 'Not set'}")
print(f"DB_NAME: {os.getenv('DB_NAME', 'moneda_db')}")
print(f"USE_MONGO: {os.getenv('USE_MONGO', 'Not set')}")
print(f"WA_SERVICE_URL: {os.getenv('WA_SERVICE_URL', 'Not set')}")
print(f"WA_SERVICE_AUTH_TOKEN: {'Set' if os.getenv('WA_SERVICE_AUTH_TOKEN') else 'Not set'}")
print("===========================\n")

# WhatsApp service configuration
WA_SERVICE_URL = os.getenv('WA_SERVICE_URL', '').rstrip('/')
WA_SERVICE_AUTH_TOKEN = os.getenv('WA_SERVICE_AUTH_TOKEN', '')

def send_whatsapp_message(to_phone, body, attachment=None):
    """Send WhatsApp message via wa_service.
    
    Args:
        to_phone: Phone number (will be normalized to E164)
        body: Message text
        attachment: Optional dict with {filename, content (base64), type}
    
    Returns:
        bool: True if sent successfully
    """
    if not WA_SERVICE_URL or not WA_SERVICE_AUTH_TOKEN:
        app.logger.warning("WhatsApp service not configured (WA_SERVICE_URL or WA_SERVICE_AUTH_TOKEN missing)")
        return False
    
    # Normalize phone to E164
    phone = (to_phone or '').strip()
    if not phone:
        return False
    if phone.startswith('+'):
        phone = phone
    elif len(phone) == 10 and phone.isdigit():
        phone = f"+91{phone}"
    elif len(phone) > 10 and phone.startswith('91'):
        phone = f"+{phone}"
    else:
        phone = f"+{phone.replace('+', '')}"
    
    payload = {
        'to': phone,
        'body': body
    }
    if attachment:
        payload['attachment'] = attachment
    
    try:
        import requests as _requests
        import time as _time

        max_attempts = int(os.getenv('WA_SEND_MAX_ATTEMPTS', '3') or '3')
        connect_timeout = float(os.getenv('WA_SEND_CONNECT_TIMEOUT', '10') or '10')
        read_timeout = float(os.getenv('WA_SEND_READ_TIMEOUT', '45') or '45')
        backoff_seconds = float(os.getenv('WA_SEND_BACKOFF_SECONDS', '2') or '2')

        last_error = None
        for attempt in range(1, max_attempts + 1):
            try:
                resp = _requests.post(
                    f"{WA_SERVICE_URL}/send",
                    json=payload,
                    headers={
                        'Authorization': f"Bearer {WA_SERVICE_AUTH_TOKEN}",
                        'Content-Type': 'application/json'
                    },
                    timeout=(connect_timeout, read_timeout)
                )

                if resp.status_code == 200:
                    app.logger.info(f"WhatsApp message sent to {phone}")
                    return True

                last_error = f"HTTP {resp.status_code}: {resp.text}"
                app.logger.warning(f"WhatsApp send failed (attempt {attempt}/{max_attempts}): {last_error}")
            except Exception as e:
                last_error = str(e)
                app.logger.error(f"Error sending WhatsApp message (attempt {attempt}/{max_attempts}): {e}")

            if attempt < max_attempts:
                _time.sleep(backoff_seconds * attempt)

        app.logger.warning(f"WhatsApp send ultimately failed after {max_attempts} attempts: {last_error}")
        return False
    except Exception as e:
        app.logger.error(f"Error sending WhatsApp message: {e}")
        return False

# Timezone for sequential IDs
IST = timezone(timedelta(hours=5, minutes=30))

def get_india_time():
    """Return current IST-aware datetime."""
    return datetime.now(IST)

def get_next_quote_id():
    """Return next sequential quotation id like CGI_Q0001, CGI_Q0002 ..."""
    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
        try:
            counter = mongo_db.counters.find_one_and_update(
                {'_id': 'quotation_sequence'},
                {'$inc': {'seq': 1}},
                upsert=True,
                return_document=True
            )
            seq = counter.get('seq', 1)
            if seq <= 0:
                seq = mongo_db.quotations.count_documents({'quote_id': {'$regex': r'^CGI_Q\d+$'}}) + 1
                mongo_db.counters.update_one({'_id': 'quotation_sequence'}, {'$set': {'seq': seq}}, upsert=True)
            return f"CGI_Q{seq:04d}"
        except Exception as e:
            app.logger.error(f"Failed to generate sequential quotation id: {e}")
    return f"CGI_Q{int(datetime.now(IST).timestamp())}"


def to_float(value):
    """Safely convert the provided value to float, returning None on failure."""
    try:
        if value is None or value == '':
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def patch_rule_metadata(product, payload=None):
    """Ensure rule-specific metadata fields exist on the product dict."""
    if product is None:
        return {}

    payload = payload or {}

    length = to_float(payload.get('length_per_unit_m'))
    if length is None or length <= 0:
        length = to_float(product.get('length_per_unit_m'))
    if length is None or length <= 0:
        length = 100.0

    rate = to_float(payload.get('rate_per_meter'))
    if rate is None or rate < 0:
        rate = to_float(product.get('rate_per_meter'))
    if rate is None or rate < 0:
        rate = 0.0

    total_length = to_float(payload.get('total_length_m'))
    if total_length is None:
        total_length = to_float(product.get('total_length_m'))

    quantity = payload.get('quantity')
    try:
        quantity = int(quantity)
    except (TypeError, ValueError):
        try:
            quantity = int(product.get('quantity', 1))
        except (TypeError, ValueError):
            quantity = 1
    if quantity < 1:
        quantity = 1

    if total_length is None:
        total_length = length * quantity

    product['length_per_unit_m'] = length
    product['rate_per_meter'] = rate
    product['quantity'] = quantity
    product['total_length_m'] = round(total_length, 2)

    for key in ['rule_category', 'profile_id', 'profile_label', 'profile_code', 'packaging', 'packaging_type']:
        value = payload.get(key)
        if value is not None:
            product[key] = value
        elif key not in product:
            product[key] = None

    return product


def recalc_rule_pricing(product):
    """Recalculate pricing totals for a rule product in-place."""
    if product is None:
        return {}

    patch_rule_metadata(product)

    length = to_float(product.get('length_per_unit_m')) or 100.0
    if length <= 0:
        length = 100.0

    rate = to_float(product.get('rate_per_meter')) or 0.0
    if rate < 0:
        rate = 0.0

    try:
        quantity = int(product.get('quantity', 1))
    except (TypeError, ValueError):
        quantity = 1
    if quantity < 1:
        quantity = 1

    discount_percent = to_float(product.get('discount_percent')) or 0.0
    discount_percent = max(0.0, min(discount_percent, 100.0))

    gst_percent = to_float(product.get('gst_percent'))
    if gst_percent is None or gst_percent < 0:
        gst_percent = 18.0

    unit_price = to_float(product.get('unit_price'))
    if unit_price is None or unit_price <= 0:
        unit_price = length * rate

    subtotal = unit_price * quantity
    discount_amount = subtotal * (discount_percent / 100)
    discounted_subtotal = subtotal - discount_amount
    gst_amount = (discounted_subtotal * gst_percent) / 100
    final_total = discounted_subtotal + gst_amount
    total_length = length * quantity

    product['unit_price'] = round(unit_price, 2)
    product['total_price'] = round(final_total, 2)
    product['quantity'] = quantity
    product['discount_percent'] = discount_percent
    product['gst_percent'] = gst_percent
    product['length_per_unit_m'] = length
    product['rate_per_meter'] = rate
    product['total_length_m'] = round(total_length, 2)

    product['calculations'] = {
        'unit_price': round(unit_price, 2),
        'quantity': quantity,
        'subtotal': round(subtotal, 2),
        'discount_percent': discount_percent,
        'discount_amount': round(discount_amount, 2),
        'discounted_subtotal': round(discounted_subtotal, 2),
        'gst_percent': gst_percent,
        'gst_amount': round(gst_amount, 2),
        'final_total': round(final_total, 2)
    }

    return product

# CORS Configuration
from flask_cors import CORS
# Initialize Flask app and login manager
app = Flask(__name__)

# Include 'templates/user' in the Jinja2 search path so that render_template('cart.html')
# and similar calls find templates stored under that sub-directory without changing
# every individual render_template path.
app.jinja_loader.searchpath.append(os.path.join(app.root_path, 'templates', 'user'))
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# -------------------- Helpers --------------------

def reset_company_selection_session():
    """Clear any stored company selection data so the user must re-select."""
    cleared = False
    for key in ('selected_company', 'company_id', 'company_name', 'company_email', 'company', 'email'):
        if key in session:
            session.pop(key, None)
            cleared = True
    if cleared:
        session.modified = True


def clear_login_prompt_flash():
    """Remove the automatic 'please log in' flash from Flask-Login if it exists."""
    flashes = session.get('_flashes')
    if not flashes:
        return

    filtered = [
        flash for flash in flashes
        if not (
            isinstance(flash, (list, tuple)) and len(flash) >= 2 and
            flash[0] == getattr(login_manager, 'login_message_category', None) and
            flash[1] == getattr(login_manager, 'login_message', None)
        )
    ]

    if len(filtered) != len(flashes):
        if filtered:
            session['_flashes'] = filtered
        else:
            session.pop('_flashes', None)
        session.modified = True

# -------------------- Company selection enforcement --------------------

def company_required(view_func):
    """Decorator to ensure a company is selected before accessing product/cart pages.
    If a `company_id` query parameter is present, it will set the selected company
    in the session on-the-fly so that the request can proceed seamlessly.
    """
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        app.logger.info("[DEBUG] company_required decorator called for %s", request.path)

        # Allow admin routes and profile page without company enforcement
        if request.path.startswith('/admin/') or request.path == '/profile' or request.path in ['/', '/index', '/company-selection', '/product_selection', '/signup', '/login', '/logout']:
            app.logger.info("[DEBUG] Bypassing company_required for admin route or public page: %s", request.path)
            return view_func(*args, **kwargs)

        # Only allow access if an explicit company has been selected in this session
        # (must be an object with a non-empty `id`).
        selected_company = session.get('selected_company', {})
        if not isinstance(selected_company, dict):
            selected_company = {}
        app.logger.info("[DEBUG] Current selected_company from session: %s", selected_company)
        active_company_id = (selected_company.get('id') or '').strip()

        # No implicit fallbacks to session legacy fields or current_user are allowed here.
        # The user must explicitly select a company.

        if active_company_id:
            app.logger.info("[DEBUG] Company already selected (id=%s), allowing access", active_company_id)
            return view_func(*args, **kwargs)

        # Otherwise, block access.
        # For AJAX/JSON requests, return a JSON error instead of redirecting to index.
        wants_json = (
            request.is_json
            or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            or 'application/json' in (request.headers.get('Accept') or '')
        )
        if wants_json:
            app.logger.warning("[DEBUG] No company selected for JSON request (%s)", request.path)
            return jsonify({'success': False, 'error': 'No company selected'}), 403

        app.logger.warning("[DEBUG] No company selected, redirecting to index")
        session['needs_company_warning'] = True
        session.modified = True
        return redirect(url_for('index'))
    return wrapped_view


def render_company_product_page(template_name, extra_context=None):
    """Render a product template while ensuring company context is available."""
    company_id = request.args.get('company_id')
    company_name = ''
    company_email = ''

    if company_id:
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')

    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id

    context = extra_context.copy() if isinstance(extra_context, dict) else {}
    context['current_company'] = {
        'id': company_id,
        'name': company_name,
        'email': company_email
    }
    return render_template(template_name, **context)


EXTENDED_DISCOUNT_ADMIN_ROLES = {'admin', 'superadmin', 'sales_admin'}
COMPANY_FULL_ACCESS_ROLES = {'admin', 'superadmin'}
GM_PRICING_ROLES = {'sales_admin', 'superadmin'}
GM_DISCOUNT_STEPS = tuple(range(0, 51, 5))
DEFAULT_GM_DISCOUNT = 0

SUPERADMIN_ROLE_NAME = 'superadmin'
DEFAULT_ROLE_DEFINITIONS = [
    {'name': SUPERADMIN_ROLE_NAME, 'label': 'Super Admin', 'is_custom': False},
    {'name': 'admin', 'label': 'Admin', 'is_custom': False},
    {'name': 'sales_admin', 'label': 'Sales Admin', 'is_custom': False},
    {'name': 'user', 'label': 'User', 'is_custom': False},
]
DEFAULT_ROLE_NAMES = {role['name'] for role in DEFAULT_ROLE_DEFINITIONS}
CUSTOM_ROLES_FILE = os.path.join(app.root_path, 'data', 'custom_roles.json')
RESTRICTED_BLANKET_PATTERNS = [
    re.compile(r'conti\s*sava', re.IGNORECASE),
    re.compile(r'web\s*x\s*press\s*g3', re.IGNORECASE),
]


def _extended_discount_user_set():
    raw = os.getenv('SAVA_EXTENDED_DISCOUNT_USERS', '')
    return {
        entry.strip().lower()
        for entry in raw.split(',')
        if entry and entry.strip()
    }


def is_restricted_blanket(name: str) -> bool:
    if not name:
        return False
    for pattern in RESTRICTED_BLANKET_PATTERNS:
        if pattern.search(str(name)):
            return True
    return False


def has_extended_discount_access(user) -> bool:
    if not user or not getattr(user, 'is_authenticated', False):
        return False

    role = (getattr(user, 'role', '') or '').strip().lower()
    if role in EXTENDED_DISCOUNT_ADMIN_ROLES:
        return True

    extended_users = _extended_discount_user_set()
    email = (getattr(user, 'email', '') or '').strip().lower()
    if email and email in extended_users:
        return True

    user_id = str(getattr(user, 'id', '') or '').strip().lower()
    if user_id and user_id in extended_users:
        return True

    return False


def get_restricted_discount_cap(user) -> float:
    return 10.0 if has_extended_discount_access(user) else 5.0


def is_superadmin(user) -> bool:
    if not user:
        return False
    return (getattr(user, 'role', '') or '').strip().lower() == SUPERADMIN_ROLE_NAME


def has_gm_pricing_access(user) -> bool:
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    raw_role = getattr(user, 'role', '') or ''
    role_key = normalize_role_key(raw_role)
    return role_key in GM_PRICING_ROLES


def sanitize_gm_discount(value) -> int:
    try:
        percent = int(float(value))
    except (TypeError, ValueError):
        return DEFAULT_GM_DISCOUNT
    percent = max(0, min(50, percent))
    # Snap to nearest lower multiple of 5 to respect allowed steps
    if percent % 5 != 0:
        percent -= percent % 5
    return percent


def set_pricing_mode(mode: str):
    if mode not in ('gm', 'standard'):
        mode = 'standard'
    session['pricing_mode'] = mode
    session.modified = True


def get_active_pricing_mode() -> str:
    return session.get('pricing_mode', 'standard')


def get_gm_discount_percent(force=False) -> int:
    if not force and get_active_pricing_mode() != 'gm':
        return DEFAULT_GM_DISCOUNT
    stored = session.get('gm_discount', DEFAULT_GM_DISCOUNT)
    percent = sanitize_gm_discount(stored)
    if stored != percent:
        session['gm_discount'] = percent
        session.modified = True
    return percent


def apply_gm_discount_to_total(total: float):
    """Apply GM flat discount (post-GST) if GM pricing mode is active."""
    if get_active_pricing_mode() != 'gm':
        return total, 0.0, 0

    percent = get_gm_discount_percent()
    if percent <= 0:
        return total, 0.0, 0

    total_value = float(total or 0)
    discount_amount = max(0.0, total_value * (percent / 100))
    adjusted_total = max(0.0, total_value - discount_amount)
    return adjusted_total, discount_amount, percent


@app.context_processor
def inject_pricing_context():
    pricing_mode = get_active_pricing_mode()
    gm_discount = get_gm_discount_percent() if pricing_mode == 'gm' else DEFAULT_GM_DISCOUNT
    return {
        'active_pricing_mode': pricing_mode,
        'session_gm_discount': gm_discount,
        'gm_pricing_roles': sorted(GM_PRICING_ROLES),
        'gm_discount_steps': list(GM_DISCOUNT_STEPS),
        'has_gm_pricing_access': has_gm_pricing_access
    }


def normalize_role_key(raw: str) -> str:
    if not raw:
        return ''
    cleaned = re.sub(r'[^a-zA-Z0-9_\-\s]+', '', str(raw)).strip().lower()
    cleaned = re.sub(r'[\s\-]+', '_', cleaned)
    return cleaned.strip('_')


def prettify_role_label(raw: str) -> str:
    if not raw:
        return ''
    parts = [part for part in str(raw).replace('-', ' ').replace('_', ' ').split(' ') if part]
    return ' '.join(part.capitalize() for part in parts) if parts else str(raw)


def _load_custom_role_records():
    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
        try:
            records = []
            cursor = mongo_db.roles.find({}, {'_id': 0, 'name': 1, 'label': 1, 'created_by': 1, 'created_at': 1})
            for doc in cursor:
                name = normalize_role_key(doc.get('name'))
                if not name:
                    continue
                label = (doc.get('label') or prettify_role_label(name)).strip()
                created_at = doc.get('created_at')
                if hasattr(created_at, 'isoformat'):
                    created_at = created_at.isoformat()
                records.append({
                    'name': name,
                    'label': label,
                    'created_by': doc.get('created_by'),
                    'created_at': created_at,
                })
            return records
        except Exception as err:
            app.logger.error(f"Error loading custom roles from MongoDB: {err}")
    try:
        if os.path.exists(CUSTOM_ROLES_FILE):
            with open(CUSTOM_ROLES_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f) or {}
            roles_data = data.get('roles', [])
        else:
            roles_data = []
    except Exception as err:
        app.logger.error(f"Error reading custom roles file: {err}")
        roles_data = []

    records = []
    for entry in roles_data:
        if isinstance(entry, dict):
            name = normalize_role_key(entry.get('name') or entry.get('label'))
            if not name:
                continue
            label = (entry.get('label') or prettify_role_label(name)).strip()
            records.append({
                'name': name,
                'label': label,
                'created_by': entry.get('created_by'),
                'created_at': entry.get('created_at'),
            })
        elif isinstance(entry, str):
            name = normalize_role_key(entry)
            if name:
                records.append({
                    'name': name,
                    'label': prettify_role_label(entry),
                    'created_by': None,
                    'created_at': None,
                })
    return records


def _save_custom_role_records(records):
    os.makedirs(os.path.dirname(CUSTOM_ROLES_FILE), exist_ok=True)
    try:
        with open(CUSTOM_ROLES_FILE, 'w', encoding='utf-8') as f:
            json.dump({'roles': records}, f, ensure_ascii=False, indent=2)
    except Exception as err:
        app.logger.error(f"Error saving custom roles file: {err}")
        raise


def get_custom_role_definitions():
    records = _load_custom_role_records()
    definitions = []
    for record in records:
        name = record.get('name')
        if not name:
            continue
        definitions.append({
            'name': name,
            'label': (record.get('label') or prettify_role_label(name)).strip(),
            'is_custom': True
        })
    return definitions


def get_role_definitions():
    roles = [dict(role) for role in DEFAULT_ROLE_DEFINITIONS]
    roles.extend(get_custom_role_definitions())
    return sorted(roles, key=lambda r: (r.get('is_custom', False), r.get('label', r.get('name', '')).lower()))


def get_all_role_names():
    return {role['name'] for role in get_role_definitions()}


def add_custom_role_definition(role_name: str, label: str = None, created_by: str = None):
    normalized = normalize_role_key(role_name)
    if not normalized:
        raise ValueError('Role name is required.')
    if normalized in DEFAULT_ROLE_NAMES:
        raise ValueError('Role already exists.')

    existing_custom = {role['name'] for role in get_custom_role_definitions()}
    if normalized in existing_custom:
        raise ValueError('Role already exists.')

    final_label = (label or prettify_role_label(normalized)).strip()
    record = {
        'name': normalized,
        'label': final_label,
        'created_by': created_by,
        'created_at': datetime.utcnow()
    }

    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
        try:
            mongo_db.roles.create_index('name', unique=True)
        except Exception:
            pass
        existing = mongo_db.roles.find_one({'name': normalized})
        if existing:
            raise ValueError('Role already exists.')
        mongo_db.roles.insert_one(record)
    else:
        records = _load_custom_role_records()
        records.append({
            'name': record['name'],
            'label': record['label'],
            'created_by': created_by,
            'created_at': record['created_at'].isoformat(),
        })
        _save_custom_role_records(records)

    return {
        'name': normalized,
        'label': final_label,
        'is_custom': True
    }


def can_assign_role(user, target_role: str) -> bool:
    if not user or not getattr(user, 'is_authenticated', False):
        return False

    normalized_role = normalize_role_key(target_role)
    user_role = (getattr(user, 'role', '') or '').strip().lower()

    if normalized_role not in get_all_role_names():
        return False

    if normalized_role == SUPERADMIN_ROLE_NAME:
        return user_role == SUPERADMIN_ROLE_NAME

    custom_roles = {role['name'] for role in get_custom_role_definitions()}
    if normalized_role in custom_roles:
        return user_role == SUPERADMIN_ROLE_NAME

    return user_role in {'admin', SUPERADMIN_ROLE_NAME}


def is_custom_role(role_name: str) -> bool:
    normalized = normalize_role_key(role_name)
    return normalized in {role['name'] for role in get_custom_role_definitions()}


def get_user_assigned_company_ids(user):
    """Return list of company IDs the user may access.

    Returns None if the user has full access (admin-like roles)."""
    if not user or not getattr(user, 'is_authenticated', False):
        return []

    role = (getattr(user, 'role', '') or '').strip().lower()
    if role in COMPANY_FULL_ACCESS_ROLES:
        return None

    assigned_ids = None

    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
        try:
            mongo_db.command('ping')
            user_doc = mu_find_user_by_id(str(user.id))
            if user_doc and user_doc.get('assigned_companies'):
                assigned_ids = normalize_assigned_companies(user_doc.get('assigned_companies'))
        except Exception as refresh_error:
            app.logger.error(f"Failed to refresh assigned companies from MongoDB: {refresh_error}")

    if assigned_ids is None:
        assigned_ids = normalize_assigned_companies(getattr(user, 'assigned_companies', []))

    user.assigned_companies = assigned_ids
    return [str(cid) for cid in assigned_ids if cid]

# -----------------------------------------------------------------------

CORS(app, resources={
    r"/api/*": {
        "origins": ["*"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "supports_credentials": True
    }
})

# -------------------- MongoDB configuration --------------------
# Admin email for alerts
ADMIN_ALERT_EMAIL = os.getenv('ADMIN_ALERT_EMAIL', 'athulnair3096@gmail.com')

# Helper to send alert email

def send_alert_email(subject: str, body: str):
    """Send an alert email to admin using the Resend API."""
    if not ADMIN_ALERT_EMAIL:
        app.logger.error('No admin email address configured')
        return False

    success = send_email_resend(
        to=ADMIN_ALERT_EMAIL,
        subject=subject,
        text=body
    )

    if success:
        app.logger.info("Alert email sent successfully via Resend")
    return success

# MongoDB Configuration
from pymongo import MongoClient, ReturnDocument
from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError
import time

MONGO_URI = os.environ.get('MONGO_URI', 'mongodb://localhost:27017/')
DB_NAME = os.environ.get('DB_NAME', 'moneda_db')
MONGO_AVAILABLE = False
USE_MONGO = os.environ.get('USE_MONGO', 'true').lower() == 'true'

# Initialize MongoDB client
mongo_client = None
mongo_db = None
users_col = None
mongo_init_lock = threading.Lock()
mongo_initialized = False

def init_mongodb():
    global mongo_client, mongo_db, users_col, MONGO_AVAILABLE
    
    if not USE_MONGO:
        print("MongoDB is disabled via environment variable")
        return
        
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            print(f"Attempting to connect to MongoDB (Attempt {attempt + 1}/{max_retries})...")
            mongo_client = MongoClient(
                MONGO_URI,
                serverSelectionTimeoutMS=3000,  # 3 second timeout
                connectTimeoutMS=3000,
                socketTimeoutMS=10000
            )
            
            # Test the connection
            mongo_client.server_info()
            mongo_db = mongo_client[DB_NAME]
            users_col = mongo_db['users']
            
            # Create indexes if they don't exist
            users_col.create_index('email', unique=True)
            users_col.create_index('username', unique=True)
            # Ensure efficient lookups during company import
            companies_col = mongo_db.get_collection('companies')
            try:
                companies_col.create_index('EmailID', name='email_ci', unique=False, collation={'locale': 'en', 'strength': 2})
                companies_col.create_index('Company Name', name='company_name_ci', unique=False, collation={'locale': 'en', 'strength': 2})
            except Exception as idx_err:
                app.logger.warning(f"Index creation on companies failed: {idx_err}")
            
            MONGO_AVAILABLE = True
            print("Successfully connected to MongoDB")
            return
            
        except (ConnectionFailure, ServerSelectionTimeoutError) as e:
            print(f"MongoDB connection attempt {attempt + 1} failed: {str(e)}")
            if attempt < max_retries - 1:
                print(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
            else:
                print("Could not connect to MongoDB after multiple attempts. Falling back to JSON storage.")
                MONGO_AVAILABLE = False

def ensure_mongo_connection_initialized():
    global mongo_initialized

    if mongo_initialized or not USE_MONGO:
        return

    with mongo_init_lock:
        if mongo_initialized or not USE_MONGO:
            return

        init_mongodb()
        mongo_initialized = True


@app.before_request
def _ensure_mongo_connection():
    ensure_mongo_connection_initialized()

def mu_find_user_by_email_or_username(identifier):
    """Find a user by email or username in MongoDB"""
    if not MONGO_AVAILABLE or users_col is None:
        return None
    try:
        return users_col.find_one({
            '$or': [
                {'email': identifier},
                {'username': identifier}
            ]
        })
    except Exception as e:
        print(f"Error finding user by email/username: {str(e)}")
        return None

def mu_create_user(email, username, password, phone=None):
    """Create a new user in MongoDB"""
    if not MONGO_AVAILABLE or users_col is None:
        return None
    
    try:
        from werkzeug.security import generate_password_hash
        user_data = {
            'email': email,
            'username': username,
            'username_lower': username.lower(),
            'password_hash': generate_password_hash(password),
            'is_verified': False,
            'otp_verified': False,
            'created_at': time.time(),
            'updated_at': time.time(),
            'role': 'user'
        }
        if phone:
            user_data['phone'] = phone
        result = users_col.insert_one(user_data)
        return str(result.inserted_id)
    except Exception as e:
        print(f"Error creating user: {str(e)}")
        return None

def mu_find_user_by_id(user_id):
    """Find a user by ID in MongoDB"""
    if not MONGO_AVAILABLE or users_col is None:
        return None
    
    try:
        from bson.objectid import ObjectId
        return users_col.find_one({'_id': ObjectId(user_id)})
    except Exception as e:
        print(f"Error finding user by ID: {str(e)}")
        return None

def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated:
            return login_manager.unauthorized()

        user_role = (getattr(current_user, 'role', 'user') or '').strip().lower()
        if user_role not in {'admin', 'superadmin'}:
            return abort(403)

        return view_func(*args, **kwargs)

    return wrapped

# Health check endpoint
@app.route('/health')
def health_check():
    """Health check endpoint"""
    mongo_status = "connected" if MONGO_AVAILABLE else "disconnected"
    status = {
        'status': 'ok',
        'timestamp': time.time(),
        'mongo_available': MONGO_AVAILABLE,
        'use_mongo': USE_MONGO,
        'mongo_status': mongo_status,
        'app': 'pathways',
        'version': '1.0.0'
    }
    
    # Add MongoDB connection status details if available
    if MONGO_AVAILABLE and mongo_client:
        try:
            # Test the connection with a quick operation
            mongo_client.admin.command('ping')
            status['mongo_connection'] = 'healthy'
            
            # Add some basic database stats
            if mongo_db:
                status['database'] = mongo_db.name
                status['collections'] = mongo_db.list_collection_names()
                
                # Add user count if users collection exists
                if 'users' in status['collections']:
                    status['user_count'] = mongo_db.users.count_documents({})
                    
        except Exception as e:
            status['mongo_connection'] = 'error'
            status['mongo_error'] = str(e)
            
    return jsonify(status)

# This was an incorrectly indented else block - removed as it's not part of any if statement
if not USE_MONGO:
    print("MongoDB not enabled in configuration")

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html', user=current_user)

@app.route('/admin/manage-users')
@login_required
@admin_required
def admin_manage_users():
    return render_template('admin/manage_users.html', user=current_user)


@app.route('/admin/quotations')
@login_required
@admin_required
def admin_quotations():
    return render_template('admin/quotations.html', user=current_user)


@app.route('/api/admin/users', methods=['GET'])
@login_required
@admin_required
def admin_list_users():
    try:
        users = load_admin_users()
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        app.logger.error(f"Error in admin_list_users: {e}")
        return jsonify({'success': False, 'error': 'Failed to load users'}), 500


@app.route('/api/admin/users/me', methods=['GET'])
@login_required
def admin_get_current_user():
    try:
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401

        user_data = serialize_admin_user({
            '_id': getattr(current_user, 'id', None),
            'username': getattr(current_user, 'username', ''),
            'email': getattr(current_user, 'email', ''),
            'role': getattr(current_user, 'role', 'user'),
            'assigned_companies': getattr(current_user, 'assigned_companies', []) or [],
            'created_at': getattr(current_user, 'created_at', datetime.utcnow()),
            'updated_at': getattr(current_user, 'updated_at', datetime.utcnow()),
        })
        return jsonify({'success': True, 'user': user_data})
    except Exception as err:
        app.logger.error(f"Error getting current admin user: {err}")
        return jsonify({'success': False, 'error': 'Failed to load current user'}), 500


@app.route('/api/admin/users/<user_id>', methods=['GET'])
@login_required
@admin_required
def admin_get_user(user_id):
    try:
        users = load_admin_users()
        user = next((u for u in users if str(u.get('id')) == str(user_id)), None)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        return jsonify({'success': True, 'user': user})
    except Exception as e:
        app.logger.error(f"Error in admin_get_user: {e}")
        return jsonify({'success': False, 'error': 'Failed to load user'}), 500


@app.route('/api/admin/users', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        email = (data.get('email') or '').strip().lower()
        password = data.get('password')
        phone = (data.get('phone') or '').strip()
        position = (data.get('position') or '').strip()
        role = (data.get('role') or 'user').lower()
        assigned_companies = normalize_assigned_companies(data.get('assigned_companies', []))

        if not username or not email or not password:
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                existing = mongo_db.users.find_one({'email': email})
                if existing:
                    return jsonify({'success': False, 'error': 'Email already registered'}), 409

                hashed_password = generate_password_hash(password)
                new_user = {
                    'username': username,
                    'username_lower': username.lower(),
                    'email': email,
                    'password_hash': hashed_password,
                    'phone': phone,
                    'position': position,
                    'role': role,
                    'is_verified': True,
                    'assigned_companies': assigned_companies,
                    'created_at': datetime.utcnow(),
                    'updated_at': datetime.utcnow()
                }

                result = mongo_db.users.insert_one(new_user)
                new_user['id'] = str(result.inserted_id)
                sync_user_company_links(new_user['id'], assigned_companies)
                return jsonify({'success': True, 'user': serialize_admin_user(new_user)})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_create_user: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        users = _load_users_json()
        if any((user.email if isinstance(user, User) else user.get('email', '')).lower() == email for user in users.values()):
            return jsonify({'success': False, 'error': 'Email already registered'}), 409

        user_id = str(uuid.uuid4())
        hashed_password = generate_password_hash(password)
        new_user = {
            'id': user_id,
            'username': username,
            'email': email,
            'password_hash': hashed_password,
            'phone': phone,
            'role': role,
            'is_verified': True,
            'assigned_companies': assigned_companies,
            'created_at': datetime.utcnow().isoformat(),
            'updated_at': datetime.utcnow().isoformat()
        }

        users[user_id] = new_user
        save_users(users)
        sync_user_company_links(user_id, assigned_companies)
        return jsonify({'success': True, 'user': serialize_admin_user(new_user)})
    except Exception as e:
        app.logger.error(f"Error in admin_create_user: {e}")
        return jsonify({'success': False, 'error': 'Failed to create user'}), 500


@app.route('/api/admin/users/<user_id>', methods=['PUT'])
@login_required
@admin_required
def admin_update_user(user_id):
    try:
        data = request.get_json() or {}

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                update_fields = {'updated_at': datetime.utcnow()}

                if 'username' in data:
                    update_fields['username'] = data['username']
                    update_fields['username_lower'] = data['username'].lower()
                if 'email' in data:
                    update_fields['email'] = data['email'].lower()
                if 'phone' in data:
                    update_fields['phone'] = data['phone'].strip()
                if 'position' in data:
                    update_fields['position'] = data['position'].strip()
                if 'role' in data:
                    role_value = normalize_role_key(data['role'])
                    if not can_assign_role(current_user, role_value):
                        return jsonify({'success': False, 'error': 'You are not allowed to assign that role'}), 403
                    update_fields['role'] = role_value
                if 'password' in data and data['password']:
                    update_fields['password_hash'] = generate_password_hash(data['password'])
                if 'assigned_companies' in data:
                    update_fields['assigned_companies'] = normalize_assigned_companies(data.get('assigned_companies', []))

                result = mongo_db.users.update_one({'_id': ObjectId(user_id)}, {'$set': update_fields})
                if result.matched_count == 0:
                    return jsonify({'success': False, 'error': 'User not found'}), 404

                user_doc = mongo_db.users.find_one({'_id': ObjectId(user_id)})
                user_doc['id'] = str(user_doc.pop('_id'))
                sync_user_company_links(user_doc['id'], user_doc.get('assigned_companies', []))
                return jsonify({'success': True, 'user': serialize_admin_user(user_doc)})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_update_user: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        users = _load_users_json()
        user = users.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        if 'username' in data:
            user['username'] = data['username']
        if 'email' in data:
            user['email'] = data['email'].lower()
        if 'phone' in data:
            user['phone'] = data['phone'].strip()
        if 'position' in data:
            user['position'] = data['position'].strip()
        if 'role' in data:
            role_value = normalize_role_key(data['role'])
            if not can_assign_role(current_user, role_value):
                return jsonify({'success': False, 'error': 'You are not allowed to assign that role'}), 403
            user['role'] = role_value
        if 'password' in data and data['password']:
            user['password_hash'] = generate_password_hash(data['password'])
        if 'assigned_companies' in data:
            user['assigned_companies'] = normalize_assigned_companies(data.get('assigned_companies', []))

        user['updated_at'] = datetime.utcnow().isoformat()
        users[user_id] = user
        save_users(users)
        sync_user_company_links(user_id, user.get('assigned_companies', []))
        return jsonify({'success': True, 'user': serialize_admin_user(user)})
    except Exception as e:
        app.logger.error(f"Error in admin_update_user: {e}")
        return jsonify({'success': False, 'error': 'Failed to update user'}), 500


@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
@login_required
@admin_required
def admin_delete_user(user_id):
    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                result = mongo_db.users.delete_one({'_id': ObjectId(user_id)})
                if result.deleted_count == 0:
                    return jsonify({'success': False, 'error': 'User not found'}), 404
                sync_user_company_links(user_id, [])
                return jsonify({'success': True})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_delete_user: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        users = _load_users_json()
        if user_id not in users:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        users.pop(user_id)
        save_users(users)
        sync_user_company_links(user_id, [])
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Error in admin_delete_user: {e}")
        return jsonify({'success': False, 'error': 'Failed to delete user'}), 500


@app.route('/api/admin/companies', methods=['GET'])
@login_required
@admin_required
def admin_list_companies():
    try:
        global USE_MONGO
        page = max(int(request.args.get('page', 0)), 0)
        limit = max(min(int(request.args.get('limit', 10)), 1000), 1)
        start = page * limit

        companies_payload = []
        total_count = 0

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')

                projection = {
                    '_id': 1,
                    'Company Name': 1,
                    'EmailID': 1,
                    'Phone': 1,
                    'Billing Attention': 1,
                    'Billing Address': 1,
                    'Billing Street': 1,
                    'Billing City': 1,
                    'Billing State': 1,
                    'Billing Postal Code': 1,
                    'Billing Phone': 1,
                    'created_at': 1,
                    'Created At': 1,
                    'name': 1,
                    'email': 1,
                    'Address': 1,
                    'last_payment_terms': 1,
                    'assigned_to': 1
                }

                total_count = mongo_db.companies.count_documents({})
                cursor = (
                    mongo_db.companies
                    .find({}, projection)
                    .sort('Company Name', 1)
                    .skip(start)
                    .limit(limit)
                )

                for doc in cursor:
                    company_doc = dict(doc)
                    company_doc['id'] = str(company_doc.pop('_id'))
                    companies_payload.append(serialize_admin_company(company_doc))
            except Exception as mongo_error:
                app.logger.error(f"Mongo pagination error in admin_list_companies: {mongo_error}")
                USE_MONGO = False

        if not companies_payload:
            companies = load_companies_data()
            total_count = len(companies)
            slice_end = start + limit
            companies_slice = companies[start:slice_end]
            companies_payload = [serialize_admin_company(company) for company in companies_slice]

        return jsonify({
            'success': True,
            'companies': companies_payload,
            'page': page,
            'limit': limit,
            'total': total_count
        })
    except Exception as e:
        app.logger.error(f"Error in admin_list_companies: {e}")
        return jsonify({'success': False, 'error': 'Failed to load companies'}), 500


@app.route('/api/admin/companies/<company_id>', methods=['GET'])
@login_required
@admin_required
def admin_get_company(company_id):
    try:
        companies = load_companies_data()
        for company in companies:
            serialized = serialize_admin_company(company)
            if serialized.get('id') == str(company_id):
                return jsonify({'success': True, 'company': serialized})
        return jsonify({'success': False, 'error': 'Company not found'}), 404
    except Exception as e:
        app.logger.error(f"Error in admin_get_company: {e}")
        return jsonify({'success': False, 'error': 'Failed to load company'}), 500


@app.route('/api/admin/companies', methods=['POST'])
@login_required
@admin_required
def admin_create_company():
    try:
        data = request.get_json() or {}
        def _required(field):
            value = (data.get(field) or '').strip()
            if not value:
                raise ValueError(field)
            return value

        try:
            name = _required('name')
            email = _required('email')
            phone = _required('phone')
            billing_attention = _required('billing_attention')
            billing_address = _required('billing_address')
            billing_city = _required('billing_city')
            billing_state = _required('billing_state')
            billing_postal_code = _required('billing_postal_code')
        except ValueError as missing_field:
            return jsonify({'success': False, 'error': f"Missing required field: {missing_field}"}), 400

        # Admin dashboard uses a single phone field; keep Billing Phone for backward compatibility.
        billing_phone = (data.get('billing_phone') or phone or '').strip()

        billing_street = (data.get('billing_street') or '').strip()
        last_payment_terms = (data.get('last_payment_terms') or '').strip()
        assigned_to = normalize_assigned_companies(data.get('assigned_to', []))

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                company_data = {
                    'Company Name': name,
                    'EmailID': email,
                    'Phone': phone,
                    'last_payment_terms': last_payment_terms,
                    'Billing Attention': billing_attention,
                    'Billing Address': billing_address,
                    'Billing Street': billing_street,
                    'Billing City': billing_city,
                    'Billing State': billing_state,
                    'Billing Postal Code': billing_postal_code,
                    'Billing Phone': billing_phone,
                    'Address': billing_address,
                    'assigned_to': assigned_to,
                    'created_at': datetime.utcnow(),
                    'updated_at': datetime.utcnow(),
                    'created_by': str(current_user.id)
                }
                result = mongo_db.companies.insert_one(company_data)
                company_data['id'] = str(result.inserted_id)
                sync_company_user_links(company_data['id'], assigned_to)
                return jsonify({'success': True, 'company': serialize_admin_company(company_data)})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_create_company: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        companies = load_companies_data()
        new_company = {
            'id': str(uuid.uuid4()),
            'Company Name': name,
            'EmailID': email,
            'Phone': phone,
            'last_payment_terms': last_payment_terms,
            'Billing Attention': billing_attention,
            'Billing Address': billing_address,
            'Billing Street': billing_street,
            'Billing City': billing_city,
            'Billing State': billing_state,
            'Billing Postal Code': billing_postal_code,
            'Billing Phone': billing_phone,
            'Address': billing_address,
            'assigned_to': assigned_to,
            'created_at': datetime.utcnow().isoformat()
        }

        companies.append(new_company)
        save_companies_data(companies)
        sync_company_user_links(new_company['id'], assigned_to)
        return jsonify({'success': True, 'company': serialize_admin_company(new_company)})
    except Exception as e:
        app.logger.error(f"Error in admin_create_company: {e}")
        return jsonify({'success': False, 'error': 'Failed to create company'}), 500


@app.route('/api/admin/companies/<company_id>', methods=['PUT'])
@login_required
@admin_required
def admin_update_company(company_id):
    try:
        data = request.get_json() or {}
        assigned_to_payload = data.get('assigned_to') if 'assigned_to' in data else None

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                update_fields = {'updated_at': datetime.utcnow()}
                if 'name' in data:
                    update_fields['Company Name'] = data['name']
                if 'email' in data:
                    update_fields['EmailID'] = data['email']
                if 'phone' in data:
                    update_fields['Phone'] = data['phone']
                if 'last_payment_terms' in data:
                    update_fields['last_payment_terms'] = data['last_payment_terms']
                if 'billing_attention' in data:
                    update_fields['Billing Attention'] = data['billing_attention']
                if 'billing_address' in data:
                    update_fields['Billing Address'] = data['billing_address']
                    update_fields['Address'] = data['billing_address']
                if 'billing_street' in data:
                    update_fields['Billing Street'] = data['billing_street']
                if 'billing_city' in data:
                    update_fields['Billing City'] = data['billing_city']
                if 'billing_state' in data:
                    update_fields['Billing State'] = data['billing_state']
                if 'billing_postal_code' in data:
                    update_fields['Billing Postal Code'] = data['billing_postal_code']
                if 'billing_phone' in data:
                    update_fields['Billing Phone'] = data['billing_phone']
                if 'assigned_to' in data:
                    update_fields['assigned_to'] = normalize_assigned_companies(data.get('assigned_to', []))

                result = mongo_db.companies.update_one({'_id': ObjectId(company_id)}, {'$set': update_fields})
                if result.matched_count == 0:
                    return jsonify({'success': False, 'error': 'Company not found'}), 404

                company_doc = mongo_db.companies.find_one({'_id': ObjectId(company_id)})
                company_doc['id'] = str(company_doc.pop('_id'))
                serialized_company = serialize_admin_company(company_doc)
                sync_company_user_links(serialized_company.get('id'), serialized_company.get('assigned_to', []))
                return jsonify({'success': True, 'company': serialized_company})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_update_company: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        companies = load_companies_data()
        updated = None
        for company in companies:
            serialized = serialize_admin_company(company)
            if serialized.get('id') == str(company_id):
                if 'name' in data:
                    company['Company Name'] = data['name']
                if 'email' in data:
                    company['EmailID'] = data['email']
                if 'phone' in data:
                    company['Phone'] = data['phone']
                if 'last_payment_terms' in data:
                    company['last_payment_terms'] = data['last_payment_terms']
                if 'billing_attention' in data:
                    company['Billing Attention'] = data['billing_attention']
                if 'billing_address' in data:
                    company['Billing Address'] = data['billing_address']
                    company['Address'] = data['billing_address']
                if 'billing_street' in data:
                    company['Billing Street'] = data['billing_street']
                if 'billing_city' in data:
                    company['Billing City'] = data['billing_city']
                if 'billing_state' in data:
                    company['Billing State'] = data['billing_state']
                if 'billing_postal_code' in data:
                    company['Billing Postal Code'] = data['billing_postal_code']
                if 'billing_phone' in data:
                    company['Billing Phone'] = data['billing_phone']
                if 'assigned_to' in data:
                    company['assigned_to'] = normalize_assigned_companies(data.get('assigned_to', []))
                updated = serialize_admin_company(company)
                break

        if not updated:
            return jsonify({'success': False, 'error': 'Company not found'}), 404

        save_companies_data(companies)
        sync_company_user_links(company_id, updated.get('assigned_to', []))
        return jsonify({'success': True, 'company': updated})
    except Exception as e:
        app.logger.error(f"Error in admin_update_company: {e}")
        return jsonify({'success': False, 'error': 'Failed to update company'}), 500


@app.route('/api/admin/companies/<company_id>', methods=['DELETE'])
@login_required
@admin_required
def admin_delete_company(company_id):
    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                result = mongo_db.companies.delete_one({'_id': ObjectId(company_id)})
                if result.deleted_count == 0:
                    return jsonify({'success': False, 'error': 'Company not found'}), 404
                sync_company_user_links(company_id, [])
                return jsonify({'success': True})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_delete_company: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        companies = load_companies_data()
        new_companies = []
        deleted = False
        for company in companies:
            serialized = serialize_admin_company(company)
            if serialized.get('id') == str(company_id):
                deleted = True
                continue
            new_companies.append(company)

        if not deleted:
            return jsonify({'success': False, 'error': 'Company not found'}), 404

        save_companies_data(new_companies)
        sync_company_user_links(company_id, [])
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Error in admin_delete_company: {e}")
        return jsonify({'success': False, 'error': 'Failed to delete company'}), 500


@app.route('/api/admin/companies/search', methods=['GET'])
@login_required
@admin_required
def admin_search_companies():
    try:
        query = (request.args.get('q') or '').strip().lower()
        if len(query) < 2:
            return jsonify({'success': True, 'companies': []})

        companies = load_companies_data()
        filtered = []
        for company in companies:
            serialized = serialize_admin_company(company)
            if query in serialized.get('name', '').lower():
                filtered.append(serialized)
        return jsonify({'success': True, 'companies': filtered})
    except Exception as e:
        app.logger.error(f"Error in admin_search_companies: {e}")
        return jsonify({'success': False, 'error': 'Failed to search companies'}), 500


@app.route('/api/admin/companies/import', methods=['POST'])
@login_required
@admin_required
def admin_import_companies():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        upload = request.files['file']
        if not upload or upload.filename.strip() == '':
            return jsonify({'success': False, 'error': 'Empty filename'}), 400

        filename = secure_filename(upload.filename)
        if not filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({'success': False, 'error': 'Unsupported file type'}), 400

        upload.stream.seek(0)
        workbook = load_workbook(upload.stream, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'error': 'Workbook is empty'}), 400

        headers_normalized, missing_headers = _validate_company_import_headers(rows[0])
        if missing_headers:
            return jsonify({'success': False, 'error': f"Missing required columns: {', '.join(missing_headers)}"}), 400

        insert_count = 0
        update_count = 0
        errors = []
        processed = []

        use_mongo = MONGO_AVAILABLE and USE_MONGO and mongo_db is not None
        if use_mongo:
            try:
                mongo_db.command('ping')
            except Exception as ping_error:
                app.logger.error(f"Mongo ping failed in admin_import_companies: {ping_error}")
                use_mongo = False

        for row_index, row in enumerate(rows[1:], start=2):
            try:
                record = _extract_row_data(headers_normalized, row)
            except Exception as extract_error:
                app.logger.error(f"Row {row_index} parsing error: {extract_error}", exc_info=True)
                errors.append(f"Row {row_index}: Failed to parse row data")
                continue
            if not any(str(value).strip() for value in record.values()):
                continue

            identifier_key, identifier_value = _resolve_company_identifier(record)
            if not identifier_value:
                errors.append(f"Row {row_index}: Missing company name/email")
                continue

            try:
                payload, created_at, assigned_to = _convert_record_to_storage(record)
            except Exception as convert_error:
                app.logger.error(f"Row {row_index} conversion error: {convert_error}", exc_info=True)
                errors.append(f"Row {row_index}: Failed to normalize data")
                continue

            try:
                if use_mongo:
                    company_id, was_inserted, previous_assigned = _upsert_company_mongo(identifier_key, identifier_value, payload, assigned_to, created_at)
                else:
                    company_id, was_inserted, previous_assigned = _upsert_company_json(identifier_key, identifier_value, payload, assigned_to, created_at)

                if not company_id:
                    errors.append(f"Row {row_index}: Failed to upsert company")
                    continue

                _sync_assigned_users(company_id, assigned_to, previous_assigned)

                if was_inserted:
                    insert_count += 1
                    processed.append({'row': row_index, 'company_id': company_id, 'status': 'inserted'})
                else:
                    update_count += 1
                    processed.append({'row': row_index, 'company_id': company_id, 'status': 'updated'})
            except Exception as row_error:
                app.logger.error(f"Error importing row {row_index}: {row_error}")
                errors.append(f"Row {row_index}: {row_error}")

        return jsonify({
            'success': True,
            'inserted': insert_count,
            'updated': update_count,
            'errors': errors,
            'processed': processed
        })
    except Exception as e:
        app.logger.error(f"Error in admin_import_companies: {e}")
        return jsonify({'success': False, 'error': 'Failed to import companies'}), 500


@app.route('/api/admin/companies/export', methods=['GET'])
@login_required
@admin_required
def admin_export_companies():
    try:
        companies = [serialize_admin_company(company) for company in load_companies_data()]
        wb = Workbook()
        ws = wb.active
        ws.title = 'Companies'

        headers = [
            'Sr No',
            'Company Name',
            'Phone',
            'Billing Attention',
            'Billing Address',
            'Billing Street',
            'Billing City',
            'Billing State',
            'Postal Code',
            'Billing Phone',
            'EmailID',
            'Users Assigned To',
            'Created Time'
        ]
        ws.append(headers)

        for idx, company in enumerate(companies, start=1):
            ws.append([
                idx,
                company.get('name', COMPANY_PLACEHOLDER),
                company.get('phone', COMPANY_PLACEHOLDER),
                company.get('billing_attention', COMPANY_PLACEHOLDER),
                company.get('billing_address', COMPANY_PLACEHOLDER),
                company.get('billing_street', COMPANY_PLACEHOLDER),
                company.get('billing_city', COMPANY_PLACEHOLDER),
                company.get('billing_state', COMPANY_PLACEHOLDER),
                company.get('billing_postal_code', COMPANY_PLACEHOLDER),
                company.get('billing_phone', COMPANY_PLACEHOLDER),
                company.get('email', COMPANY_PLACEHOLDER),
                ','.join(company.get('assigned_to', [])) or COMPANY_PLACEHOLDER,
                company.get('created_at', COMPANY_PLACEHOLDER)
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'companies_export_{timestamp}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        app.logger.error(f"Error exporting companies: {e}")
        return jsonify({'success': False, 'error': 'Failed to export companies'}), 500


@app.route('/api/admin/quotations', methods=['GET'])
@login_required
@admin_required
def admin_list_quotations():
    try:
        quotations = []
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                cursor = mongo_db.quotations.find({}).sort('created_at', -1)
                for doc in cursor:
                    mongo_id = str(doc.get('_id'))
                    doc['mongo_id'] = mongo_id
                    doc['id'] = doc.get('quote_id') or mongo_id
                    doc.pop('_id', None)
                    quotations.append(serialize_admin_quotation(doc))
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_list_quotations: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500
        return jsonify({'success': True, 'quotations': quotations})
    except Exception as e:
        app.logger.error(f"Error in admin_list_quotations: {e}")
        return jsonify({'success': False, 'error': 'Failed to load quotations'}), 500


@app.route('/api/admin/quotations/<quotation_id>', methods=['GET'])
@login_required
@admin_required
def admin_get_quotation(quotation_id):
    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                query = {}
                if ObjectId.is_valid(quotation_id):
                    query['_id'] = ObjectId(quotation_id)
                else:
                    query['quote_id'] = quotation_id

                doc = mongo_db.quotations.find_one(query)

                if not doc and '_id' in query:
                    # If lookup by ObjectId failed or quotation_id was formatted, try quote_id as fallback
                    doc = mongo_db.quotations.find_one({'quote_id': quotation_id})

                if not doc:
                    return jsonify({'success': False, 'error': 'Quotation not found'}), 404

                mongo_id = str(doc.get('_id'))
                doc['mongo_id'] = mongo_id
                doc['id'] = doc.get('quote_id') or mongo_id
                doc.pop('_id', None)
                return jsonify({'success': True, 'quotation': serialize_admin_quotation(doc)})
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_get_quotation: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500
        return jsonify({'success': False, 'error': 'Quotation not found'}), 404
    except Exception as e:
        app.logger.error(f"Error in admin_get_quotation: {e}")
        return jsonify({'success': False, 'error': 'Failed to load quotation'}), 500


@app.route('/admin/quotations/<quotation_id>/pdf')
@login_required
@admin_required
def admin_quotation_pdf(quotation_id):
    if HTML is None:
        flash('PDF generation is not available on this server.', 'danger')
        return redirect(url_for('admin_quotations'))

    if not (MONGO_AVAILABLE and USE_MONGO and mongo_db is not None):
        flash('Database not available.', 'danger')
        return redirect(url_for('admin_quotations'))

    try:
        mongo_db.command('ping')
        query = {'quote_id': quotation_id}
        if ObjectId.is_valid(quotation_id):
            query = {'_id': ObjectId(quotation_id)}

        doc = mongo_db.quotations.find_one(query)
        if not doc and '_id' in query:
            doc = mongo_db.quotations.find_one({'quote_id': quotation_id})

        if not doc:
            flash('Quotation not found.', 'warning')
            return redirect(url_for('admin_quotations'))

        created_at = doc.get('created_at')
        if isinstance(created_at, dict) and '$date' in created_at:
            try:
                created_at = datetime.fromisoformat(created_at['$date'].replace('Z', '+00:00'))
            except Exception:
                created_at = None

        current_datetime = created_at if isinstance(created_at, datetime) else get_india_time()
        if isinstance(current_datetime, datetime) and current_datetime.tzinfo is None:
            current_datetime = current_datetime.replace(tzinfo=IST)
        quote_date = current_datetime.strftime('%d-%m-%Y')
        quote_time = current_datetime.strftime('%H:%M:%S')

        products = doc.get('products') or []
        cart = {'products': products}

        selected_company = {
            'id': doc.get('company_id') or '',
            'name': doc.get('company_name') or '',
            'email': doc.get('company_email') or ''
        }

        subtotal_before_discount = doc.get('subtotal_before_discount') or 0
        total_discount = doc.get('total_discount') or 0
        subtotal_after_discount = doc.get('subtotal_after_discount') or doc.get('total_amount_pre_gst') or max(0, float(subtotal_before_discount or 0) - float(total_discount or 0))
        total_gst = doc.get('total_gst') or doc.get('gst_amount') or 0
        total_after_gst = doc.get('total_amount_post_gst') or (float(subtotal_after_discount or 0) + float(total_gst or 0))

        calculations = {
            'subtotal_before_discount': subtotal_before_discount,
            'total_discount': total_discount,
            'subtotal_after_discount': subtotal_after_discount,
            'gst_breakdown': {
                'total_gst': total_gst
            },
            'total': total_after_gst
        }

        payment_terms = doc.get('payment_terms') or doc.get('paymentTerms') or ''

        context = {
            'cart': cart,
            'quote_date': quote_date,
            'quote_time': quote_time,
            'company_name': selected_company.get('name') or 'Not specified',
            'company_email': selected_company.get('email') or '',
            'payment_terms': payment_terms,
            'company_details': build_quotation_company_details(
                selected_company,
                selected_company.get('id'),
                selected_company.get('email'),
                fallback={
                    'name': selected_company.get('name') or 'Not specified',
                    'email': selected_company.get('email') or ''
                }
            ),
            'calculations': calculations,
            'now': current_datetime
        }

        html = render_template('quotation_pdf.html', **context)
        pdf_bytes = HTML(string=html, base_url=request.url_root).write_pdf()

        safe_id = doc.get('quote_id') or str(doc.get('_id'))
        filename = f"quotation_{safe_id}.pdf"
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        app.logger.error(f"Error generating admin quotation PDF: {e}", exc_info=True)
        flash('Failed to generate PDF.', 'danger')
        return redirect(url_for('admin_quotations'))


@app.route('/api/customers', methods=['GET'])
@login_required
@admin_required
def admin_list_customers():
    try:
        customers = []
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
                cursor = mongo_db.companies.find({}, {'Company Name': 1, 'EmailID': 1})
                for doc in cursor:
                    customers.append({
                        'id': str(doc['_id']),
                        'name': doc.get('Company Name', ''),
                        'email': doc.get('EmailID', '')
                    })
            except Exception as mongo_error:
                app.logger.error(f"Mongo error in admin_list_customers: {mongo_error}")
                return jsonify({'success': False, 'error': 'Database error'}), 500

        if not customers:
            customers = load_companies_data()

        return jsonify({'success': True, 'customers': [
            {
                'id': customer.get('id') or customer.get('_id') or str(uuid.uuid4()),
                'name': customer.get('name') or customer.get('Company Name') or 'Unknown',
                'email': customer.get('email') or customer.get('EmailID') or ''
            }
            for customer in customers
        ]})
    except Exception as e:
        app.logger.error(f"Error in admin_list_customers: {e}")
        return jsonify({'success': False, 'error': 'Failed to load customers'}), 500

@app.route('/api/admin/stats')
@login_required
@admin_required
def admin_stats():
    """Return aggregate statistics for the admin dashboard."""
    try:
        stats = {
            'total_accounts': 0,
            'total_users': 0,
            'active_sessions': 0,
            'total_quotations': 0,
            'quotations_on_hold': 0,
            'recent_activity': []
        }

        sessions_snapshot = get_active_sessions_snapshot()
        stats['active_sessions'] = len(sessions_snapshot)

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
            except Exception as ping_error:
                app.logger.error(f"MongoDB ping failed in admin_stats: {ping_error}")
                return jsonify({'error': 'Database unavailable'}), 503

            collection_names = set(mongo_db.list_collection_names())

            if 'companies' in collection_names:
                stats['total_accounts'] = mongo_db.companies.count_documents({})

            stats['total_users'] = mongo_db.users.count_documents({})

            if 'quotations' in collection_names:
                stats['total_quotations'] = mongo_db.quotations.count_documents({})

            recent_cursor = mongo_db.users.find(
                {},
                {
                    'username': 1,
                    'role': 1,
                    'email': 1,
                    'last_login': 1,
                    'updated_at': 1,
                    'created_at': 1
                }
            ).sort('updated_at', -1).limit(10)

            for doc in recent_cursor:
                ts = doc.get('last_login') or doc.get('updated_at') or doc.get('created_at')
                if ts and hasattr(ts, 'isoformat'):
                    timestamp = ts.isoformat()
                else:
                    timestamp = str(ts) if ts else ''

                stats['recent_activity'].append({
                    'message': f"{doc.get('username', doc.get('email', 'Unknown user'))} activity",
                    'role': doc.get('role', 'user'),
                    'timestamp': timestamp
                })

        else:
            users_json = _load_users_json()
            if isinstance(users_json, dict):
                stats['total_users'] = len(users_json)
            statistics_companies = 0
            try:
                file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as f:
                        companies_payload = json.load(f)
                    if isinstance(companies_payload, dict):
                        companies_list = companies_payload.get('companies', [])
                    else:
                        companies_list = companies_payload
                    statistics_companies = len(companies_list) if isinstance(companies_list, list) else 0
            except Exception as company_error:
                app.logger.warning(f"Failed to load fallback company data: {company_error}")
            stats['total_accounts'] = statistics_companies

        return jsonify(stats)
    except Exception as e:
        app.logger.error(f"Error in admin_stats: {e}")
        return jsonify({'error': 'Failed to load stats'}), 500


@app.route('/api/admin/active-sessions')
@login_required
@admin_required
def admin_get_active_sessions():
    """Return current active session snapshot for admins."""
    try:
        sessions = get_active_sessions_snapshot()
        return jsonify({'success': True, 'sessions': sessions})
    except Exception as e:
        app.logger.error(f"Error getting active sessions: {e}")
        return jsonify({'success': False, 'error': 'Failed to load active sessions'}), 500


@app.route('/api/admin/roles', methods=['GET'])
@login_required
def admin_list_roles():
    try:
        if not can_assign_role(current_user, 'admin') and not is_superadmin(current_user):
            return jsonify({'success': False, 'error': 'Forbidden'}), 403

        return jsonify({
            'success': True,
            'roles': get_role_definitions()
        })
    except Exception as err:
        app.logger.error(f"Error listing roles: {err}")
        return jsonify({'success': False, 'error': 'Failed to load roles'}), 500


@app.route('/api/admin/roles', methods=['POST'])
@login_required
def admin_create_role():
    if not is_superadmin(current_user):
        return jsonify({'success': False, 'error': 'Only superadmins can create roles'}), 403

    payload = request.get_json(silent=True) or {}
    raw_name = payload.get('name') or payload.get('role')
    label = payload.get('label')

    if not raw_name:
        return jsonify({'success': False, 'error': 'Role name is required'}), 400

    try:
        definition = add_custom_role_definition(raw_name, label=label, created_by=str(current_user.id))
        return jsonify({'success': True, 'role': definition})
    except ValueError as ve:
        return jsonify({'success': False, 'error': str(ve)}), 400
    except Exception as err:
        app.logger.error(f"Error creating custom role: {err}")
        return jsonify({'success': False, 'error': 'Failed to create role'}), 500


@app.route('/api/admin/chart-data')
@login_required
@admin_required
def admin_chart_data():
    """Return chart data for the admin dashboard."""
    try:
        result = {
            'users_by_month': [],
            'user_roles': []
        }

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
            except Exception as ping_error:
                app.logger.error(f"MongoDB ping failed in admin_chart_data: {ping_error}")
                return jsonify({'error': 'Database unavailable'}), 503

            users_cursor = mongo_db.users.find({}, {'created_at': 1, 'role': 1})
            quotations_cursor = []
            if 'quotations' in mongo_db.list_collection_names():
                quotations_cursor = list(mongo_db.quotations.find({}, {'created_at': 1}))

            month_names = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]

            monthly_totals = {}
            role_counts = {}

            for doc in users_cursor:
                dt = _parse_datetime(doc.get('created_at'))
                if not dt:
                    continue
                year, month = dt.year, dt.month
                if not (1 <= month <= 12):
                    continue
                key = (year, month)
                monthly_entry = monthly_totals.setdefault(key, {'users': 0, 'quotations': 0})
                monthly_entry['users'] += 1

                role = (doc.get('role') or 'user').lower()
                role_counts[role] = role_counts.get(role, 0) + 1

            for doc in quotations_cursor:
                dt = _parse_datetime(doc.get('created_at'))
                if not dt:
                    continue
                year, month = dt.year, dt.month
                if not (1 <= month <= 12):
                    continue
                key = (year, month)
                monthly_entry = monthly_totals.setdefault(key, {'users': 0, 'quotations': 0})
                monthly_entry['quotations'] += 1

            for (year, month) in sorted(monthly_totals.keys()):
                month_label = f"{month_names[month - 1]} {year}"
                result['users_by_month'].append({
                    'month': month_label,
                    'users': monthly_totals[(year, month)]['users'],
                    'quotations': monthly_totals[(year, month)]['quotations']
                })

            result['user_roles'] = [
                {'role': role, 'count': count}
                for role, count in sorted(role_counts.items(), key=lambda item: item[1], reverse=True)
            ]

        else:
            users_json = _load_users_json()
            month_names = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]
            monthly_totals = {}
            role_counts = {}

            for user in users_json.values():
                created_at = user.created_at if isinstance(user, User) else user.get('created_at')
                dt = _parse_datetime(created_at)
                if dt:
                    year, month = dt.year, dt.month
                    if 1 <= month <= 12:
                        key = (year, month)
                        monthly_entry = monthly_totals.setdefault(key, {'users': 0, 'quotations': 0})
                        monthly_entry['users'] += 1

                role = (user.role if isinstance(user, User) else user.get('role')) or 'user'
                role_counts[role.lower()] = role_counts.get(role.lower(), 0) + 1

            for (year, month) in sorted(monthly_totals.keys()):
                month_label = f"{month_names[month - 1]} {year}"
                result['users_by_month'].append({
                    'month': month_label,
                    'users': monthly_totals[(year, month)]['users'],
                    'quotations': monthly_totals[(year, month)]['quotations']
                })

            result['user_roles'] = [
                {'role': role, 'count': count}
                for role, count in sorted(role_counts.items(), key=lambda item: item[1], reverse=True)
            ]

        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in admin_chart_data: {e}")
        return jsonify({'error': 'Failed to load chart data'}), 500
# User profile page
@app.route('/profile')
@login_required
def profile():
    """Render the user profile page."""
    user = current_user
    # Pass helper so template can look up company name by ID
    return render_template('profile/profile.html', user=user, get_company_name_by_id=get_company_name_by_id)

print("==============================\n")
JWT_SECRET = os.getenv('JWT_SECRET', 'your-secret-key')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION = 3600  # 1 hour

# Email Configuration (Resend)
RESEND_API_KEY = os.getenv('RESEND_API_KEY')
EMAIL_FROM = os.getenv('EMAIL_FROM')
EMAIL_FROM_NAME = os.getenv('EMAIL_FROM_NAME', 'Product Calculator')
RESEND_FROM_ADDRESS = os.getenv('RESEND_FROM') or (
    f"{EMAIL_FROM_NAME} <{EMAIL_FROM}>" if EMAIL_FROM else None
)

# Frontend URL for email links
FRONTEND_URL = os.getenv('FRONTEND_URL', 'http://localhost:3000')

# Persistent file paths (Render users can attach a Disk at /var/data or set USERS_FILE_PATH/CART_FILE_PATH)
def _resolve_data_dir():
    # Determine writable directory for persistence
    preferred = os.getenv('DATA_DIR', '/var/data')
    try:
        os.makedirs(preferred, exist_ok=True)
        test_path = os.path.join(preferred, '.write_test')
        with open(test_path, 'w') as fp:
            fp.write('ok')
        os.remove(test_path)
        return preferred
    except Exception:
        fallback = os.path.join('static', 'data')
        os.makedirs(fallback, exist_ok=True)
        return fallback

DATA_DIR = _resolve_data_dir()
USERS_FILE = os.getenv('USERS_FILE_PATH', os.path.join(DATA_DIR, 'users.json'))
CART_FILE = os.getenv('CART_FILE_PATH', os.path.join(DATA_DIR, 'cart.json'))

# User class
def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value)
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            try:
                return datetime.fromtimestamp(float(value))
            except (ValueError, TypeError):
                return None
    return None


def log_time(message: str):
    print(f"[{datetime.utcnow().isoformat()}] {message}")


def normalize_assigned_companies(assigned):
    if not assigned:
        return []
    normalized = []
    for company_id in assigned:
        if not company_id:
            continue
        if isinstance(company_id, ObjectId):
            normalized.append(str(company_id))
        else:
            normalized.append(str(company_id))
    return normalized


class User(UserMixin):
    def __init__(self, id, email, username, password_hash, is_verified=False, otp_verified=False, cart=None, reset_token=None, reset_token_expiry=None, company_id=None, role='user', created_at=None, assigned_companies=None, phone=None):
        self.id = id
        self.email = email
        self.username = username
        self.password_hash = password_hash
        self.is_verified = is_verified
        self.otp_verified = otp_verified
        self.cart = cart if cart is not None else []
        self.reset_token = reset_token
        self.reset_token_expiry = reset_token_expiry
        self.company_id = company_id
        self.phone = phone
        self.role = role or 'user'
        self.created_at = created_at or datetime.utcnow()
        self.assigned_companies = normalize_assigned_companies(assigned_companies)

    def to_dict(self):
        return {
            'id': self.id,
            'email': self.email,
            'username': self.username,
            'password_hash': self.password_hash,
            'is_verified': self.is_verified,
            'phone': self.phone,
            'reset_token': self.reset_token,
            'reset_token_expiry': self.reset_token_expiry.isoformat() if self.reset_token_expiry else None,
            'otp_verified': self.otp_verified,
            'company_id': self.company_id,
            'assigned_companies': [str(cid) for cid in (self.assigned_companies or []) if cid],
            'role': self.role,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def generate_auth_token(self, expires_in=JWT_EXPIRATION):
        return jwt.encode(
            {'user_id': self.id, 'exp': datetime.utcnow() + timedelta(seconds=expires_in)},
            JWT_SECRET,
            algorithm=JWT_ALGORITHM
        )

    @staticmethod
    def verify_auth_token(token):
        try:
            data = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            return data.get('user_id')
        except:
            return None

# ---------------------------------------------------------------------------
# JSON persistence helpers
# ---------------------------------------------------------------------------

# --- Admin helper utilities (serialization + fallback saves) ---


def _parse_float(value, default=0.0):
    """Safely parse a number from string/decimal/None."""
    try:
        return float(value)
    except Exception:
        return default


COMPANY_PLACEHOLDER = '-'


def _normalize_company_text(value, placeholder=COMPANY_PLACEHOLDER):
    if value is None:
        return placeholder
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else placeholder
    if value == '' or value == []:
        return placeholder
    return str(value)


def _normalize_company_datetime(value, placeholder=COMPANY_PLACEHOLDER):
    if isinstance(value, datetime):
        return value.astimezone(IST).strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, dict) and '$date' in value:
        raw = value.get('$date')
        if isinstance(raw, str):
            try:
                parsed = datetime.fromisoformat(raw.replace('Z', '+00:00'))
                return parsed.astimezone(IST).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return raw
        return str(raw)
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else placeholder
    return placeholder


def _get_company_field(company_doc, *keys):
    for key in keys:
        if key in company_doc:
            value = company_doc.get(key)
            if isinstance(value, str):
                value = value.strip()
            if value not in (None, '', []):
                return value
    return None


def normalize_company_record(company_doc):
    if not company_doc:
        return {}
    cid = str(company_doc.get('id') or company_doc.get('_id') or company_doc.get('Company ID', '') or '')
    name = _get_company_field(company_doc, 'name', 'Company Name')
    email = _get_company_field(company_doc, 'email', 'EmailID')
    phone = _get_company_field(company_doc, 'phone', 'Phone')
    last_payment_terms = _get_company_field(company_doc, 'last_payment_terms', 'Last Payment Terms')
    billing_attention = _get_company_field(company_doc, 'billing_attention', 'Billing Attention')
    billing_address = _get_company_field(
        company_doc,
        'billing_address',
        'Billing Address',
        'address',
        'Address'
    )
    billing_street = _get_company_field(company_doc, 'billing_street', 'Billing Street')
    billing_city = _get_company_field(company_doc, 'billing_city', 'Billing City')
    billing_state = _get_company_field(company_doc, 'billing_state', 'Billing State')
    billing_postal_code = _get_company_field(company_doc, 'billing_postal_code', 'Billing Postal Code')
    billing_phone = _get_company_field(company_doc, 'billing_phone', 'Billing Phone')
    gst_registered = company_doc.get('GST Registered', company_doc.get('gst_registered'))
    gst_number = company_doc.get('GST Number', company_doc.get('gst_number'))
    created_at = company_doc.get('created_at') or company_doc.get('Created At')
    assigned = normalize_assigned_companies(company_doc.get('assigned_to', []))

    normalized = {
        'id': cid,
        'name': _normalize_company_text(name),
        'email': _normalize_company_text(email),
        'phone': _normalize_company_text(phone),
        'last_payment_terms': _normalize_company_text(last_payment_terms),
        'billing_attention': _normalize_company_text(billing_attention),
        'billing_address': _normalize_company_text(billing_address),
        'billing_street': _normalize_company_text(billing_street),
        'billing_city': _normalize_company_text(billing_city),
        'billing_state': _normalize_company_text(billing_state),
        'billing_postal_code': _normalize_company_text(billing_postal_code),
        'billing_phone': _normalize_company_text(billing_phone),
        'gst_registered': bool(gst_registered) if gst_registered is not None else False,
        'gst_number': (gst_number or '').strip().upper() if isinstance(gst_number, str) else (gst_number if isinstance(gst_number, (int, float)) else ''),
        'assigned_to': assigned,
        'assigned_to_count': len(assigned),
        'created_at': _normalize_company_datetime(created_at),
        'address': _normalize_company_text(billing_address)
    }
    return normalized


def build_quotation_company_details(selected_company=None, session_company_id=None, session_company_email=None, fallback=None):
    """Resolve company details (address & GST) for quotation displays."""
    placeholder = "--"
    gst_unregistered_label = "URP"

    selected_company = selected_company or {}
    fallback = fallback or {}
    company_id = selected_company.get('id') or session_company_id
    company_email = selected_company.get('email') or session_company_email or fallback.get('email')
    company_name = selected_company.get('name') or fallback.get('name')

    company_record = None
    try:
        companies = load_companies_data()
    except Exception as e:
        app.logger.error(f"Failed to load companies for quotation details: {e}")
        companies = []

    if company_id and companies:
        company_id_str = str(company_id)
        for record in companies:
            if str(record.get('id')) == company_id_str or str(record.get('_id')) == company_id_str:
                company_record = record
                break

    if company_record is None and company_email and companies:
        email_lower = company_email.strip().lower()
        for record in companies:
            record_email = (record.get('email') or record.get('EmailID') or '').strip().lower()
            if record_email and record_email == email_lower:
                company_record = record
                break

    name = (company_record.get('name') or company_record.get('Company Name')) if company_record else None
    if not name:
        name = company_name or 'Not specified'

    email = (company_record.get('email') or company_record.get('EmailID')) if company_record else None
    if not email:
        email = company_email or placeholder

    address = ''
    if company_record:
        address = (
            company_record.get('billing_address')
            or company_record.get('Billing Address')
            or company_record.get('address')
            or company_record.get('Address')
            or ''
        )
    address = address.strip() if isinstance(address, str) else ''
    if not address:
        address = placeholder

    gst_registered = False
    gst_registered_value = None
    if company_record:
        gst_registered_value = (
            company_record.get('gst_registered')
            if 'gst_registered' in company_record
            else company_record.get('GST Registered')
        )
    if gst_registered_value is not None:
        if isinstance(gst_registered_value, str):
            gst_registered = gst_registered_value.strip().lower() in {'true', '1', 'yes', 'y'}
        else:
            gst_registered = bool(gst_registered_value)

    gst_number_raw = ''
    if company_record:
        gst_number_raw = company_record.get('gst_number') or company_record.get('GST Number') or ''
    gst_number = ''
    if isinstance(gst_number_raw, (int, float)):
        gst_number = str(gst_number_raw)
    elif isinstance(gst_number_raw, str):
        gst_number = gst_number_raw.strip().upper()

    if gst_registered:
        gst_display = gst_number if gst_number else placeholder
    else:
        gst_display = gst_unregistered_label

    return {
        'name': name,
        'email': email or placeholder,
        'address': address,
        'gst_registered': gst_registered,
        'gst_number': gst_number,
        'gst_display': gst_display
    }


COMPANY_IMPORT_REQUIRED_HEADERS = [
    'Company Name',
    'EmailID',
    'Phone',
    'Billing Attention',
    'Billing Address',
    'Billing City',
    'Billing State',
    'Postal Code',
    'Billing Phone',
    'Users Assigned To'
]

COMPANY_IMPORT_OPTIONAL_HEADERS = [
    'Sr No',
    'Billing Street',
    'Created Time'
]


def _clean_import_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned == '-' or cleaned.lower() == 'none':
            return ''
        return cleaned
    if isinstance(value, (int, float)):
        text = str(value).strip()
        return text
    if isinstance(value, datetime):
        return value
    return str(value).strip()


def _parse_import_datetime(value):
    if isinstance(value, datetime):
        return value
    cleaned = _clean_import_value(value)
    if not cleaned:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y %H:%M:%S'):
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(cleaned)
    except Exception:
        return None


def _parse_import_assigned_to(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    cleaned = _clean_import_value(value)
    if not cleaned:
        return []
    parts = [part.strip() for part in cleaned.split(',')]
    return [part for part in parts if part]


def _build_company_store_payload(record):
    payload = {
        'Company Name': record.get('Company Name', ''),
        'EmailID': record.get('EmailID', ''),
        'Phone': record.get('Phone', ''),
        'Billing Attention': record.get('Billing Attention', ''),
        'Billing Address': record.get('Billing Address', ''),
        'Billing Street': record.get('Billing Street', ''),
        'Billing City': record.get('Billing City', ''),
        'Billing State': record.get('Billing State', ''),
        'Billing Postal Code': record.get('Billing Postal Code', ''),
        'Billing Phone': record.get('Billing Phone', ''),
        'Address': record.get('Billing Address', '')
    }
    return payload


def _validate_company_import_headers(headers):
    normalized = [str(header).strip() if header else '' for header in headers]
    missing = [required for required in COMPANY_IMPORT_REQUIRED_HEADERS if required not in normalized]
    return normalized, missing


def _extract_row_data(headers, row):
    record = {}
    for header, value in zip(headers, row):
        if not header:
            continue
        record[header] = _clean_import_value(value)
    return record


def _resolve_company_identifier(record):
    email = _clean_import_value(record.get('EmailID'))
    if email:
        return 'email', email
    name = _clean_import_value(record.get('Company Name'))
    if name:
        return 'name', name
    return None, None


def _convert_record_to_storage(record):
    payload = {
        'Company Name': _clean_import_value(record.get('Company Name')),
        'EmailID': _clean_import_value(record.get('EmailID')),
        'Phone': _clean_import_value(record.get('Phone')),
        'Billing Attention': _clean_import_value(record.get('Billing Attention')),
        'Billing Address': _clean_import_value(record.get('Billing Address')),
        'Billing Street': _clean_import_value(record.get('Billing Street')),
        'Billing City': _clean_import_value(record.get('Billing City')),
        'Billing State': _clean_import_value(record.get('Billing State')),
        'Billing Postal Code': _clean_import_value(record.get('Postal Code') or record.get('Billing Postal Code')),
        'Billing Phone': _clean_import_value(record.get('Billing Phone')),
        'Address': _clean_import_value(record.get('Billing Address'))
    }

    created_at = _parse_import_datetime(record.get('Created Time'))
    assigned_to = _parse_import_assigned_to(record.get('Users Assigned To'))
    return payload, created_at, assigned_to


def _upsert_company_mongo(identifier_key, identifier_value, payload, assigned_to, created_at):
    query = {}
    if identifier_key == 'email':
        query['EmailID'] = {'$regex': f'^{re.escape(identifier_value)}$', '$options': 'i'}
    elif identifier_key == 'name':
        query['Company Name'] = {'$regex': f'^{re.escape(identifier_value)}$', '$options': 'i'}
    else:
        return None, False, []

    existing_doc = mongo_db.companies.find_one(query, {'_id': 1, 'assigned_to': 1})
    previous_assigned = normalize_assigned_companies(existing_doc.get('assigned_to', [])) if existing_doc else []
    created_at_value = created_at or datetime.utcnow()

    update_doc = {
        '$set': {
            **payload,
            'assigned_to': assigned_to,
            'updated_at': datetime.utcnow()
        },
        '$setOnInsert': {
            'created_at': created_at_value,
            'created_by': str(current_user.id)
        }
    }

    updated_doc = mongo_db.companies.find_one_and_update(
        query,
        update_doc,
        upsert=True,
        return_document=ReturnDocument.AFTER,
        projection={'_id': 1}
    )

    company_id = str(updated_doc['_id']) if updated_doc else None
    was_inserted = existing_doc is None and company_id is not None

    return company_id, was_inserted, previous_assigned


def _upsert_company_json(identifier_key, identifier_value, payload, assigned_to, created_at):
    companies = load_companies_data()
    identifier_value_lower = (identifier_value or '').lower()
    assigned_normalized = normalize_assigned_companies(assigned_to)
    created_at_iso = created_at.isoformat() if isinstance(created_at, datetime) else (created_at or datetime.utcnow().isoformat())

    for company in companies:
        if identifier_key == 'email':
            compare_value = (company.get('EmailID') or company.get('email') or '').lower()
        else:
            compare_value = (company.get('Company Name') or company.get('name') or '').lower()

        if compare_value and compare_value == identifier_value_lower:
            previous_assigned = normalize_assigned_companies(company.get('assigned_to', []))
            company.update(payload)
            company['assigned_to'] = assigned_normalized
            company['updated_at'] = datetime.utcnow().isoformat()
            if created_at:
                company['created_at'] = created_at_iso
            save_companies_data(companies)
            return company.get('id') or company.get('_id'), False, previous_assigned

    new_company = payload.copy()
    new_company['id'] = str(uuid.uuid4())
    new_company['assigned_to'] = assigned_normalized
    new_company['created_at'] = created_at_iso
    new_company['created_by'] = str(current_user.id)
    companies.append(new_company)
    save_companies_data(companies)
    return new_company['id'], True, []


def _sync_assigned_users(company_id, assigned_to, previous_assigned=None):
    if not company_id:
        return

    normalized_current = normalize_assigned_companies(assigned_to)
    normalized_previous = normalize_assigned_companies(previous_assigned or [])

    if set(normalized_current) == set(normalized_previous):
        return

    sync_company_user_links(str(company_id), normalized_current)


def serialize_admin_user(user_doc):
    """Normalize user document (Mongo or JSON) for admin UI."""
    if not user_doc:
        return {}
    # Mongo docs may have ObjectId; JSON already string id
    uid = str(user_doc.get('id') or user_doc.get('_id') or user_doc.get('user_id', ''))
    return {
        'id': uid,
        'username': user_doc.get('username') or user_doc.get('email', '')[: user_doc.get('email', '').find('@')],
        'email': user_doc.get('email') or user_doc.get('Email', ''),
        'phone': user_doc.get('phone') or '',
        'position': user_doc.get('position') or '',
        'role': user_doc.get('role', 'user'),
        'is_verified': bool(user_doc.get('is_verified', True)),
        'created_at': str(user_doc.get('created_at') or user_doc.get('Created', '')),
        'customers': user_doc.get('customers_assigned', user_doc.get('customers', [])),
        'assigned_companies': [str(cid) for cid in user_doc.get('assigned_companies', []) if cid]
    }


def serialize_admin_company(company_doc):
    return normalize_company_record(company_doc)


def _extract_company_id(company_dict):
    if not company_dict:
        return ''
    return str(
        company_dict.get('id') or
        company_dict.get('_id') or
        company_dict.get('Company ID') or ''
    )


def _extract_company_assigned_users(company_dict):
    existing = company_dict.get('assigned_to', []) if isinstance(company_dict, dict) else []
    return normalize_assigned_companies(existing)


def sync_user_company_links(user_id, assigned_company_ids):
    user_id = str(user_id or '').strip()
    assigned_set = set(normalize_assigned_companies(assigned_company_ids))

    if not user_id:
        return

    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
            except Exception as ping_error:
                app.logger.error(f"Mongo ping failed in sync_user_company_links: {ping_error}")
                return

            object_ids = []
            invalid_ids = []
            for cid in assigned_set:
                if ObjectId.is_valid(cid):
                    object_ids.append(ObjectId(cid))
                else:
                    invalid_ids.append(cid)

            try:
                remove_filter = {'assigned_to': user_id}
                if object_ids:
                    remove_filter['_id'] = {'$nin': object_ids}
                mongo_db.companies.update_many(remove_filter, {'$pull': {'assigned_to': user_id}})
            except Exception as remove_error:
                app.logger.error(f"Error removing user-company links: {remove_error}")

            if object_ids:
                try:
                    mongo_db.companies.update_many({'_id': {'$in': object_ids}}, {'$addToSet': {'assigned_to': user_id}})
                except Exception as add_error:
                    app.logger.error(f"Error bulk-adding user {user_id} to companies {assigned_set}: {add_error}")

            if invalid_ids:
                app.logger.warning(f"Cannot sync company assignment for user {user_id} due to invalid company ids: {invalid_ids}")
        else:
            companies = load_companies_data()
            updated = False
            for company in companies:
                cid = _extract_company_id(company)
                if not cid:
                    continue
                current_users = _extract_company_assigned_users(company)
                if cid in assigned_set:
                    if user_id not in current_users:
                        current_users.append(user_id)
                        updated = True
                else:
                    if user_id in current_users:
                        current_users = [uid for uid in current_users if uid != user_id]
                        updated = True
                company['assigned_to'] = current_users

            if updated:
                save_companies_data(companies)
    except Exception as e:
        app.logger.error(f"Unexpected error in sync_user_company_links: {e}", exc_info=True)


def _extract_user_assigned_companies(user_record):
    if isinstance(user_record, User):
        return normalize_assigned_companies(getattr(user_record, 'assigned_companies', []))
    if isinstance(user_record, dict):
        return normalize_assigned_companies(user_record.get('assigned_companies', []))
    return []


def _set_user_assigned_companies(user_record, assignments):
    normalized = normalize_assigned_companies(assignments)
    if isinstance(user_record, User):
        user_record.assigned_companies = normalized
    elif isinstance(user_record, dict):
        user_record['assigned_companies'] = normalized


def sync_company_user_links(company_id, assigned_user_ids):
    company_id = str(company_id or '').strip()
    assigned_set = set(normalize_assigned_companies(assigned_user_ids))

    if not company_id:
        return

    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.command('ping')
            except Exception as ping_error:
                app.logger.error(f"Mongo ping failed in sync_company_user_links: {ping_error}")
                return

            try:
                current_cursor = mongo_db.users.find({'assigned_companies': company_id}, {'assigned_companies': 1})
                processed = set()
                for user in current_cursor:
                    uid_str = str(user.get('_id'))
                    processed.add(uid_str)
                    if uid_str not in assigned_set:
                        mongo_db.users.update_one({'_id': user['_id']}, {'$pull': {'assigned_companies': company_id}})

                for uid in assigned_set:
                    try:
                        if ObjectId.is_valid(uid):
                            mongo_db.users.update_one({'_id': ObjectId(uid)}, {'$addToSet': {'assigned_companies': company_id}})
                        else:
                            mongo_db.users.update_one({'_id': uid}, {'$addToSet': {'assigned_companies': company_id}})
                    except Exception as add_error:
                        app.logger.error(f"Error syncing company {company_id} to user {uid}: {add_error}")
            except Exception as mongo_error:
                app.logger.error(f"Error syncing company-user links (mongo): {mongo_error}")
        else:
            users_dict = load_users()
            updated = False
            for uid, record in users_dict.items():
                uid_str = str(getattr(record, 'id', uid))
                current = _extract_user_assigned_companies(record)
                if uid_str in assigned_set:
                    if company_id not in current:
                        current.append(company_id)
                        updated = True
                        _set_user_assigned_companies(record, current)
                else:
                    if company_id in current:
                        current = [cid for cid in current if cid != company_id]
                        _set_user_assigned_companies(record, current)
                        updated = True

            if updated:
                save_users(users_dict)
    except Exception as e:
        app.logger.error(f"Unexpected error in sync_company_user_links: {e}", exc_info=True)


def serialize_admin_quotation(q_doc):
    if not q_doc:
        return {}
    qid = str(q_doc.get('quote_id') or q_doc.get('id') or q_doc.get('_id') or '')
    total_pre = _parse_float(q_doc.get('total_amount_pre_gst') or q_doc.get('subtotal_pre_gst') or q_doc.get('total_pre_gst'))
    total_post = _parse_float(q_doc.get('total_amount_post_gst') or q_doc.get('total_post_gst'))
    total_gst = _parse_float(q_doc.get('total_gst') or q_doc.get('gst_amount'))
    subtotal_before_discount = _parse_float(q_doc.get('subtotal_before_discount'))
    subtotal_after_discount = _parse_float(q_doc.get('subtotal_after_discount') or q_doc.get('total_amount_pre_gst'))
    total_discount = _parse_float(q_doc.get('total_discount'))
    return {
        'id': qid,
        'username': q_doc.get('username') or q_doc.get('user_name', ''),
        'user_email': q_doc.get('user_email', ''),
        'company_name': q_doc.get('company_name') or q_doc.get('Company Name', ''),
        'company_email': q_doc.get('company_email') or q_doc.get('EmailID', ''),
        'products_count': q_doc.get('products_count') or len(q_doc.get('products', [])),
        'total_amount_pre_gst': total_pre,
        'total_amount_post_gst': total_post,
        'total_gst': total_gst,
        'subtotal_before_discount': subtotal_before_discount,
        'subtotal_after_discount': subtotal_after_discount,
        'total_discount': total_discount,
        'discount_text': q_doc.get('discount_text', ''),
        'from_company': q_doc.get('from_company', ''),
        'from_email': q_doc.get('from_email', ''),
        'prepared_by_name': q_doc.get('prepared_by_name', q_doc.get('username', '')),
        'prepared_by_email': q_doc.get('prepared_by_email', q_doc.get('user_email', '')),
        'notes': q_doc.get('notes', ''),
        'products': q_doc.get('products', []),
        'created_at': str(q_doc.get('created_at') or ''),
        'created_at_iso': q_doc.get('created_at_iso', ''),
        'created_at_ist': q_doc.get('created_at_ist', ''),
        'created_at_utc': q_doc.get('created_at_utc', ''),
        'generated_at_date_display': q_doc.get('generated_at_date_display', ''),
        'generated_at_time_display': q_doc.get('generated_at_time_display', '')
    }


# ----- Admin data loaders -----

def load_admin_users():
    """Return list of serialized users from Mongo or JSON fallback."""
    users_list = []
    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            mongo_db.command('ping')
            cursor = mongo_db.users.find({})
            for doc in cursor:
                doc['id'] = str(doc.pop('_id'))
                users_list.append(serialize_admin_user(doc))
        else:
            users_json = _load_users_json()
            for u in users_json.values():
                users_list.append(serialize_admin_user(u))
    except Exception as e:
        app.logger.error(f"Error in load_admin_users: {e}")
    return users_list


# ----- Company JSON saver for fallback -----

def save_companies_data(companies):
    """Save companies list to JSON file as fallback when Mongo not used."""
    try:
        data_dir = os.path.join(app.root_path, 'static', 'data')
        os.makedirs(data_dir, exist_ok=True)
        target = os.path.join(data_dir, 'companies.json')
        with open(target, 'w', encoding='utf-8') as f:
            json.dump({'companies': companies}, f, ensure_ascii=False, indent=2, default=str)
        return True
    except Exception as e:
        app.logger.error(f"Error saving companies JSON: {e}")
        return False
# Existing private helpers (_load_users_json / _save_users_json) are used by
# the rest of the code via these thin wrappers so the earlier calls to
# load_users()/save_users() continue to work without refactor.

def _load_users_json():
    """Load users from JSON file."""
    try:
        # Ensure the directory exists
        os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
        
        # Try to read the file
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                if not content.strip():
                    content = '{}'
                users_data = json.loads(content)
        except (FileNotFoundError, json.JSONDecodeError):
            # If file doesn't exist or is invalid JSON, create a new empty file
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                f.write('{}')
            users_data = {}
            
        users = {}
        for user_id, user_data in users_data.items():
            try:
                # Ensure all required fields exist
                if not all(key in user_data for key in ['email', 'username', 'password_hash']):
                    print(f"Skipping invalid user data: missing required fields")
                    continue
                
                users[user_id] = User(
                    id=user_id,
                    email=user_data['email'],
                    username=user_data['username'],
                    password_hash=user_data['password_hash'],
                    is_verified=user_data.get('is_verified', False),
                    otp_verified=user_data.get('otp_verified', False),
                    cart=user_data.get('cart', []),
                    reset_token=user_data.get('reset_token'),
                    reset_token_expiry=datetime.fromisoformat(user_data.get('reset_token_expiry')) if user_data.get('reset_token_expiry') else None,
                    company_id=user_data.get('company_id'),
                    role=user_data.get('role', 'user'),
                    created_at=_parse_datetime(user_data.get('created_at')),
                    assigned_companies=user_data.get('assigned_companies', [])
                )
            except Exception as e:
                print(f"Error loading user {user_id}: {e}")
                continue
        return users
    except Exception as e:
        print(f"Error loading users: {e}")
        try:
            # Create a fresh empty file with proper encoding
            os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                f.write('{}')
            return {}  # Return empty dict after creating new file
        except Exception as e:
            print(f"Error creating users file: {e}")
        return {}

# ... (rest of the code remains the same)

@login_manager.user_loader
def load_user(user_id):
    """Load user by ID from either MongoDB or JSON."""
    if MONGO_AVAILABLE and USE_MONGO:
        # Try MongoDB first
        try:
            print(f'Loading user from MongoDB with ID: {user_id}')
            doc = mu_find_user_by_id(user_id)
            if not doc:
                print(f'User not found in MongoDB with ID: {user_id}')
                return None
                
            user = User(
                id=str(doc['_id']),  # Convert ObjectId to string
                email=doc['email'],
                username=doc['username'],
                password_hash=doc['password_hash'],
                is_verified=doc.get('is_verified', False),
                otp_verified=doc.get('otp_verified', False),
                company_id=doc.get('company_id'),
                phone=doc.get('phone'),
                role=doc.get('role', 'user'),
                created_at=_parse_datetime(doc.get('created_at')),
                assigned_companies=doc.get('assigned_companies', [])
            )
            print(f'Successfully loaded user: {user.email} (ID: {user.id})')
            return user
        except Exception as e:
            print(f"Error loading user {user_id}: {e}")
            return None
    
    # Fall back to JSON users
    users = _load_users_json()
    user_data = users.get(user_id) if hasattr(users, 'get') else None
    if user_data:
        return User(
            id=user_id,
            email=user_data['email'],
            username=user_data.get('username', user_data['email'].split('@')[0]),
            password_hash=user_data['password_hash'],
            is_verified=user_data.get('is_verified', False),
            otp_verified=user_data.get('otp_verified', False),
            cart=user_data.get('cart', []),
            reset_token=user_data.get('reset_token'),
            reset_token_expiry=user_data.get('reset_token_expiry'),
            company_id=user_data.get('company_id'),
            phone=user_data.get('phone'),
            role=user_data.get('role', 'user'),
            created_at=_parse_datetime(user_data.get('created_at')),
            assigned_companies=user_data.get('assigned_companies', [])
        )
    return None

def save_users(users_dict=None):
    """Legacy wrapper around _save_users_json."""
    return _save_users_json(users_dict)

def _save_users_json(users_dict=None):
    """Save users to JSON file. If no argument is provided, saves the global users dictionary."""
    try:
        # Ensure the directory exists
        os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
        
        # Get the users data to save
        if users_dict is None:
            users_dict = users
            
        # Create a temporary file
        temp_file = USERS_FILE + '.tmp'
        
        # Convert users to dictionary format
        user_data = {user_id: user.to_dict() for user_id, user in users_dict.items()}
        
        # Write to temporary file
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(user_data, f, indent=2)
        
        # Replace original file with temporary file atomically
        try:
            os.replace(temp_file, USERS_FILE)
        except FileNotFoundError:
            # If file doesn't exist, just rename the temp file
            os.rename(temp_file, USERS_FILE)
        
        # Verify the file was saved correctly
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                saved_data = json.load(f)
                if len(saved_data) != len(user_data):
                    raise Exception("File save verification failed")
        except Exception as e:
            print(f"Error verifying saved file: {e}")
            return False
        
        return True
    except Exception as e:
        print(f"Error saving users: {e}")
        try:
            # Clean up temp file if it exists
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception as cleanup_error:
            print(f"Error cleaning up temp file: {cleanup_error}")
        return False

# ----- Mongo wrappers overriding JSON if USE_MONGO -----
if USE_MONGO:
    def load_users():
        users_local = {}
        try:
            for doc in users_col.find():
                users_local[doc['_id']] = User(
                    id=doc['_id'],
                    email=doc.get('email'),
                    username=doc.get('username'),
                    password_hash=doc.get('password_hash'),
                    is_verified=doc.get('is_verified', False),
                    otp_verified=doc.get('otp_verified', False)
                )
        except Exception as e:
            print(f"Error loading users from MongoDB: {e}")
        return users_local

    def save_users(users_dict=None):
        try:
            if users_dict is None:
                users_dict = users
            for uid, user in users_dict.items():
                users_col.update_one({'_id': uid}, {'$set': user.to_dict()}, upsert=True)
            return True
        except Exception as e:
            print(f"Error saving users to MongoDB: {e}")
            return False

    ensure_mongo_connection_initialized()
    users = load_users() # Initialize users from MongoDB
else:
    # Fallback to JSON versions defined above
    load_users = _load_users_json
    save_users = _save_users_json
    users = load_users() # Initialize users from JSON

# Add logging for debugging

print("Resend Configuration:\n"
      f"RESEND_API_KEY: {'Set' if RESEND_API_KEY else 'Not set'}\n"
      f"RESEND_FROM_ADDRESS: {RESEND_FROM_ADDRESS}\n"
      f"EMAIL_FROM: {EMAIL_FROM}\n"
      f"EMAIL_FROM_NAME: {EMAIL_FROM_NAME}")

def check_email_config():
    """Check if Resend email configuration is valid."""
    if not RESEND_API_KEY:
        print("Warning: Resend API key is not configured")
        return False
    if not RESEND_FROM_ADDRESS:
        print("Warning: Resend from address is not configured")
        return False
    resend.api_key = RESEND_API_KEY
    return True

# Initialize email configuration
email_config_valid = check_email_config()

def refresh_email_config():
    """Periodically refresh email configuration."""
    global email_config_valid
    email_config_valid = check_email_config()


def send_email_resend(
    to,
    subject: str,
    html: str | None = None,
    text: str | None = None,
    from_email: str | None = None,
    cc=None,
    bcc=None,
    reply_to: str | None = None,
    attachments=None
) -> bool:
    """Send an email via Resend API."""
    if not email_config_valid:
        app.logger.error("Resend email configuration is invalid")
        return False

    recipients = to if isinstance(to, list) else [to]
    from_address = from_email or RESEND_FROM_ADDRESS

    if not from_address:
        app.logger.error("Resend sender address is not configured")
        return False

    payload = {
        "from": from_address,
        "sender": from_address,
        "to": recipients,
        "subject": subject
    }

    if html:
        payload["html"] = html
    if text:
        payload["text"] = text
    if cc:
        payload["cc"] = cc if isinstance(cc, list) else [cc]
    if bcc:
        payload["bcc"] = bcc if isinstance(bcc, list) else [bcc]
    if reply_to:
        payload["reply_to"] = reply_to

    if attachments:
        payload["attachments"] = attachments

    # Resend requires at least one of html/text
    if "html" not in payload and "text" not in payload:
        app.logger.error("Resend payload missing both html and text content")
        return False

    try:
        resend.Emails.send(payload)
        return True
    except Exception as exc:
        app.logger.error(f"Error sending email via Resend: {exc}", exc_info=True)
        return False

# Initialize Flask app with logging
import logging
import sys
from logging.handlers import RotatingFileHandler

# Configure root logger
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        RotatingFileHandler('app.log', maxBytes=10000, backupCount=1)
    ]
)

# Suppress Flask debug pin console output
logging.getLogger('werkzeug').setLevel(logging.WARNING)

# Configure secret key for existing app instance
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

# Configure session
app.secret_key = os.getenv('SECRET_KEY', 'dev-key-123')
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Helps with CSRF protection
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)  # Session expires after 1 day

# Add regex_search filter to Jinja2 environment
@app.template_filter('regex_search')
def regex_search_filter(s, pattern):
    """Check if the pattern matches the string."""
    if not s or not pattern:
        return False
    return bool(re.search(pattern, str(s)))

app.logger.info("Flask app initialized")

# Initialize cart store
# -------------------- Cart storage abstractions --------------------
class MongoCartStore:
    """MongoDB-backed cart store with one cart document per user."""

    def __init__(self, db):
        self.col = db.get_collection('carts')
        app.logger.info("[DEBUG] Initialized MongoCartStore with collection: %s", self.col.name)

    def _doc(self, user_id):
        doc = self.col.find_one({"user_id": user_id})
        app.logger.debug(
            "[DEBUG] _doc(user_id=%s) - %s",
            user_id,
            f"Found document with {len(doc.get('products', []))} products" if doc else "No document found"
        )
        return doc or {}

    def get_cart(self, user_id):
        app.logger.debug("[DEBUG] get_cart(user_id=%s)", user_id)
        doc = self._doc(user_id)
        products = doc.get('products', [])
        app.logger.debug(
            "[DEBUG] Retrieved cart for user %s with %d products",
            user_id,
            len(products)
        )
        if products:
            app.logger.debug("[DEBUG] Sample product data: %s", str(products[0])[:200])
        return products

    def save_cart(self, user_id, products):
        app.logger.debug(
            "[DEBUG] save_cart(user_id=%s) - Saving %d products",
            user_id,
            len(products)
        )
        if products:
            app.logger.debug("[DEBUG] Sample product being saved: %s", str(products[0])[:200])
            
        result = self.col.update_one(
            {"user_id": user_id},
            {
                "$set": {
                    "products": products,
                    "updated_at": datetime.utcnow(),
                    "user_id": user_id  # Ensure user_id is set
                }
            },
            upsert=True
        )
        app.logger.debug(
            "[DEBUG] Cart save result - Matched: %d, Modified: %d, Upserted ID: %s",
            result.matched_count,
            result.modified_count,
            getattr(result, 'upserted_id', 'N/A')
        )
        return True

    def clear_cart(self, user_id):
        app.logger.info("[DEBUG] Clearing cart for user: %s", user_id)
        return self.save_cart(user_id, [])


# Fallback JSON/in-memory version ----------------------------------
class CartStore:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.cart = cls._load_cart()
        return cls._instance
    
    @staticmethod
    def _load_cart():
        try:
            if os.path.exists(CART_FILE):
                with open(CART_FILE, 'r') as f:
                    return json.load(f)
            return {"products": []}
        except Exception as e:
            print(f"Error loading cart: {e}")
            return {"products": []}
    
    @staticmethod
    def _save_cart(cart):
        try:
            os.makedirs(os.path.dirname(CART_FILE), exist_ok=True)
            with open(CART_FILE, 'w') as f:
                json.dump(cart, f, indent=2)
            return True
        except Exception as e:
            print(f"Error saving cart: {e}")
            return False
    
    def get_cart(self, *_args, **_kwargs):
        """Return stored cart; ignores user since json cart is global."""
        return self.cart
    
    def save_cart(self, cart):
        self.cart = cart
        return self._save_cart(cart)

# Choose the appropriate cart store implementation dynamically
cart_store = None


def get_cart_store():
    global cart_store

    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
        if not isinstance(cart_store, MongoCartStore):
            print("Using MongoCartStore for cart persistence")
            cart_store = MongoCartStore(mongo_db)
    else:
        if not isinstance(cart_store, CartStore):
            print("Using local JSON CartStore for cart persistence")
            cart_store = CartStore()

    return cart_store

# -------------------- Cart helper wrappers --------------------

def get_user_cart():
    """Return a dict with a products list for the current user using MongoDB."""
    try:
        app.logger.info(f"[DEBUG] get_user_cart() called for user: {getattr(current_user, 'id', 'no-user')}")
        
        if not hasattr(current_user, 'id'):
            app.logger.warning("[DEBUG] No current_user.id, returning empty cart")
            return {"products": []}
            
        store = get_cart_store()

        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and isinstance(store, MongoCartStore):
            app.logger.info("[DEBUG] Using MongoDB for cart storage")
            app.logger.info(f"[DEBUG] MongoDB status - MONGO_AVAILABLE: {MONGO_AVAILABLE}, USE_MONGO: {USE_MONGO}, mongo_db: {'available' if mongo_db is not None else 'None'}")
            
            try:
                products = store.get_cart(current_user.id)
                app.logger.info(f"[DEBUG] Retrieved {len(products) if products else 0} products from MongoDB")
                if products:
                    app.logger.debug(f"[DEBUG] Sample product from MongoDB: {str(products[0])[:200]}...")
            except Exception as e:
                app.logger.error(f"[DEBUG] Error fetching cart from MongoDB: {str(e)}")
                products = []
            # Ensure products is a list; fallback if malformed
            if not isinstance(products, list):
                app.logger.warning("[DEBUG] Expected list from cart, got %s. Resetting to empty list.", type(products).__name__)
                products = []

            # Ensure all items in products are dicts with calculations
            sanitized_products = []
            for product in products:
                if not isinstance(product, dict):
                    app.logger.warning("[DEBUG] Skipping malformed product entry: %s", str(product)[:100])
                    continue
                sanitized_products.append(product)

                if product.get('type') == 'mpack':
                    try:
                        underpacking_type = (product.get('underpacking_type') or '').strip().lower()
                        format_label = (product.get('format_label') or '').strip().lower()
                        name_lower = (product.get('name') or '').strip().lower()

                        base_rate_per_100 = 75.0
                        if underpacking_type == 'polipack':
                            if 'self' in format_label or 'adhesive' in format_label or 'self' in name_lower:
                                base_rate_per_100 = 925.0
                            elif 'non' in format_label or 'wa' in format_label or 'non' in name_lower:
                                base_rate_per_100 = 425.0

                        thickness_raw = product.get('thickness') or ''
                        thickness_digits = ''.join(ch for ch in str(thickness_raw) if ch.isdigit() or ch == '.')
                        try:
                            thickness_micron = float(thickness_digits) if thickness_digits else 0.0
                        except ValueError:
                            thickness_micron = 0.0

                        sqm_per_sheet = product.get('standard_area_sqm') or product.get('custom_area_sqm')
                        try:
                            sqm_per_sheet = float(sqm_per_sheet) if sqm_per_sheet is not None else 0.0
                        except (TypeError, ValueError):
                            sqm_per_sheet = 0.0

                        if not sqm_per_sheet:
                            w = product.get('display_width_mm') or product.get('standard_width_mm') or product.get('custom_width_mm')
                            l = product.get('display_length_mm') or product.get('standard_length_mm') or product.get('custom_length_mm')
                            try:
                                w = float(w) if w is not None else 0.0
                            except (TypeError, ValueError):
                                w = 0.0
                            try:
                                l = float(l) if l is not None else 0.0
                            except (TypeError, ValueError):
                                l = 0.0
                            if w and l:
                                sqm_per_sheet = (w * l) / 1_000_000

                        calculations = product.get('calculations') if isinstance(product.get('calculations'), dict) else {}
                        stored_rate = calculations.get('rate_per_sqm')
                        stored_unit_price = calculations.get('unit_price', product.get('unit_price', 0))
                        try:
                            stored_unit_price = float(stored_unit_price) if stored_unit_price is not None else 0.0
                        except (TypeError, ValueError):
                            stored_unit_price = 0.0

                        should_recalc = False
                        if stored_unit_price <= 0 or not sqm_per_sheet or thickness_micron <= 0:
                            should_recalc = True
                        if underpacking_type == 'polipack':
                            try:
                                stored_rate_val = float(stored_rate) if stored_rate is not None else 0.0
                            except (TypeError, ValueError):
                                stored_rate_val = 0.0
                            desired_rate = base_rate_per_100 * (thickness_micron / 100.0) if thickness_micron else 0.0
                            if desired_rate and (not stored_rate_val or abs(stored_rate_val - desired_rate) > 0.01):
                                should_recalc = True

                        if should_recalc:
                            rate_per_sqm = base_rate_per_100 * (thickness_micron / 100.0) if thickness_micron else 0.0
                            unit_price = rate_per_sqm * sqm_per_sheet
                            try:
                                quantity = int(product.get('quantity', 1) or 1)
                            except (TypeError, ValueError):
                                quantity = 1
                            try:
                                discount_percent = float(product.get('discount_percent', 0) or 0)
                            except (TypeError, ValueError):
                                discount_percent = 0.0
                            try:
                                gst_percent = float(product.get('gst_percent', 18) or 18)
                            except (TypeError, ValueError):
                                gst_percent = 18.0

                            subtotal = unit_price * quantity
                            discount_amount = subtotal * (discount_percent / 100.0)
                            discounted_subtotal = subtotal - discount_amount
                            gst_amount = discounted_subtotal * (gst_percent / 100.0)
                            final_total = discounted_subtotal + gst_amount

                            product['unit_price'] = round(unit_price, 2)
                            product['standard_area_sqm'] = round(sqm_per_sheet, 6) if sqm_per_sheet else product.get('standard_area_sqm')
                            product['discount_amount'] = round(discount_amount, 2)
                            product['discounted_subtotal'] = round(discounted_subtotal, 2)
                            product['gst_amount'] = round(gst_amount, 2)
                            product['total_price'] = round(final_total, 2)
                            product['total'] = product['total_price']

                            product['calculations'] = {
                                'rate_per_sqm': round(rate_per_sqm, 2),
                                'sqm_per_sheet': round(sqm_per_sheet, 3),
                                'unit_price': round(unit_price, 2),
                                'quantity': quantity,
                                'subtotal': round(subtotal, 2),
                                'discount_percent': discount_percent,
                                'discount_amount': round(discount_amount, 2),
                                'discounted_subtotal': round(discounted_subtotal, 2),
                                'gst_percent': gst_percent,
                                'gst_amount': round(gst_amount, 2),
                                'final_total': round(final_total, 2)
                            }
                    except Exception as _mpack_recalc_error:
                        app.logger.warning(f"MPack recalc skipped due to error: {_mpack_recalc_error}")

                if 'calculations' not in product or not isinstance(product.get('calculations'), dict):
                    # If calculations are missing or invalid, recalculate them
                    if product.get('type') == 'blanket':
                        base_price = float(product.get('base_price', 0))
                        bar_price = float(product.get('bar_price', 0))
                        quantity = int(product.get('quantity', 1))
                        discount_percent = float(product.get('discount_percent', 0))
                        gst_percent = float(product.get('gst_percent', 18))
                        
                        price_per_unit = base_price + bar_price
                        subtotal = price_per_unit * quantity
                        discount_amount = subtotal * (discount_percent / 100)
                        discounted_subtotal = subtotal - discount_amount
                        gst_amount = (discounted_subtotal * gst_percent) / 100
                        final_total = discounted_subtotal + gst_amount
                        
                        product['unit_price'] = round(price_per_unit, 2)
                        product['calculations'] = {
                            'base_price': round(base_price, 2),
                            'bar_price': round(bar_price, 2),
                            'unit_price': round(price_per_unit, 2),
                            'quantity': quantity,
                            'subtotal': round(subtotal, 2),
                            'discount_percent': discount_percent,
                            'discount_amount': round(discount_amount, 2),
                            'discounted_subtotal': round(discounted_subtotal, 2),
                            'gst_percent': gst_percent,
                            'gst_amount': round(gst_amount, 2),
                            'final_total': round(final_total, 2)
                        }
                    elif product.get('type') == 'mpack':
                        price = float(product.get('unit_price', 0))
                        quantity = int(product.get('quantity', 1))
                        discount_percent = float(product.get('discount_percent', 0))
                        gst_percent = float(product.get('gst_percent', 18))
                        
                        subtotal = price * quantity
                        discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
                        price_after_discount = subtotal - discount_amount
                        gst_amount = (price_after_discount * gst_percent / 100) if gst_percent else 0
                        final_total = price_after_discount + gst_amount
                        
                        product['calculations'] = {
                            'unit_price': round(price, 2),
                            'quantity': quantity,
                            'subtotal': round(subtotal, 2),
                            'discount_percent': discount_percent,
                            'discount_amount': round(discount_amount, 2),
                            'price_after_discount': round(price_after_discount, 2),
                            'gst_percent': gst_percent,
                            'gst_amount': round(gst_amount, 2),
                            'final_total': round(final_total, 2)
                        }
                    elif product.get('type') == 'rule':
                        recalc_rule_pricing(product)
                    else:
                        # Generic fallback for any other item type with missing/zero calculations
                        unit_price = float(product.get('unit_price', 0))
                        quantity = int(product.get('quantity', 1))
                        discount_percent = float(product.get('discount_percent', 0))
                        gst_percent = float(product.get('gst_percent', 18))
                        
                        subtotal = unit_price * quantity
                        discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
                        discounted_subtotal = subtotal - discount_amount
                        gst_amount = (discounted_subtotal * gst_percent / 100) if gst_percent else 0
                        final_total = discounted_subtotal + gst_amount
                        
                        product['calculations'] = {
                            'unit_price': round(unit_price, 2),
                            'quantity': quantity,
                            'subtotal': round(subtotal, 2),
                            'discount_percent': discount_percent,
                            'discount_amount': round(discount_amount, 2),
                            'discounted_subtotal': round(discounted_subtotal, 2),
                            'gst_percent': gst_percent,
                            'gst_amount': round(gst_amount, 2),
                            'final_total': round(final_total, 2)
                        }
            
            products = sanitized_products
            return {"products": products}

        # Fallback to JSON cart store when Mongo is unavailable
        app.logger.warning("[DEBUG] MongoDB is not available for cart storage")
        app.logger.warning(f"[DEBUG] MONGO_AVAILABLE: {MONGO_AVAILABLE}, USE_MONGO: {USE_MONGO}, mongo_db: {'available' if 'mongo_db' in globals() and mongo_db is not None else 'None'}")
        products = store.get_cart()
        if not isinstance(products, list):
            products = []
        return {"products": products}
        
    except Exception as e:
        print(f"Error in get_user_cart: {e}")
        import traceback
        traceback.print_exc()
        return {"products": []}


def get_active_sessions_snapshot():
    """Return a lightweight snapshot of active sessions for dashboard use."""
    sessions = []

    try:
        if not current_user.is_authenticated:
            return sessions

        cart_data = get_user_cart() or {}
        products = cart_data.get('products', []) if isinstance(cart_data, dict) else []

        cart_total = 0.0
        for product in products:
            calculations = product.get('calculations', {}) if isinstance(product, dict) else {}
            if 'final_total' in calculations:
                cart_total += float(calculations.get('final_total', 0) or 0)
            elif 'total_price' in product:
                cart_total += float(product.get('total_price', 0) or 0)

        cart_items_count = len(products)

        selected_company = session.get('selected_company') if isinstance(session.get('selected_company'), dict) else {}
        company_name = session.get('company_name') or selected_company.get('name') or 'Not selected'
        company_email = session.get('company_email') or selected_company.get('email') or ''

        last_activity = datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')

        sessions.append({
            'user_id': getattr(current_user, 'id', None),
            'username': getattr(current_user, 'username', 'Unknown'),
            'email': getattr(current_user, 'email', ''),
            'role': getattr(current_user, 'role', 'user'),
            'company_name': company_name,
            'company_email': company_email,
            'cart_amount': round(cart_total, 2),
            'cart_items_count': cart_items_count,
            'last_activity': last_activity
        })

    except Exception as session_error:
        app.logger.error(f"Error building active sessions snapshot: {session_error}")

    return sessions

def save_user_cart(cart_dict):
    """Persist cart for current user using MongoDB."""
    try:
        if not hasattr(current_user, 'id'):
            print("Cannot save cart: No user ID available")
            return
            
        if not isinstance(cart_dict, dict) or 'products' not in cart_dict:
            print("Invalid cart format")
            return
            
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            cart_store.save_cart(current_user.id, cart_dict['products'])
        else:
            print("MongoDB is not available for cart storage")
            
    except Exception as e:
        print(f"Error in save_user_cart: {e}")
        import traceback
        traceback.print_exc()

# Initialize users dictionary (only for JSON fallback)
if USE_MONGO:
    users = {}
else:
    # Removed this line - it's causing the error
    # users = load_users()
    pass

@login_manager.user_loader
def load_user(user_id):
    if MONGO_AVAILABLE and USE_MONGO:
        try:
            print(f'Loading user from MongoDB with ID: {user_id}')
            doc = mu_find_user_by_id(user_id)
            if not doc:
                print(f'User not found in MongoDB with ID: {user_id}')
                return None
                
            user = User(
                id=str(doc['_id']),  # Convert ObjectId to string
                email=doc['email'],
                username=doc['username'],
                password_hash=doc['password_hash'],
                is_verified=doc.get('is_verified', False),
                otp_verified=doc.get('otp_verified', False),
                company_id=doc.get('company_id'),
                phone=doc.get('phone'),
                role=doc.get('role', 'user'),
                created_at=_parse_datetime(doc.get('created_at')),
                assigned_companies=doc.get('assigned_companies', [])
            )
            print(f'Successfully loaded user: {user.email} (ID: {user.id})')
            return user
        except Exception as e:
            print(f"Error loading user {user_id}: {e}")
            return None
    else:
        print('MongoDB not available, falling back to JSON users')
        user_data = users.get(user_id) if hasattr(users, 'get') else None
        if user_data:
            return User(
                id=user_id,
                email=user_data['email'],
                username=user_data.get('username', user_data['email'].split('@')[0]),
                password_hash=user_data['password_hash'],
                is_verified=user_data.get('is_verified', False),
                otp_verified=user_data.get('otp_verified', False),
                cart=user_data.get('cart', []),
                reset_token=user_data.get('reset_token'),
                reset_token_expiry=user_data.get('reset_token_expiry'),
                company_id=user_data.get('company_id'),
                phone=user_data.get('phone'),
                role=user_data.get('role', 'user'),
                created_at=_parse_datetime(user_data.get('created_at')),
                assigned_companies=user_data.get('assigned_companies', [])
            )
        return None

@app.route('/cart')
@login_required
@company_required
def cart():
    """Render the cart page with current cart contents and calculated totals.
    
    The Jinja template expects a cart object with products list and calculated totals.
    """
    try:
        # Get the current cart
        cart_data = get_user_cart()
        if not isinstance(cart_data, dict):
            cart_data = {"products": []}

        # Verify company selection even after decorator guard (belt-and-suspenders)
        selected_company = session.get('selected_company', {}) or {}
        active_company_id = selected_company.get('id')
        if not active_company_id:
            return redirect(url_for('index'))

        # Ensure products list exists
        cart_data.setdefault("products", [])

        # Calculate cart totals using the actual total field from each product
        total = 0
        if cart_data.get('products'):
            total = sum(
                float(p.get('calculations', {}).get('final_total', 0))
                for p in cart_data['products']
            )
            
            # If no calculations exist, fall back to total field
            if not total:
                total = sum(
                    float(p.get('total', 0))
                    for p in cart_data['products']
                )
            
            # Calculate discount amount if needed
            discount_amount = sum(
                float(p.get('calculations', {}).get('discount_amount', 0))
                for p in cart_data['products']
            )
            
            # Add calculated totals to the cart data
            cart_data['calculations'] = {
                'discount_amount': round(discount_amount, 2),
                'total': round(total, 2)
            }
        
        # Company details must come from an explicit selection (enforced by company_required)
        company_name = (selected_company.get('name') or session.get('company_name') or '')
        company_email = (selected_company.get('email') or session.get('company_email') or '')
            
        # Log the company info for debugging
        app.logger.info(f"Cart - Company: {company_name}, Email: {company_email}")
        return render_template('cart.html',
                           cart=cart_data,
                            products=cart_data.get('products', []),
                            company_name=company_name,
                            company_email=company_email,
                            # Calculate GST rates for each product
                            products_with_gst=[
                                {**p, 'gst_percent': 18.0}
                                for p in cart_data.get('products', [])
                            ],
                            calculations=cart_data.get('calculations', {
                                'subtotal': 0,
                                'gst_percent': 0,  # Will be calculated per product
                                'gst_amount': 0,
                                'total': 0
                            }))
        
    except Exception as e:
        error_msg = f"Error in cart route: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        app.logger.error(error_msg)
        # Return empty cart with error message
        return render_template('cart.html', 
                           cart={"products": []}, 
                           error=str(e),
                           company_name='',
                           company_email='',
                           products_with_gst=[],
                           calculations={
                               'subtotal': 0,
                               'gst_percent': 0,
                               'gst_amount': 0,
                               'total': 0
                           })

@app.route('/clear_cart', methods=['POST'])
@login_required
def clear_cart():
    """Clear current user's cart"""
    try:
        if current_user.is_authenticated:
            # For logged-in users, clear the cart from the database
            if USE_MONGO and MONGO_AVAILABLE and mongo_db is not None:
                mongo_db.carts.update_one(
                    {'user_id': str(current_user.id)},
                    {'$set': {'products': []}},
                    upsert=True
                )
            else:
                # Fallback to session for non-MongoDB
                session['cart'] = {'products': []}
        else:
            # For non-logged-in users, clear the session cart
            session['cart'] = {'products': []}
        
        session.modified = True
        return jsonify({'success': True, 'message': 'Cart cleared successfully'})
    except Exception as e:
        print(f"Error clearing cart: {e}")
        return jsonify({'error': 'Failed to clear cart', 'message': str(e)})

@app.route('/add_to_cart', methods=['POST'])
@login_required
@company_required
def add_to_cart():
    try:
        # Get request data and validate
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400

        def to_bool(value):
            if isinstance(value, bool):
                return value
            if value is None:
                return False
            return str(value).strip().lower() in ('1', 'true', 'yes', 'y', 'on')

        # Validate required fields for blanket
        required_fields = ['type', 'name', 'machine', 'length', 'width', 'unit', 'quantity', 'base_price', 'bar_price', 'gst_percent']
        if data.get('type') == 'blanket' and not all(data.get(field) is not None for field in required_fields):
            return jsonify({
                'success': False,
                'error': f'Missing required fields: {required_fields}'
            }), 400

        # Calculate prices based on product type
        if data.get('type') == 'blanket':
            # Get all required fields with proper defaults
            base_price = float(data.get('base_price', 0))
            bar_price = float(data.get('bar_price', 0))
            quantity = int(data.get('quantity', 1))
            discount_percent = float(data.get('discount_percent', 0))
            blanket_name = data.get('name') or data.get('blanket_name')
            if is_restricted_blanket(blanket_name):
                cap = get_restricted_discount_cap(current_user)
                if discount_percent > cap:
                    discount_percent = cap
            gst_percent = float(data.get('gst_percent', 18))
            
            # Calculate prices
            unit_price = base_price + bar_price
            subtotal = unit_price * quantity
            discount_amount = subtotal * (discount_percent / 100)
            discounted_subtotal = subtotal - discount_amount
            gst_amount = (discounted_subtotal * gst_percent) / 100
            final_total = discounted_subtotal + gst_amount
            
            # Get dimensions and other details
            length = float(data.get('length', 0))
            width = float(data.get('width', 0))
            unit = data.get('unit', 'mm')
            
            # Convert area to square meters if needed
            area_sq_m = length * width
            if unit == 'mm':
                area_sq_m = (length / 1000) * (width / 1000)
            elif unit == 'in':
                area_sq_m = (length * 0.0254) * (width * 0.0254)
            
            # Create product with all details
            import uuid
            product = {
                'id': str(uuid.uuid4()),  # Add unique ID
                'type': 'blanket',
                'name': data.get('name', 'Custom Blanket'),
                'machine': data.get('machine', 'Unknown Machine'),
                'thickness': data.get('thickness', ''),
                'length': length,
                'width': width,
                'unit': unit,
                'bar_type': data.get('bar_type', 'None'),
                'bar_price': bar_price,
                'quantity': quantity,
                'base_price': base_price,
                'discount_percent': discount_percent,
                'gst_percent': gst_percent,
                'unit_price': round(unit_price, 2),
                'total_price': round(final_total, 2),
                'calculations': {
                    'areaSqM': round(area_sq_m, 4),
                    'ratePerSqMt': round(base_price / area_sq_m, 2) if area_sq_m > 0 else 0,
                    'basePrice': round(base_price, 2),
                    'pricePerUnit': round(unit_price, 2),
                    'subtotal': round(subtotal, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_price': round(final_total, 2)
                },
                'added_at': datetime.utcnow().isoformat()
            }
        else:
            # Handle other product types (mpack, etc.)
            import uuid

            standard_length = to_float(data.get('standard_length_mm'))
            standard_width = to_float(data.get('standard_width_mm'))
            standard_area = to_float(data.get('standard_area_sqm'))
            if standard_area is None and standard_length is not None and standard_width is not None:
                standard_area = (standard_length * standard_width) / 1_000_000

            custom_length = to_float(data.get('custom_length_mm'))
            custom_width = to_float(data.get('custom_width_mm'))
            custom_area = to_float(data.get('custom_area_sqm'))
            if custom_area is None and custom_length is not None and custom_width is not None:
                custom_area = (custom_length * custom_width) / 1_000_000

            display_length = to_float(data.get('display_length_mm'))
            display_width = to_float(data.get('display_width_mm'))
            cut_to_custom = to_bool(data.get('cut_to_custom_size'))

            if display_length is None:
                display_length = custom_length if cut_to_custom and custom_length is not None else standard_length
            if display_width is None:
                display_width = custom_width if cut_to_custom and custom_width is not None else standard_width

            standard_size_label = data.get('standard_size_label') or data.get('size', '')
            custom_size_label = data.get('custom_size_label') or ''
            display_size_label = data.get('display_size_label') or (
                custom_size_label if cut_to_custom and custom_size_label else standard_size_label
            )

            width_value = display_width if display_width is not None else (standard_width if standard_width is not None else 0)
            height_value = display_length if display_length is not None else (standard_length if standard_length is not None else 0)

            product = {
                'id': str(uuid.uuid4()),  # Add unique ID
                'type': data.get('type'),
                'name': data.get('name'),
                'unit_price': float(data.get('unit_price', 0)),
                'quantity': int(data.get('quantity', 1)),
                'discount_percent': float(data.get('discount_percent', 0)),
                'gst_percent': float(data.get('gst_percent', 18)),
                'machine': data.get('machine', ''),
                'thickness': data.get('thickness', ''),
                'size': display_size_label or data.get('size', ''),
                'underpacking_type': data.get('underpacking_type', ''),
                'added_at': datetime.utcnow().isoformat(),
                'cut_to_custom_size': cut_to_custom,
                'custom_length_mm': custom_length,
                'custom_width_mm': custom_width,
                'custom_area_sqm': custom_area,
                'standard_length_mm': standard_length,
                'standard_width_mm': standard_width,
                'standard_area_sqm': standard_area,
                'standard_size_label': standard_size_label,
                'custom_size_label': custom_size_label,
                'display_size_label': display_size_label,
                'display_length_mm': display_length,
                'display_width_mm': display_width,
                'width': width_value,
                'height': height_value
            }

            product_type = product.get('type')

            if product_type == 'rule':
                patch_rule_metadata(product, data)

            if product_type == 'litho_perforation':
                quantity = int(product.get('quantity', 1) or 1)
                unit_price = float(product.get('unit_price', 0) or 0)
                discount_percent = float(product.get('discount_percent', 0) or 0)
                gst_percent = float(product.get('gst_percent', 18) or 18)
                subtotal = unit_price * quantity
                discount_amount = subtotal * (discount_percent / 100)
                discounted_subtotal = subtotal - discount_amount
                gst_amount = discounted_subtotal * (gst_percent / 100)
                final_total = discounted_subtotal + gst_amount

                product.update({
                    'tpi': data.get('tpi'),
                    'brand': data.get('brand', ''),
                    'brand_id': data.get('brand_id', ''),
                    'rule_type': data.get('rule_type', ''),
                    'rule_type_id': data.get('rule_type_id', ''),
                    'product_code': data.get('product_code', ''),
                    'packets': quantity,
                    'subtotal': round(subtotal, 2),
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_amount': round(gst_amount, 2),
                    'total_price': round(final_total, 2),
                    'total': round(final_total, 2)
                })

                product['calculations'] = {
                    'unit_price': round(unit_price, 2),
                    'quantity': quantity,
                    'subtotal': round(subtotal, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(final_total, 2)
                }

            if product_type == 'autowash_cloth':
                quantity = int(product.get('quantity', 1) or 1)
                unit_price = float(product.get('unit_price', 0) or 0)
                discount_percent = float(product.get('discount_percent', 0) or 0)
                gst_percent = float(product.get('gst_percent', 18) or 18)
                subtotal = unit_price * quantity
                discount_amount = subtotal * (discount_percent / 100)
                discounted_subtotal = subtotal - discount_amount
                gst_amount = discounted_subtotal * (gst_percent / 100)
                final_total = discounted_subtotal + gst_amount

                product.update({
                    'category': data.get('category', ''),
                    'product_id': data.get('product_id', ''),
                    'format_label': data.get('format_label', ''),
                    'autowash_type': data.get('autowash_type', ''),
                    'width_mm': data.get('width_mm', ''),
                    'length_m': data.get('length_m', ''),
                    'packaging': data.get('packaging', ''),
                    'pcs_per_box': data.get('pcs_per_box', ''),
                    'unit': data.get('unit', ''),
                    'subtotal': round(subtotal, 2),
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_amount': round(gst_amount, 2),
                    'total_price': round(final_total, 2),
                    'total': round(final_total, 2)
                })

                product['calculations'] = {
                    'unit_price': round(unit_price, 2),
                    'quantity': quantity,
                    'subtotal': round(subtotal, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(final_total, 2)
                }

            if product_type in ('chemical', 'maintenance'):
                pack_size = to_float(data.get('pack_size_litre'))
                quantity_litre = to_float(data.get('quantity_litre'))
                total_litre = to_float(data.get('total_litre'))
                surplus_litre = to_float(data.get('surplus_litre'))

                packs_needed_raw = data.get('packs_needed', product.get('quantity'))
                try:
                    packs_needed = int(packs_needed_raw) if packs_needed_raw is not None else product.get('quantity')
                except (TypeError, ValueError):
                    packs_needed = product.get('quantity')

                subtotal = product['unit_price'] * product.get('quantity', 1)
                discount_percent = product.get('discount_percent', 0.0)
                discount_amount = subtotal * (discount_percent / 100)
                discounted_subtotal = subtotal - discount_amount
                gst_percent = product.get('gst_percent', 18.0)
                gst_amount = discounted_subtotal * (gst_percent / 100)
                final_total = discounted_subtotal + gst_amount

                product.update({
                    'category': data.get('category', ''),
                    'product_id': data.get('product_id'),
                    'format_id': data.get('format_id'),
                    'format_label': data.get('format_label'),
                    'pack_size_litre': pack_size,
                    'quantity_litre': quantity_litre,
                    'packs_needed': packs_needed,
                    'total_litre': total_litre,
                    'surplus_litre': surplus_litre,
                    'pricing_tier': data.get('pricing_tier', 'standard'),
                    'subtotal': round(subtotal, 2),
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_amount': round(gst_amount, 2),
                    'total': round(final_total, 2),
                    'total_price': round(final_total, 2)
                })

                product['calculations'] = {
                    'unit_price': round(product['unit_price'], 2),
                    'quantity': product.get('quantity', 1),
                    'subtotal': round(subtotal, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'discounted_subtotal': round(discounted_subtotal, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(final_total, 2)
                }

                # Ensure pack-specific metrics are reflected in quantity fields
                product['quantity'] = packs_needed

            # Calculate prices for other product types if needed
            if product_type == 'mpack':
                unit_price = float(product.get('unit_price', 0) or 0)
                quantity = int(product.get('quantity', 1) or 1)
                discount_percent = float(product.get('discount_percent', 0) or 0)
                gst_percent = float(product.get('gst_percent', 18) or 18)

                subtotal = unit_price * quantity
                discount_amount = subtotal * (discount_percent / 100)
                discounted_subtotal = subtotal - discount_amount
                gst_amount = discounted_subtotal * (gst_percent / 100)
                final_total = discounted_subtotal + gst_amount

                product['unit_price'] = round(unit_price, 2)
                product['discount_amount'] = round(discount_amount, 2)
                product['discounted_subtotal'] = round(discounted_subtotal, 2)
                product['gst_amount'] = round(gst_amount, 2)
                product['total_price'] = round(final_total, 2)
                product['total'] = product['total_price']

                product['calculations'] = {
                    'unit_price': product['unit_price'],
                    'quantity': quantity,
                    'subtotal': round(subtotal, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': product['discount_amount'],
                    'discounted_subtotal': product['discounted_subtotal'],
                    'gst_percent': gst_percent,
                    'gst_amount': product['gst_amount'],
                    'final_total': product['total_price'],
                    'machine': product.get('machine', ''),
                    'thickness': product.get('thickness', ''),
                    'size': product.get('size', ''),
                    'standard_size_label': standard_size_label,
                    'custom_size_label': custom_size_label,
                    'display_size_label': display_size_label,
                    'cut_to_custom_size': cut_to_custom
                }
            elif product_type == 'rule':
                patch_rule_metadata(product, data)
                recalc_rule_pricing(product)
        
        # Get existing cart or create new one
        try:
            cart = get_user_cart()
            if not isinstance(cart, dict):
                cart = {'products': []}
            if 'products' not in cart:
                cart['products'] = []
            
            # Check if this is an update to an existing item
            item_id = data.get('item_id')
            if item_id:
                # Find and update the existing item
                item_updated = False
                for idx, item in enumerate(cart['products']):
                    if str(item.get('_id', '')) == str(item_id) or str(item.get('id', '')) == str(item_id):
                        # Update all fields from the new product data
                        cart['products'][idx].update(product)
                        item_updated = True
                        break
                
                if not item_updated:
                    return jsonify({
                        'success': False,
                        'error': 'Item not found in cart',
                        'message': 'The item you are trying to update was not found in your cart.'
                    }), 404
            else:
                # Check for duplicate products with same dimensions if force_add is not True
                if not data.get('force_add'):
                    duplicate_index = -1
                    product_type = product.get('type')
                    
                    if product_type == 'blanket':
                        for idx, item in enumerate(cart['products']):
                            if (item.get('type') == 'blanket' and 
                                abs(float(item.get('length', 0)) - float(product.get('length', 0))) < 0.01 and 
                                abs(float(item.get('width', 0)) - float(product.get('width', 0))) < 0.01 and 
                                item.get('thickness') == product.get('thickness') and
                                item.get('bar_type') == product.get('bar_type')):
                                duplicate_index = idx
                                break
                    
                    # Check for duplicate MPacks with same specifications
                    elif product_type == 'mpack':
                        for idx, item in enumerate(cart['products']):
                            if (item.get('type') == 'mpack' and 
                                item.get('machine') == product.get('machine') and
                                item.get('thickness') == product.get('thickness') and
                                item.get('size') == product.get('size') and
                                item.get('underpacking_type') == product.get('underpacking_type')):
                                duplicate_index = idx
                                break
                    
                    if duplicate_index >= 0:
                        # Return info about duplicate product
                        return jsonify({
                            'success': False,
                            'is_duplicate': True,
                            'duplicate_index': duplicate_index,
                            'message': 'A product with the same dimensions already exists in your cart.'
                        })
                
                # If no duplicate found and not an update, add the product to cart
                cart['products'].append(product)
            
            # Save updated cart
            save_user_cart(cart)
            
            # Get updated cart count
            updated_cart = get_user_cart()
            cart_count = len(updated_cart.get('products', [])) if updated_cart and isinstance(updated_cart, dict) else 0
            
            return jsonify({
                'success': True,
                'is_duplicate': False,
                'message': 'Product added to cart successfully',
                'cart_count': cart_count
            })
        except Exception as e:
            app.logger.error(f"Error saving cart: {str(e)}")
            return jsonify({
                'success': False,
                'error': f'Failed to save cart: {str(e)}'
            }), 500
    except Exception as e:
        app.logger.error(f"Error adding to cart: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Internal server error: {str(e)}'
        }), 500
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/get_cart')
@login_required
def get_cart():
    """Return the current user's cart as JSON."""
    try:
        cart = get_user_cart()
        return jsonify(cart)
    except Exception as e:
        print(f"Error get_cart: {e}")
        return jsonify({'error': 'Failed to get cart', 'products': []}), 500

@app.route('/remove_from_cart', methods=['POST'])
@login_required
def remove_from_cart():
    """Remove the product with the given ID from the user's cart."""
    data = request.get_json() or {}
    item_id = data.get('item_id')
    
    if not item_id:
        return jsonify({'error': 'Missing item_id'}), 400

    try:
        cart = get_user_cart()
        products = cart.get('products', [])
        
        # Find the item by ID
        initial_count = len(products)
        products = [p for p in products if p.get('id') != item_id]
        
        if len(products) < initial_count:
            # Item was found and removed
            save_user_cart({'products': products})
            return jsonify({
                'success': True,
                'cart_count': len(products),
                'message': 'Item removed from cart'
            })
            
        return jsonify({
            'success': False,
            'error': 'Item not found in cart',
            'cart_count': len(products)
        }), 404
        
    except Exception as e:
        app.logger.error(f'Error in remove_from_cart: {e}')
        return jsonify({
            'success': False,
            'error': 'Failed to remove item from cart',
            'details': str(e)
        }), 500


@app.route('/update_cart_item', methods=['POST'])
@login_required
def update_cart_item():
    """Update an existing cart item with new data."""
    if not current_user.is_authenticated:
        return jsonify({
            'success': False,
            'error': 'User not authenticated',
            'redirect': url_for('login')
        }), 401

    data = request.get_json()
    item_id = data.get('item_id')
    
    if not item_id:
        return jsonify({
            'success': False,
            'error': 'Item ID is required',
            'message': 'Please provide a valid item ID'
        }), 400
    
    # Get the current cart
    cart = get_user_cart()
    products = cart.get('products', [])
    
    # Find the item to update
    item_index = next((i for i, item in enumerate(products) 
                      if str(item.get('id')) == str(item_id) or 
                         str(item.get('_id')) == str(item_id)), None)
    
    if item_index is None:
        return jsonify({
            'success': False,
            'error': 'Item not found in cart',
            'message': 'The item you are trying to update was not found in your cart'
        }), 404
    
    # Update the item with new data
    item = products[item_index]
    
    # Update fields from the form data
    for key in ['quantity', 'length', 'width', 'thickness', 'size', 'machine', 'bar_type',
               'discount_percent', 'gst_percent', 'unit_price', 'base_price', 'bar_price', 'name', 'type',
               'unit', 'blanket_name', 'underpacking_type', 'category', 'format_label',
               'custom_length_mm', 'custom_width_mm', 'custom_area_sqm',
               'standard_length_mm', 'standard_width_mm', 'standard_area_sqm',
               'display_length_mm', 'display_width_mm', 'display_size_label',
               'standard_size_label', 'custom_size_label', 'cut_to_custom_size',
               'along_mm', 'across_mm', 'tpi', 'brand', 'brand_id',
               'rule_type', 'rule_type_id', 'product_code', 'packets']:
        if key in data:
            item[key] = data[key]

    if 'calculations' in data and isinstance(data.get('calculations'), dict):
        item['calculations'] = data['calculations']
    
    # Recalculate any calculated fields
    if 'quantity' in data or 'unit_price' in data or 'discount_percent' in data or 'gst_percent' in data:
        quantity = item.get('quantity', 1)
        discount_percent = float(item.get('discount_percent', 0))
        gst_percent = float(item.get('gst_percent', 18))  # Default to 18% GST if not specified
        
        # Handle blanket vs other product types differently
        if item.get('type') == 'blanket':
            # For blankets: keep base_price and bar_price separate for display
            base_price = float(item.get('base_price', 0)) or float(item.get('unit_price', 0))
            bar_price = float(item.get('bar_price', 0))
            
            # Calculate unit price (base + bar)
            unit_price = base_price + bar_price
            
            # Calculate subtotal (unit_price * quantity)
            subtotal = unit_price * quantity
            
            # Update the stored values
            item['base_price'] = base_price
            item['bar_price'] = bar_price
            item['unit_price'] = unit_price
        else:
            # For other products (mpack, etc.)
            unit_price = float(item.get('unit_price', 0))
            subtotal = unit_price * quantity
        
        # Calculate discount and final amounts
        discount_amount = (subtotal * discount_percent) / 100
        discounted_subtotal = subtotal - discount_amount
        gst_amount = (discounted_subtotal * gst_percent) / 100
        final_total = discounted_subtotal + gst_amount
        
        # Update calculations
        item['calculations'] = {
            'unit_price': unit_price,
            'quantity': quantity,
            'subtotal': subtotal,
            'discount_percent': discount_percent,
            'discount_amount': discount_amount,
            'discounted_subtotal': discounted_subtotal,
            'gst_percent': gst_percent,
            'gst_amount': gst_amount,
            'final_total': final_total
        }
        
        # Update the item's total_price field
        item['total_price'] = final_total
    
    # Save the updated cart
    cart['products'][item_index] = item
    save_user_cart(cart)
    
    return jsonify({
        'success': True,
        'message': 'Item updated successfully',
        'cart': cart
    })

@app.route('/update_cart_quantity', methods=['POST'])
@login_required
def update_cart_quantity():
    """Update the quantity of a product in the user's cart."""
    def to_float(value):
        """Safely convert the provided value to float, returning None on failure."""
        try:
            if value is None or value == '':
                return None
            return float(value)
        except (TypeError, ValueError):
            return None
            
    if not current_user.is_authenticated:
        return jsonify({
            'success': False,
            'error': 'User not authenticated',
            'redirect': url_for('login')
        }), 401

    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
            
        product_type = data.get('type', 'mpack')
        item_id = data.get('item_id')
        if product_type in ('chemical', 'maintenance'):
            quantity_value = to_float(data.get('quantity_litre') or data.get('quantity') or 0)
            quantity = quantity_value if quantity_value is not None else 0
        else:
            quantity = int(data.get('quantity', 1))

        # Validate quantity
        if product_type in ('chemical', 'maintenance'):
            if quantity <= 0:
                return jsonify({
                    'success': False,
                    'message': 'Quantity must be greater than 0 for chemical items'
                }), 400
        elif quantity < 1:
            return jsonify({
                'success': False,
                'message': 'Quantity must be at least 1'
            }), 400
            
        # Get current cart
        cart = get_user_cart()
        products = cart.get('products', [])
        
        # Find the item by ID
        item_updated = False
        updated_item = None
        
        for item in products:
            if str(item.get('id')) == str(item_id):
                # Update the quantity
                item['quantity'] = quantity
                if product_type in ('chemical', 'maintenance'):
                    item['quantity_litre'] = quantity

                # Recalculate prices if needed (for blankets)
                if item.get('type') == 'blanket':
                    # Recalculate blanket prices
                    base_price = item.get('base_price', 0)
                    bar_price = item.get('bar_price', 0)
                    discount_percent = item.get('discount_percent', 0)
                    gst_percent = item.get('gst_percent', 18)
                    
                    # Recalculate all values
                    price_per_unit = base_price + bar_price
                    subtotal = price_per_unit * quantity
                    discount_amount = subtotal * (discount_percent / 100)
                    discounted_subtotal = subtotal - discount_amount
                    gst_amount = (discounted_subtotal * gst_percent) / 100
                    final_total = discounted_subtotal + gst_amount
                    
                    # Update all price fields
                    item.update({
                        'unit_price': round(price_per_unit, 2),
                        'total_price': round(final_total, 2),
                        'calculations': {
                            **item.get('calculations', {}),
                            'subtotal': round(subtotal, 2),
                            'discount_amount': round(discount_amount, 2),
                            'discounted_subtotal': round(discounted_subtotal, 2),
                            'gst_amount': round(gst_amount, 2),
                            'final_price': round(final_total, 2)
                        }
                    })
                
                if item.get('type') in ('chemical', 'maintenance'):
                    price_per_litre = item.get('price_per_litre') or item.get('unit_price') or 0
                    discount_percent = item.get('discount_percent', 0)
                    gst_percent = item.get('gst_percent', 18)

                    subtotal = price_per_litre * quantity
                    discount_amount = subtotal * (discount_percent / 100)
                    discounted_subtotal = subtotal - discount_amount
                    gst_amount = (discounted_subtotal * gst_percent / 100)
                    final_total = discounted_subtotal + gst_amount

                    item['total_price'] = round(final_total, 2)
                    item['calculations'] = {
                        **item.get('calculations', {}),
                        'unit_price': round(price_per_litre, 2),
                        'quantity': quantity,
                        'subtotal': round(subtotal, 2),
                        'discount_percent': discount_percent,
                        'discount_amount': round(discount_amount, 2),
                        'discounted_subtotal': round(discounted_subtotal, 2),
                        'gst_percent': gst_percent,
                        'gst_amount': round(gst_amount, 2),
                        'final_total': round(final_total, 2)
                    }

                updated_item = item
                item_updated = True
                break
        
        if item_updated:
            # Save the updated cart
            save_user_cart({'products': products})
            
            return jsonify({
                'success': True,
                'message': 'Cart quantity updated',
                'cart_count': sum(1 for _ in products),
                'updated_item': updated_item
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Item not found in cart',
                'cart_count': len(products)
            }), 404
    except Exception as e:
        app.logger.error(f'Error updating cart quantity: {str(e)}')
        return jsonify({
            'success': False,
            'message': 'An error occurred while updating the cart quantity',
            'error': str(e)
        }), 500


@app.route('/update_cart_discount', methods=['POST'])
@login_required
def update_cart_discount():
    """Update the discount percentage of a product in the user's cart."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'message': 'No data provided'
            }), 400
            
        item_id = data.get('item_id')
        discount_percent = float(data.get('discount_percent', 0))
        
        # Validate discount percentage
        if discount_percent < 0 or discount_percent > 100:
            return jsonify({
                'success': False,
                'message': 'Discount percentage must be between 0 and 100'
            }), 400
            
        if not item_id:
            return jsonify({
                'success': False,
                'message': 'Item ID is required'
            }), 400
        
        # Get current cart
        cart = get_user_cart()
        products = cart.get('products', [])
        
        # Find the item by ID
        item_updated = False
        updated_item = None
        
        for item in products:
            if str(item.get('id')) == str(item_id):
                # Update the discount percentage
                item['discount_percent'] = discount_percent
                
                # Recalculate prices based on product type
                if item.get('type') == 'blanket':
                    blanket_name = item.get('blanket_name') or item.get('name')
                    if is_restricted_blanket(blanket_name):
                        cap = get_restricted_discount_cap(current_user)
                        if discount_percent > cap:
                            discount_percent = cap
                    base_price = item.get('base_price', 0)
                    bar_price = item.get('bar_price', 0)
                    quantity = item.get('quantity', 1)
                    gst_percent = item.get('gst_percent', 18)
                    
                    price_per_unit = base_price + bar_price
                    subtotal = price_per_unit * quantity
                    discount_amount = subtotal * (discount_percent / 100)
                    discounted_subtotal = subtotal - discount_amount
                    gst_amount = (discounted_subtotal * gst_percent) / 100
                    final_total = discounted_subtotal + gst_amount
                    
                    # Update all price fields
                    item.update({
                        'unit_price': round(price_per_unit, 2),
                        'total_price': round(final_total, 2),
                        'calculations': {
                            **item.get('calculations', {}),
                            'subtotal': round(subtotal, 2),
                            'discount_amount': round(discount_amount, 2),
                            'discounted_subtotal': round(discounted_subtotal, 2),
                            'gst_amount': round(gst_amount, 2),
                            'final_price': round(final_total, 2)
                        }
                    })
                else:
                    # For mpacks and other product types
                    unit_price = item.get('unit_price', 0)
                    quantity = item.get('quantity', 1)
                    gst_percent = item.get('gst_percent', 18)
                    
                    subtotal = unit_price * quantity
                    discount_amount = subtotal * (discount_percent / 100)
                    discounted_subtotal = subtotal - discount_amount
                    gst_amount = (discounted_subtotal * gst_percent) / 100
                    final_total = discounted_subtotal + gst_amount
                    
                    # Update all price fields
                    item.update({
                        'total_price': round(final_total, 2),
                        'calculations': {
                            **item.get('calculations', {}),
                            'subtotal': round(subtotal, 2),
                            'discount_amount': round(discount_amount, 2),
                            'discounted_subtotal': round(discounted_subtotal, 2),
                            'gst_amount': round(gst_amount, 2),
                            'final_price': round(final_total, 2)
                        }
                    })
                
                updated_item = item
                item_updated = True
                break
        
        if item_updated:
            # Save the updated cart
            save_user_cart({'products': products})
            
            return jsonify({
                'success': True,
                'message': 'Cart discount updated',
                'cart_count': len(products),
                'updated_item': updated_item,
                'applied_discount_percent': discount_percent
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Item not found in cart',
                'cart_count': len(products)
            }), 404
    except Exception as e:
        app.logger.error(f'Error updating cart discount: {str(e)}')
        return jsonify({
            'success': False,
            'message': 'An error occurred while updating the cart discount',
            'error': str(e)
        }), 500


@app.route('/get_cart_count')
def get_cart_count():
    """Return the number of products currently in the user's cart."""
    try:
        if not current_user.is_authenticated:
            return jsonify({'count': 0})
            
        cart = get_user_cart()
        return jsonify({'count': len(cart.get('products', []))})
    except Exception as e:
        print(f"Error in get_cart_count: {e}")
        return jsonify({'count': 0})

def load_companies_data():
    """Load companies data from MongoDB or fall back to JSON file."""
    global mongo_db, USE_MONGO
    
    try:
        # Log MongoDB status
        app.logger.info(f"Loading companies - MongoDB status: Available={MONGO_AVAILABLE}, Using={USE_MONGO}, Connected={'Yes' if mongo_db is not None else 'No'}")
        
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                # Test the connection first
                mongo_db.command('ping')
                
                # Get companies from MongoDB with only the fields we need
                projection = {
                    '_id': 1,
                    'Company Name': 1,
                    'EmailID': 1,
                    'Phone': 1,
                    'Billing Attention': 1,
                    'Billing Address': 1,
                    'Billing Street': 1,
                    'Billing City': 1,
                    'Billing State': 1,
                    'Billing Postal Code': 1,
                    'Billing Phone': 1,
                    'created_at': 1,
                    'Created At': 1,
                    'name': 1,
                    'email': 1,
                    'address': 1,
                    'Address': 1,
                    'assigned_to': 1,
                    'created_by': 1
                }
                
                # Find all companies and sort by name for consistent ordering
                companies_cursor = mongo_db.companies.find({}, projection).sort('Company Name', 1)
                companies = list(companies_cursor)
                
                mapped_companies = []
                for company in companies:
                    try:
                        company_id = str(company.pop('_id'))
                        
                        # Get company data (we only store one set of fields now)
                        name = company.get('Company Name') or company.get('name')
                        email = company.get('EmailID') or company.get('email', '')
                        
                        # Skip if we don't have a valid name
                        if not name:
                            app.logger.warning(f"Skipping company with missing name: {company_id}")
                            continue
                        
                        # Ensure email is a string and properly formatted
                        email = str(email).strip() if email else ''

                        gst_registered_raw = company.get('GST Registered')
                        if gst_registered_raw is None:
                            gst_registered_raw = company.get('gst_registered')
                        if isinstance(gst_registered_raw, str):
                            gst_registered = gst_registered_raw.strip().lower() in {'true', '1', 'yes', 'y'}
                        else:
                            gst_registered = bool(gst_registered_raw)

                        gst_number_raw = company.get('GST Number') or company.get('gst_number') or ''
                        gst_number = str(gst_number_raw).strip().upper()
                        
                        mapped_companies.append({
                            'id': company_id,
                            'Company Name': name,
                            'EmailID': email,
                            'name': name,
                            'email': email,
                            'Phone': company.get('Phone'),
                            'phone': company.get('Phone'),
                            'Billing Attention': company.get('Billing Attention'),
                            'billing_attention': company.get('Billing Attention'),
                            'Billing Address': company.get('Billing Address'),
                            'billing_address': company.get('Billing Address'),
                            'Billing Street': company.get('Billing Street'),
                            'billing_street': company.get('Billing Street'),
                            'Billing City': company.get('Billing City'),
                            'billing_city': company.get('Billing City'),
                            'Billing State': company.get('Billing State'),
                            'billing_state': company.get('Billing State'),
                            'Billing Postal Code': company.get('Billing Postal Code'),
                            'billing_postal_code': company.get('Billing Postal Code'),
                            'Billing Phone': company.get('Billing Phone'),
                            'billing_phone': company.get('Billing Phone'),
                            'GST Registered': gst_registered,
                            'gst_registered': gst_registered,
                            'GST Number': gst_number,
                            'gst_number': gst_number,
                            'Address': company.get('Address') or company.get('address', ''),
                            'address': company.get('Address') or company.get('address', ''),
                            'assigned_to': normalize_assigned_companies(company.get('assigned_to', [])),
                            'created_at': company.get('created_at') or company.get('Created At'),
                            'created_by': company.get('created_by')
                        })
                        
                    except Exception as e:
                        app.logger.error(f"Error processing company {company.get('_id')}: {str(e)}")
                        continue
                        
                app.logger.info(f"Successfully loaded {len(mapped_companies)} companies from MongoDB")
                return mapped_companies
                
            except Exception as db_error:
                app.logger.error(f"MongoDB error in load_companies_data: {str(db_error)}")
                # Fall through to JSON fallback
                USE_MONGO = False
                
        # Fall back to JSON file if MongoDB is not available or there was an error
        companies_file = os.path.join(app.root_path, 'static', 'data', 'companies.json')
        app.logger.info(f"Falling back to loading companies from: {companies_file}")
        
        if os.path.exists(companies_file):
            try:
                with open(companies_file, 'r', encoding='utf-8') as f:
                    companies_data = json.load(f)
                    raw_companies = companies_data.get('companies', companies_data if isinstance(companies_data, list) else [])

                    normalized_companies = []
                    for company in raw_companies:
                        try:
                            if isinstance(company, dict):
                                # Determine ID
                                company_id = (
                                    str(company.get('id'))
                                    or str(company.get('_id', {}).get('$oid'))
                                    or str(company.get('_id'))
                                )

                                if not company_id or company_id.lower() == 'none':
                                    continue

                                name = (
                                    company.get('name')
                                    or company.get('Company Name')
                                    or company.get('company_name')
                                )
                                if not name:
                                    app.logger.debug(f"Skipping company {company_id} due to missing name")
                                    continue

                                email = (
                                    company.get('email')
                                    or company.get('EmailID')
                                    or company.get('email_id')
                                    or ''
                                )

                                normalized_companies.append({
                                    'id': company_id,
                                    'Company Name': name,
                                    'EmailID': str(email).strip() if email else '',
                                    'name': name,
                                    'email': str(email).strip() if email else '',
                                    'Phone': company.get('Phone') or company.get('phone'),
                                    'phone': company.get('Phone') or company.get('phone'),
                                    'Billing Attention': company.get('Billing Attention') or company.get('billing_attention'),
                                    'billing_attention': company.get('Billing Attention') or company.get('billing_attention'),
                                    'Billing Address': company.get('Billing Address') or company.get('billing_address') or company.get('address') or company.get('Address', ''),
                                    'billing_address': company.get('Billing Address') or company.get('billing_address') or company.get('address') or company.get('Address', ''),
                                    'Billing Street': company.get('Billing Street') or company.get('billing_street'),
                                    'billing_street': company.get('Billing Street') or company.get('billing_street'),
                                    'Billing City': company.get('Billing City') or company.get('billing_city'),
                                    'billing_city': company.get('Billing City') or company.get('billing_city'),
                                    'Billing State': company.get('Billing State') or company.get('billing_state'),
                                    'billing_state': company.get('Billing State') or company.get('billing_state'),
                                    'Billing Postal Code': company.get('Billing Postal Code') or company.get('billing_postal_code'),
                                    'billing_postal_code': company.get('Billing Postal Code') or company.get('billing_postal_code'),
                                    'Billing Phone': company.get('Billing Phone') or company.get('billing_phone'),
                                    'billing_phone': company.get('Billing Phone') or company.get('billing_phone'),
                                    'Address': company.get('Billing Address') or company.get('billing_address') or company.get('address') or company.get('Address', ''),
                                    'address': company.get('Billing Address') or company.get('billing_address') or company.get('address') or company.get('Address', ''),
                                    'created_at': company.get('created_at'),
                                    'created_by': company.get('created_by'),
                                    'assigned_to': normalize_assigned_companies(company.get('assigned_to', []))
                                })
                        except Exception as norm_error:
                            app.logger.error(f"Failed to normalize company entry: {norm_error}")
                            continue

                    app.logger.info(f"Loaded {len(normalized_companies)} companies from JSON file")
                    return normalized_companies
            except Exception as e:
                app.logger.error(f"Error reading companies JSON file: {str(e)}")
        
        app.logger.warning("No companies data found in MongoDB or JSON file")
        return []
        
    except Exception as e:
        app.logger.error(f"Unexpected error in load_companies_data: {str(e)}", exc_info=True)
        return []

@app.route('/')
@app.route('/index')
@login_required
def index():
    # If privileged and NO force_user flag, send to GM page; if force_user present, stay on user page.
    if has_gm_pricing_access(current_user):
        if 'force_user' in request.args:
            set_pricing_mode('standard')
        else:
            return redirect(url_for('gm_page'))
    try:
        set_pricing_mode('standard')
        reset_company_selection_session()
        companies = load_companies_data()
        
        needs_warning = session.pop('needs_company_warning', False)
        if needs_warning:
            flash('Please select a company first.', 'warning')
            session.modified = True

        if not current_user.is_authenticated:
            companies = []
        else:
            assigned_ids = get_user_assigned_company_ids(current_user)
            if assigned_ids is not None:
                if assigned_ids:
                    assigned_set = {str(cid) for cid in assigned_ids}
                    companies = [company for company in companies if str(company.get('id')) in assigned_set]
                else:
                    companies = []

        # Ensure companies is a list before passing to template
        if not isinstance(companies, list):
            companies = []
            
        return render_template(
            'user/index.html',
            companies=companies,
            gm_mode=False,
            gm_discount=DEFAULT_GM_DISCOUNT,
            gm_discount_steps=GM_DISCOUNT_STEPS
        )
        
    except Exception as e:
        app.logger.error(f"Error in index route: {str(e)}")
        # Return empty companies list on error
        return render_template(
            'user/index.html',
            companies=[],
            gm_mode=False,
            gm_discount=DEFAULT_GM_DISCOUNT,
            gm_discount_steps=GM_DISCOUNT_STEPS
        )


@app.route('/gm-page')
@login_required
def gm_page():
    try:
        if not has_gm_pricing_access(current_user):
            flash('You do not have permission to access GM pricing.', 'danger')
            return redirect(url_for('index'))

        set_pricing_mode('gm')
        companies = load_companies_data()

        if not current_user.is_authenticated:
            companies = []
        else:
            assigned_ids = get_user_assigned_company_ids(current_user)
            if assigned_ids is not None:
                if assigned_ids:
                    assigned_set = {str(cid) for cid in assigned_ids}
                    companies = [company for company in companies if str(company.get('id')) in assigned_set]
                else:
                    companies = []

        if not isinstance(companies, list):
            companies = []

        discount = get_gm_discount_percent(force=True)
        return render_template(
            'user/index.html',
            companies=companies,
            gm_mode=True,
            gm_discount=discount,
            gm_discount_steps=GM_DISCOUNT_STEPS
        )

    except Exception as e:
        app.logger.error(f"Error in gm_page route: {str(e)}", exc_info=True)
        flash('Unable to load GM page right now.', 'danger')
        return redirect(url_for('index'))


@app.route('/gm/discount', methods=['POST'])
@login_required
def update_gm_discount():
    if not has_gm_pricing_access(current_user):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    payload = request.get_json(silent=True) or request.form or {}
    new_percent = sanitize_gm_discount(
        payload.get('discount') or payload.get('percent') or payload.get('value') or DEFAULT_GM_DISCOUNT
    )
    session['gm_discount'] = new_percent
    session.modified = True

    response_data = {'success': True, 'discount': new_percent}
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify(response_data)

    flash(f'GM discount updated to {new_percent}%.', 'success')
    return redirect(request.referrer or url_for('gm_page'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return jsonify({'success': True, 'redirectTo': '/index'}) if request.method == 'POST' else redirect(url_for('index'))
    
    if request.method == 'POST':
        # Handle POST request - this should never happen since we use API route
        return jsonify({'error': 'Use /api/auth/login for POST requests'}), 400
    
    return render_template('login.html')

@app.route('/signup')
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('signup.html')

@app.route('/company-selection', methods=['GET', 'POST'])
@login_required
def company_selection():
    if request.method == 'POST':
        company = request.form.get('company')
        email = request.form.get('email')
        
        if not company or not email:
            flash('Please select a company and enter an email', 'error')
            return redirect(url_for('company_selection'))
        
        # Save company info in session for convenience
        # NOTE: this legacy flow does not provide an id, so it should not count as a selection
        # for company_required-protected pages.
        session['company'] = company
        session['email'] = email
        
        # Redirect to product selection
        return redirect(url_for('product_selection'))
    
    return render_template('company_selection.html')

@app.route('/product_selection', methods=['GET', 'POST'])
@login_required
def product_selection():
    if request.method == 'POST':
        product_type = request.form.get('product_type')
        
        if not product_type:
            flash('Please select a product type', 'error')
            return redirect(url_for('product_selection'))
            
        # Save product type in session
        session['product_type'] = product_type
        
        # Redirect to appropriate product details page
        if product_type == 'blanket':
            return redirect(url_for('blankets'))
        elif product_type == 'mpack':
            return redirect(url_for('mpacks'))
            
    # Check if company is selected
    selected_company = session.get('selected_company')
    if not selected_company:
        return redirect(url_for('company_selection'))
    
    return render_template('product_selection.html')


@app.route('/select_company', methods=['GET', 'POST'])
@login_required
def select_company():
    app.logger.info(f"Select company request: {request.method}")
    
    if request.method == 'POST':
        try:
            # Get form data
            company_id = request.form.get('company_id')
            company_name = request.form.get('company_name')
            company_email = request.form.get('company_email')
            
            app.logger.info(f"Company selection - ID: {company_id}, Name: {company_name}")
            
            if not all([company_id, company_name, company_email]):
                app.logger.warning("Missing company information in form")
                flash('Please select a valid company', 'error')
                return redirect(url_for('company_selection'))
            
            # Update user's company in the database
            if USE_MONGO and MONGO_AVAILABLE:
                result = users_col.update_one(
                    {'_id': current_user.id},
                    {'$set': {
                        'company_id': company_id,
                        'company_name': company_name,
                        'company_email': company_email,
                        'updated_at': datetime.utcnow()
                    }}
                )
                app.logger.info(f"MongoDB update result: {result.matched_count} documents modified")
            else:
                # Fallback to JSON storage
                users = _load_users_json()
                user_id_str = str(current_user.id)
                if user_id_str in users:
                    users[user_id_str]['company_id'] = company_id
                    users[user_id_str]['company_name'] = company_name
                    users[user_id_str]['company_email'] = company_email
                    _save_users_json(users)
            
            # Update session
            session['company_id'] = company_id
            session['company_name'] = company_name
            session['company_email'] = company_email
            session['selected_company'] = {
                'id': company_id,
                'name': company_name,
                'email': company_email
            }
            
            # Ensure session is saved
            session.modified = True
            
            app.logger.info(f"Company selected: {company_name} ({company_id})")
            flash('Company selected successfully!', 'success')
            return redirect(url_for('product_selection'))
            
        except Exception as e:
            app.logger.error(f"Error in select_company: {str(e)}", exc_info=True)
            flash('An error occurred while processing your request. Please try again.', 'error')
            return redirect(url_for('company_selection'))
    
    # For GET requests, just render the template
    return render_template('company_selection.html')

# -------------------- Cart helper wrappers --------------------

def get_companies():
    try:
        # Load companies from static JSON file
        file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
        app.logger.info(f"Loading companies from: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            companies = json.load(f)
            
        # Create a list to store unique companies (by email)
        unique_companies = {}
        
        # Process companies and ensure unique emails
        for i, company in enumerate(companies, 1):
            email = company.get('EmailID', '').strip().lower()
            name = company.get('Company Name', '').strip()
            
            # Skip if email is missing or already processed
            if not email or email in unique_companies:
                continue
                
            # Add to unique companies with a consistent ID based on email hash
            unique_id = hashlib.md5(email.encode('utf-8')).hexdigest()
            unique_companies[email] = {
                'id': unique_id,
                'name': name,
                'email': email
            }
        
        # Convert to list and sort by company name
        result = sorted(unique_companies.values(), key=lambda x: x['name'].lower())
        return result
        
    except Exception as e:
        app.logger.error(f"Error loading companies: {str(e)}")
        return []

# Redirect old forgot-password URL to reset-password
@app.route('/forgot-password')
def forgot_password_redirect():
    return redirect(url_for('reset_password_page'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    return redirect(url_for('reset_password_page'))

# API Routes

# Company Management
@app.route('/api/companies', methods=['GET'])
@app.route('/get_companies', methods=['GET'])  # Add this line to support both endpoints
@login_required
def api_get_companies():
    """Get all companies from the database"""
    try:
        companies = load_companies_data()  # Use the helper function directly
        if not companies:
            return jsonify({'error': 'No companies found'}), 404

        # Convert to list of dicts if it's a dict
        if isinstance(companies, dict):
            companies = [
                normalize_company_record({**v, 'id': k})
                for k, v in companies.items()
                if isinstance(v, dict)
            ]
        elif isinstance(companies, list):
            companies = [normalize_company_record(c) for c in companies if isinstance(c, dict)]

        assigned_ids = get_user_assigned_company_ids(current_user)
        if assigned_ids is not None:
            if assigned_ids:
                assigned_set = {str(cid) for cid in assigned_ids}
                companies = [company for company in companies if str(company.get('id')) in assigned_set]
            else:
                companies = []

        return jsonify(companies)
    except Exception as e:
        app.logger.error(f"Error getting companies: {str(e)}")
        return jsonify({'error': 'Failed to load companies'}), 500

# Machines list endpoint
@app.route('/api/machines', methods=['GET'])
@login_required
def api_get_machines():
    """Return list of machines.
    Primary design: store machines inside a single *master* document that has an
    array field called `machines`.  If such a document doesn’t exist (e.g. data
    migrated differently), fall back to scanning the whole collection and
    returning each document’s id / name pair.  This guarantees the endpoint
    always returns an array of objects like: [{"id": 1, "name": "Heidelberg"}, …]
    """
    if not (MONGO_AVAILABLE and USE_MONGO and mongo_db is not None):
        return jsonify([])

    try:
        # Preferred structure – one master document with `machines` array
        master_doc = mongo_db.machine.find_one({'machines': {'$exists': True}})
        if master_doc and isinstance(master_doc.get('machines'), list):
            return jsonify(master_doc.get('machines', []))

        # Fallback: each machine as its own document
        cursor = mongo_db.machine.find({}, {'_id': 0, 'id': 1, 'name': 1})
        machines = []
        for doc in cursor:
            # Some datasets might store ObjectIds or missing incremental id.
            # Ensure we always provide an `id` (string) and `name`.
            m_id = str(doc.get('id', doc.get('_id')))
            m_name = doc.get('name')
            if m_name:
                machines.append({'id': m_id, 'name': m_name})

        return jsonify(machines)
    except Exception as e:
        app.logger.error(f"Error fetching machines: {str(e)}")
        return jsonify([])

@app.route('/api/session/update', methods=['POST'])
@login_required
def api_update_session():
    """Update session data such as selected_company from the frontend."""
    if not request.is_json:
        return jsonify({'status': 'error', 'message': 'Request must be JSON'}), 400

    data = request.get_json()

    # Update any keys that the frontend sends (e.g., selected_company)
    allowed_keys = {'selected_company', 'company_id', 'company_name', 'company_email'}
    updated_any = False
    for key in allowed_keys:
        if key in data:
            session[key] = data[key]
            updated_any = True

    if updated_any:
        session.modified = True
        return jsonify({'status': 'success', 'message': 'Session updated'}), 200
    else:
        return jsonify({'status': 'error', 'message': 'No valid keys provided'}), 400

# ---------------------- Static JSON Data Endpoints ----------------------

@app.route('/blanket_categories')
@login_required
def api_blanket_categories():
    """Serve blanket categories JSON to frontend."""
    try:
        file_path = os.path.join(app.root_path, 'static', 'products', 'blankets', 'blanket_categories.json')
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except FileNotFoundError:
        app.logger.error("blanket_categories.json not found at %s", file_path)
        return jsonify({'error': 'Blanket categories data not found'}), 404
    except Exception as e:
        app.logger.error("Error reading blanket_categories.json: %s", e)
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/blanket_data')
@login_required
def api_blanket_data():
    """Serve blankets data JSON to frontend."""
    try:
        file_path = os.path.join(app.root_path, 'static', 'products', 'blankets', 'blankets.json')
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except FileNotFoundError:
        app.logger.error("blankets.json not found at %s", file_path)
        return jsonify({'error': 'Blankets data not found'}), 404
    except Exception as e:
        app.logger.error("Error reading blankets.json: %s", e)
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/thickness_data')
@login_required
def api_thickness_data():
    """Serve thickness data JSON to frontend."""
    try:
        # Prefer blankets folder thickness.json, fallback to static/data/thickness.json
        primary_path = os.path.join(app.root_path, 'static', 'products', 'blankets', 'thickness.json')
        fallback_path = os.path.join(app.root_path, 'static', 'data', 'thickness.json')
        file_path = primary_path if os.path.exists(primary_path) else fallback_path
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except FileNotFoundError:
        app.logger.error("thickness.json not found at %s", file_path)
        return jsonify({'error': 'Thickness data not found'}), 404
    except Exception as e:
        app.logger.error("Error reading thickness.json: %s", e)
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/bar_data')
@login_required
def api_bar_data():
    """Serve bar data JSON to frontend."""
    try:
        file_path = os.path.join(app.root_path, 'static', 'products', 'blankets', 'bar.json')
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify(data)
    except FileNotFoundError:
        app.logger.error("bar.json not found at %s", file_path)
        return jsonify({'error': 'Bar data not found'}), 404
    except Exception as e:
        app.logger.error("Error reading bar.json: %s", e)
        return jsonify({'error': 'Internal server error'}), 500

# Company Search Endpoint
@app.route('/api/companies/search', methods=['GET'])
@login_required
def search_companies():
    """Search for companies by name"""
    query = request.args.get('q', '').lower().strip()
    if not query or len(query) < 2:
        return jsonify([])
    
    try:
        # Search in MongoDB if available
        if MONGO_AVAILABLE and USE_MONGO:
            # Search in both name and address fields
            regex_pattern = f'.*{re.escape(query)}.*'
            companies = list(mongo_db.companies.find({
                '$or': [
                    {'name': {'$regex': regex_pattern, '$options': 'i'}},
                    {'address': {'$regex': regex_pattern, '$options': 'i'}}
                ]
            }).limit(10))
            
            # Convert ObjectId to string for JSON serialization
            for company in companies:
                company['id'] = str(company.pop('_id'))
        else:
            # Fallback to JSON file if needed
            companies_file = os.path.join('data', 'companies.json')
            if os.path.exists(companies_file):
                with open(companies_file, 'r') as f:
                    all_companies = json.load(f)
                
                # Simple case-insensitive search
                companies = [
                    {**c, 'id': c['id']} for c in all_companies 
                    if query in c.get('name', '').lower() or 
                       query in c.get('address', '').lower()
                ][:10]
            else:
                companies = []
        
        return jsonify(companies)
    except Exception as e:
        app.logger.error(f"Error searching companies: {str(e)}")
        return jsonify({'error': 'Failed to search companies'}), 500

@app.route('/api/update_company', methods=['POST'])
@login_required
def update_user_company():
    """Update the current user's company"""
    if not request.is_json:
        return jsonify({'status': 'error', 'message': 'Request must be JSON'}), 400
    
    data = request.get_json()
    company_id = data.get('company_id')
    
    if not company_id:
        return jsonify({'status': 'error', 'message': 'Company ID is required'}), 400
    
    try:
        # Update user's company in the database
        if MONGO_AVAILABLE and USE_MONGO and users_col is not None:
            # Ensure company_id is stored in the correct BSON type when possible
            try:
                company_id_casted = ObjectId(company_id)
            except Exception:
                # Keep the original value if it is not a valid ObjectId
                company_id_casted = company_id

            # Update in MongoDB only – do not create new user docs here
            result = users_col.update_one(
                {'_id': current_user.id},
                {'$set': {'company_id': company_id_casted}}
            )
            if result.matched_count == 0:
                # User document missing – create it so that the company can be saved
                user_doc = users_col.find_one({'_id': current_user.id})
                if not user_doc:
                    new_doc = {
                        '_id': current_user.id,
                        'username': getattr(current_user, 'username', str(current_user.id)),
                        'username_lower': (getattr(current_user, 'username', '') or str(current_user.id)).lower(),
                        'email': getattr(current_user, 'email', ''),
                        'company_id': company_id_casted,
                    }
                    try:
                        users_col.insert_one(new_doc)
                    except Exception as dup_err:
                        # Handle duplicate key gracefully
                        try:
                            from pymongo.errors import DuplicateKeyError
                            if isinstance(dup_err, DuplicateKeyError):
                                app.logger.warning("Duplicate key on user insert, falling back to update: %s", dup_err)
                                # Ensure username_lower is unique by appending user id
                                safe_username_lower = f"user_{current_user.id}"
                                users_col.update_one(
                                    {'_id': current_user.id},
                                    {'$set': {
                                        'company_id': company_id_casted,
                                        'username_lower': safe_username_lower
                                    }},
                                    upsert=True
                                )
                            else:
                                app.logger.error("Error inserting user doc: %s", dup_err)
                        except ImportError:
                            app.logger.error("pymongo DuplicateKeyError not available; error: %s", dup_err)
                else:
                    # If a document exists but was not matched (unlikely), update company_id
                    users_col.update_one({'_id': current_user.id}, {'$set': {'company_id': company_id_casted}})
            # No error even if modified_count == 0 (company already set)
        else:
            # Update in JSON file
            users = load_users()
            if str(current_user.id) not in users:
                return jsonify({'status': 'error', 'message': 'User not found'}), 404
                
            users[str(current_user.id)]['company_id'] = company_id
            save_users()
        
        # Update session
        session['company_id'] = company_id
        
        # Get company details for response
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
        
        # Update session with company details
        session['company_name'] = company_name
        session['company_email'] = company_email
        
        return jsonify({
            'status': 'success',
            'message': 'Company updated successfully',
            'company': {
                'id': company_id,
                'name': company_name,
                'email': company_email
            }
        })
        
    except Exception as e:
        app.logger.error(f"Error updating user company: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal server error'}), 500

def load_users():
    """Load users from JSON file"""
    try:
        users_file = os.path.join(os.path.dirname(__file__), 'data', 'users.json')
        if os.path.exists(users_file):
            with open(users_file, 'r') as f:
                return json.load(f)
        return {}
    except Exception as e:
        app.logger.error(f"Error loading users: {str(e)}")
        return {}

def save_users(users):
    """Save users to JSON file"""
    try:
        users_dir = os.path.join(os.path.dirname(__file__), 'data')
        os.makedirs(users_dir, exist_ok=True)
        users_file = os.path.join(users_dir, 'users.json')
        with open(users_file, 'w') as f:
            json.dump(users, f, indent=2)
    except Exception as e:
        app.logger.error(f"Error saving users: {str(e)}")

@app.route('/update_company', methods=['POST'])
@app.route('/api/user/update-company', methods=['POST'])
@login_required
def update_company():
    """Update the current user's company from product pages"""
    if not request.is_json:
        return jsonify({'status': 'error', 'message': 'Request must be JSON'}), 400
    
    data = request.get_json()
    company_id = data.get('company_id')
    company_name = data.get('company_name')
    company_email = data.get('company_email')
    
    if not all([company_id, company_name, company_email]):
        return jsonify({'status': 'error', 'message': 'Company ID, name, and email are required'}), 400
    
    try:
        # Update user's company in the database
        if MONGO_AVAILABLE and USE_MONGO and users_col is not None:
            # Update or create in MongoDB
            # Ensure we target the correct user document and avoid inserting duplicates
            from bson import ObjectId
            try:
                user_filter = {'_id': ObjectId(current_user.id)} if ObjectId.is_valid(str(current_user.id)) else {'_id': current_user.id}
            except Exception:
                user_filter = {'_id': current_user.id}

            # First, check if the user exists and handle the username_lower field
            user = users_col.find_one(user_filter)
            update_data = {
                'company_id': company_id,
                'company_name': company_name,
                'company_email': company_email,
                'updated_at': datetime.utcnow()
            }
            
            # If user exists, update the document
            if user:
                # If username_lower is missing or null, set a default value to avoid index conflicts
                if 'username_lower' not in user or user.get('username_lower') is None:
                    update_data['username_lower'] = str(user.get('username', '')).lower() or f'user_{current_user.id}'.lower()
                
                # Update the document
                result = users_col.update_one(
                    user_filter,
                    {'$set': update_data}
                )
            else:
                # If user doesn't exist, try to find by email
                if hasattr(current_user, 'email') and current_user.email:
                    user = users_col.find_one({'email': current_user.email})
                    if user:
                        # Update the found user
                        if 'username_lower' not in user or user.get('username_lower') is None:
                            update_data['username_lower'] = str(user.get('username', '')).lower() or f'user_{user["_id"]}'.lower()
                        
                        result = users_col.update_one(
                            {'_id': user['_id']},
                            {'$set': update_data}
                        )
                    else:
                        # If no user found by email, create a new one with required fields
                        update_data.update({
                            '_id': current_user.id,
                            'email': current_user.email,
                            'username': getattr(current_user, 'username', f'user_{current_user.id}'),
                            'username_lower': getattr(current_user, 'username', f'user_{current_user.id}').lower(),
                            'created_at': datetime.utcnow()
                        })
                        users_col.insert_one(update_data)
                        result = type('obj', (object,), {'matched_count': 1})  # Mock result object
        else:
            # Update in JSON file
            users = load_users()
            user_id = str(current_user.id)
            if user_id not in users:
                users[user_id] = {}
                
            users[user_id]['company_id'] = company_id
            users[user_id]['company_name'] = company_name
            users[user_id]['company_email'] = company_email
            users[user_id]['updated_at'] = datetime.utcnow().isoformat()
            save_users(users)
        
        # Update session with company information
        session['company_id'] = company_id
        session['company_name'] = company_name
        session['company_email'] = company_email
        session['selected_company'] = {
            'id': company_id,
            'name': company_name,
            'email': company_email
        }
        session.modified = True  # Ensure session is saved
        
        return jsonify({
            'status': 'success',
            'message': 'Company updated successfully',
            'company': {
                'id': company_id,
                'name': company_name,
                'email': company_email
            }
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error updating company: {str(e)}\n{error_trace}")
        
        # Prepare error details for the response
        error_details = {
            'status': 'error',
            'message': 'Failed to update company information.',
            'error': str(e),
            'error_type': type(e).__name__
        }
        
        # Add more context based on error type
        if 'MongoDB' in error_details['error_type'] or 'pymongo' in error_details['error_type']:
            error_details['message'] = 'Database connection error. Please try again later.'
            
        # Log the full error for debugging
        app.logger.error(f"Returning error response: {error_details}")
        
        return jsonify(error_details), 500

# ---------------------------------------------------------------------------
# Company and Machine Creation Routes
# ---------------------------------------------------------------------------

@app.route('/add_company')
@login_required
def add_company():
    """Render form page to add a new company"""
    return render_template('add_company.html')


@app.route('/add_machine')
@login_required
def add_machine():
    """Render form page to add a new machine"""
    return render_template('add_machine.html')


# ----------------------------- API Endpoints ------------------------------

@app.route('/api/add_company', methods=['POST'])
@login_required
def api_add_company():
    """Handle AJAX request to create a new company"""
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Invalid request, JSON expected.'}), 400

    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    phone = data.get('phone', '').strip()
    billing_phone = data.get('billing_phone', '').strip()
    billing_attention = data.get('billing_attention', '').strip()
    billing_address = data.get('billing_address', '').strip()
    billing_street = data.get('billing_street', '').strip()
    billing_city = data.get('billing_city', '').strip()
    billing_state = data.get('billing_state', '').strip()
    billing_postal_code = data.get('billing_postal_code', '').strip()
    gst_registered = bool(data.get('gst_registered', False))
    gst_number = data.get('gst_number', '').strip().upper() if gst_registered else ''
    app.logger.info(f"Received request to add company: {name} <{email}>")

    required_fields = {
        'Name': name,
        'Email': email,
        'Phone': phone,
        'Billing Phone': billing_phone,
        'Billing Attention': billing_attention,
        'Billing Address': billing_address,
        'Billing City': billing_city,
        'Billing State': billing_state,
        'Billing Postal Code': billing_postal_code
    }

    missing = [label for label, value in required_fields.items() if not value]
    if missing:
        app.logger.warning(f"Missing required fields in add_company request: {missing}")
        return jsonify({'success': False, 'message': f"Missing required fields: {', '.join(missing)}."}), 400

    if gst_registered and (not gst_number or not re.match(r'^[0-9A-Z]{15}$', gst_number)):
        app.logger.warning("Invalid GST number provided")
        return jsonify({'success': False, 'message': 'A valid 15 character GSTIN is required when GST registered is Yes.'}), 400

    try:
        # Use the global mongo_db connection
        global mongo_db, USE_MONGO
        
        # Log MongoDB connection status
        app.logger.info(f"MongoDB status - Available: {MONGO_AVAILABLE}, Using: {USE_MONGO}, Connection: {'Yes' if mongo_db is not None else 'No'}")
        
        # Check MongoDB connection
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                # Test the connection
                mongo_db.command('ping')
                app.logger.info("Successfully pinged MongoDB")
            except Exception as e:
                app.logger.error(f"MongoDB ping failed: {str(e)}")
                USE_MONGO = False
                mongo_db = None

        # Check for existing company with same name or email
        if mongo_db is not None:
            try:
                # Check for existing company in MongoDB (case-insensitive)
                existing_company = mongo_db.companies.find_one({
                    '$or': [
                        {'Company Name': {'$regex': f'^{name}$', '$options': 'i'}},
                        {'EmailID': {'$regex': f'^{email}$', '$options': 'i'}}
                    ]
                })
                
                if existing_company:
                    return jsonify({
                        'success': False, 
                        'message': 'A company with this name or email already exists.'
                    }), 400
                    
                # Insert new company with consistent field names (only one set of fields)
                company_data = {
                    'Company Name': name,
                    'EmailID': email,
                    'Phone': phone,
                    'Billing Phone': billing_phone,
                    'Billing Attention': billing_attention,
                    'Billing Address': billing_address,
                    'Billing Street': billing_street,
                    'Billing City': billing_city,
                    'Billing State': billing_state,
                    'Billing Postal Code': billing_postal_code,
                    'GST Registered': gst_registered,
                    'GST Number': gst_number,
                    'created_at': datetime.utcnow(),
                    'created_by': str(current_user.id)
                }
                app.logger.info(f"Inserting company data: {company_data}")
                result = mongo_db.companies.insert_one(company_data)
                company_id = str(result.inserted_id)
                app.logger.info(f"Successfully inserted company into MongoDB with ID: {company_id}")
                
            except Exception as db_error:
                app.logger.error(f"Database error in api_add_company: {str(db_error)}", exc_info=True)
                # Fall through to JSON fallback
                app.logger.info("Falling back to JSON storage due to database error")
                mongo_db = None  # Force fallback to JSON
                raise db_error
        else:
            # JSON fallback implementation
            companies_file = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
            os.makedirs(os.path.dirname(companies_file), exist_ok=True)
            
            # Load existing companies
            companies = []
            if os.path.exists(companies_file):
                with open(companies_file, 'r', encoding='utf-8') as f:
                    companies = json.load(f) or []
            
            # Check for duplicates
            if any(company.get('Company Name') == name or company.get('EmailID') == email 
                  for company in companies):
                return jsonify({
                    'success': False, 
                    'message': 'A company with this name or email already exists.'
                }), 400
            
            # Add new company
            company_id = str(len(companies) + 1)
            companies.append({
                'id': company_id,
                'Company Name': name,
                'EmailID': email,
                'Phone': phone,
                'Billing Phone': billing_phone,
                'Billing Attention': billing_attention,
                'Billing Address': billing_address,
                'Billing Street': billing_street,
                'Billing City': billing_city,
                'Billing State': billing_state,
                'Billing Postal Code': billing_postal_code,
                'GST Registered': gst_registered,
                'GST Number': gst_number,
                'created_at': datetime.utcnow().isoformat(),
                'created_by': str(current_user.id)
            })
            
            # Save back to file
            with open(companies_file, 'w', encoding='utf-8') as f:
                json.dump(companies, f, ensure_ascii=False, indent=2)

        # Log the successful addition
        app.logger.info(f"Company added successfully - Name: {name}, Email: {email}")
        
        # Try to send notification email (non-blocking)
        try:
            user_identity = getattr(current_user, 'email', getattr(current_user, 'username', 'Unknown User'))
            email_sent = send_alert_email(
                subject='Database Update: New Company Added',
                body=f"{user_identity} added a new company ({name}, {email}) on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            )
        except Exception as email_error:
            app.logger.error(f"Error sending notification email: {str(email_error)}")
            email_sent = False
        
        # Return success response regardless of email status
        response = {
            'success': True, 
            'message': 'Company added successfully',
            'id': company_id
        }
        
        if email_sent:
            app.logger.info("Notification email sent successfully")
            response['message'] += '. Notification email sent.'
        else:
            app.logger.warning("Company added but failed to send notification email")
            response['message'] += '. Failed to send notification email.'
            response['warning'] = 'Email notification failed'
            
        return jsonify(response)
        
    except Exception as e:
        app.logger.error(f"Error adding company: {str(e)}", exc_info=True)
        error_message = str(e)
        
        # Provide more specific error messages for common issues
        if "duplicate key error" in error_message.lower():
            error_message = "A company with this name or email already exists."
        elif "timed out" in error_message.lower() or "connection" in error_message.lower():
            error_message = "Could not connect to the database. Please try again later."
            
        return jsonify({
            'success': False, 
            'message': f'Failed to add company: {error_message}'
        }), 500

@app.route('/api/add_machine', methods=['POST'])
@login_required
def api_add_machine():
    """Handle AJAX request to create a new machine"""
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Invalid request, JSON expected.'}), 400

    data = request.get_json()
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    gst_registered = bool(data.get('gst_registered', False))
    gst_number = data.get('gst_number', '').strip().upper() if gst_registered else ''
    billing_address = data.get('billing_address', '').strip() if gst_registered else ''

    if gst_registered:
        if not gst_number or not re.match(r'^[0-9A-Z]{15}$', gst_number):
            return jsonify({'success': False, 'message': 'A valid 15 character GST number is required.'}), 400
        if not billing_address:
            return jsonify({'success': False, 'message': 'Billing address is required when GST is provided.'}), 400

    if not name:
        return jsonify({'success': False, 'message': 'Machine name is required.'}), 400

    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            # Store machines in a single document that contains an array field `machines`
            # Find the document that holds the array (first document that has `machines`)
            master_doc = mongo_db.machine.find_one({'machines': {'$exists': True}})
            if master_doc is None:
                # Create master doc if it doesn't exist
                next_id = 1
                mongo_db.machine.insert_one({'machines': [{'id': next_id, 'name': name,
                                                           'description': description,
                                                           'gst_registered': gst_registered,
                                                           'gst_number': gst_number,
                                                           'billing_address': billing_address,
                                                           'created_at': datetime.utcnow()}]})
            else:
                machines_arr = master_doc.get('machines', [])
                # Check if machine with this name already exists
                if any(m.get('name') == name for m in machines_arr):
                    return jsonify({'success': False, 'message': 'A machine with this name already exists.'}), 400
                    
                # Determine next incremental id based on existing array length / max id
                if machines_arr:
                    next_id = max([m.get('id', 0) for m in machines_arr]) + 1
                else:
                    next_id = 1
                mongo_db.machine.update_one(
                    {'_id': master_doc['_id']},
                    {'$push': {'machines': {
                        'id': next_id,
                        'name': name,
                        'description': description,
                        'gst_registered': gst_registered,
                        'gst_number': gst_number,
                        'billing_address': billing_address,
                        'created_at': datetime.utcnow()
                    }}},
                    upsert=True
                )
            machine_id = str(next_id)
            # Send alert email
            user_identity = getattr(current_user, 'email', getattr(current_user, 'username', 'Unknown User'))
            send_alert_email(
                subject='Database Update: New Machine Added',
                body=f"{user_identity} added a new machine ({name}) on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
            )
        else:
            # File-based storage fallback
            os.makedirs(os.path.dirname(machines_file), exist_ok=True)
            machines_data = {"machines": []}
            if os.path.exists(machines_file):
                with open(machines_file, 'r', encoding='utf-8') as f:
                    machines_data = json.load(f) or {"machines": []}
            
            machines = machines_data.get('machines', [])
            
            # Check if machine with this name already exists
            if any(m.get('name') == name for m in machines):
                return jsonify({'success': False, 'message': 'A machine with this name already exists.'}), 400

            # Determine next ID
            next_id = (machines[-1]['id'] + 1) if machines else 1
            machines.append({
                'id': next_id, 
                'name': name, 
                'description': description,
                'gst_registered': gst_registered,
                'gst_number': gst_number,
                'billing_address': billing_address,
                'created_at': datetime.utcnow().isoformat()
            })
            machines_data['machines'] = machines
            with open(machines_file, 'w', encoding='utf-8') as f:
                json.dump(machines_data, f, ensure_ascii=False, indent=2)

        return jsonify({'success': True, 'message': 'Machine added successfully.', 'id': machine_id})
    except Exception as e:
        app.logger.error(f"Error adding machine: {e}")
        return jsonify({'success': False, 'message': 'Failed to add machine.'}), 500

# Step 1: Request Password Reset - Send OTP to email
# Step 2: Verify OTP
@app.route('/api/auth/request-password-reset', methods=['POST'])
def api_request_password_reset():
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    provided_phone = data.get('phone', '').strip()

    app.logger.info(f"Password reset request received: email={email}, phone={provided_phone}")

    if not email:
        app.logger.info("No email provided")
        return jsonify({'success': False, 'error': 'Email is required'}), 400

    app.logger.info("Starting user lookup")

    # Find user by email
    user = None
    if MONGO_AVAILABLE and USE_MONGO:
        doc = mu_find_user_by_email_or_username(email)
        if doc and doc.get('email', '').lower() == email:
            user = doc
    else:
        for u in users.values():
            if u.email.lower() == email:
                user = u
                break

    app.logger.info(f"User found: {user is not None}, user data: {user}")

    if not user:
        app.logger.info("User not found, returning security response")
        # For security, don't reveal if email exists
        return jsonify({'success': False}), 200

    # Get phone from user document if available
    if isinstance(user, dict):
        linked_phone = (
            user.get('phone')
            or user.get('Phone')
            or user.get('mobile')
            or user.get('Mobile')
            or user.get('contact')
            or user.get('Contact')
        )
    else:
        linked_phone = getattr(user, 'phone', None) or getattr(user, 'mobile', None)

    app.logger.info(f"Password reset request for email: {email}, linked_phone: {linked_phone}, provided_phone: {provided_phone}")

    target_phone = (linked_phone or '').strip() or provided_phone
    if not target_phone:
        app.logger.info(f"No phone found for email: {email}, returning phone_required")
        return jsonify({
            'success': False,
            'error': 'phone_required',
            'message': 'No phone number is linked to this email. Please enter your WhatsApp number to receive the OTP.'
        }), 400

    app.logger.info(f"Using phone: {target_phone} for OTP delivery")

    if MONGO_AVAILABLE and USE_MONGO and isinstance(user, dict) and provided_phone:
        try:
            if not (linked_phone or '').strip():
                users_col.update_one(
                    {'_id': user['_id']},
                    {'$set': {'phone': provided_phone, 'updated_at': time.time()}}
                )
        except Exception as e:
            app.logger.warning(f"Failed to persist phone number for {email}: {e}")

    # Generate OTP
    otp = ''.join(random.choices('0123456789', k=6))
    otp_expiry = datetime.utcnow() + timedelta(minutes=10)

    # Store OTP in user's record
    if MONGO_AVAILABLE and USE_MONGO:
        users_col.update_one(
            {'_id': user['_id']},
            {'$set': {
                'reset_token': otp,
                'reset_token_expiry': otp_expiry
            }}
        )
    else:
        user.reset_token = otp
        user.reset_token_expiry = otp_expiry
        save_users()

    # Send email with OTP
    body = (
        "<h2>Password Reset Request</h2>"
        "<p>You have requested to reset your password. Please use the following OTP to proceed:</p>"
        f"<h3 style=\"font-size: 24px; letter-spacing: 5px; margin: 20px 0;\">{otp}</h3>"
        "<p>This OTP will expire in 10 minutes.</p>"
        "<p>If you did not request this, please ignore this email.</p>"
    )

    email_ok = send_email_resend(
        to=email,
        subject='Password Reset OTP',
        html=body
    )

    if not email_ok:
        app.logger.error('Failed to send password reset email via Resend')
        return jsonify({'error': 'Failed to send password reset email. Please try again later.'}), 500

    wa_body = f"Your password reset OTP is: {otp}. This OTP will expire in 10 minutes."
    if not WA_SERVICE_URL or not WA_SERVICE_AUTH_TOKEN:
        app.logger.warning("WhatsApp service not configured (WA_SERVICE_URL or WA_SERVICE_AUTH_TOKEN missing)")
        return jsonify({'error': 'wa_not_configured', 'message': 'WhatsApp OTP is required but WhatsApp service is not configured on the server.'}), 500

    wa_ok = send_whatsapp_message(target_phone, wa_body)
    if not wa_ok:
        return jsonify({'error': 'Failed to send OTP to WhatsApp number. Please try again later.'}), 500

    return jsonify({
        'success': True,
        'message': 'If an account with that email exists, a password reset OTP has been sent.'
    })

@app.route('/api/auth/verify-reset-otp', methods=['POST'])
def api_verify_reset_otp():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        otp = data.get('otp', '').strip()

        if not all([email, otp]):
            return jsonify({'error': 'Email and OTP are required'}), 400

        user = None
        if MONGO_AVAILABLE and USE_MONGO:
            doc = mu_find_user_by_email_or_username(email)
            if doc and doc.get('email', '').lower() == email:
                user = doc
        else:
            for u in users.values():
                if u.email.lower() == email:
                    user = u
                    break

        if not user:
            return jsonify({'error': 'No account found with that email'}), 404

        stored_otp = user.get('reset_token') if isinstance(user, dict) else user.reset_token
        stored_expiry = user.get('reset_token_expiry') if isinstance(user, dict) else user.reset_token_expiry

        if not stored_otp or stored_otp != otp:
            return jsonify({'error': 'Invalid OTP'}), 400

        if stored_expiry and stored_expiry < datetime.utcnow():
            return jsonify({'error': 'OTP has expired'}), 400

        return jsonify({'success': True, 'message': 'OTP verified successfully'})

    except Exception as e:
        print(f"OTP verification error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/auth/reset-password', methods=['POST'])
def api_reset_password():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        otp = data.get('otp', '').strip()
        new_password = data.get('new_password', '').strip()
        
        if not all([email, otp, new_password]):
            return jsonify({'error': 'Email, OTP, and new password are required'}), 400
            
        print(f"Resetting password for email: {email}")
            
        # Find user by email
        user = None
        if MONGO_AVAILABLE and USE_MONGO:
            print("Looking up user in MongoDB...")
            doc = mu_find_user_by_email_or_username(email)
            if doc and doc.get('email', '').lower() == email:
                user = doc
                print(f"User found in MongoDB: {user['email']} (ID: {user['_id']})")
        else:
            # Fallback to JSON storage
            for u in users.values():
                if u.email.lower() == email:
                    user = u
                    break
                
        if not user:
            print(f"No user found with email: {email}")
            return jsonify({'error': 'No account found with that email'}), 404
            
        # Verify OTP
        stored_otp = user.get('reset_token') if isinstance(user, dict) else user.reset_token
        stored_expiry = user.get('reset_token_expiry') if isinstance(user, dict) else user.reset_token_expiry
        
        print(f"Verifying OTP - Stored: {stored_otp}, Provided: {otp}")
        
        if not stored_otp or stored_otp != otp:
            print("Invalid OTP")
            return jsonify({'error': 'Invalid OTP'}), 400
            
        if stored_expiry and stored_expiry < datetime.utcnow():
            print("OTP expired")
            return jsonify({'error': 'OTP has expired'}), 400
            
        # Update password
        if MONGO_AVAILABLE and USE_MONGO:
            print(f"Updating password for user {user['_id']}")
            # Create a new User instance to use its password hashing
            temp_user = User(
                id=str(user['_id']),
                email=user['email'],
                username=user['username'],
                password_hash=user.get('password_hash', '')
            )
            temp_user.set_password(new_password)
            
            # Update the user in MongoDB
            users_col.update_one(
                {'_id': user['_id']},
                {'$set': {
                    'password_hash': temp_user.password_hash,
                    'reset_token': None,
                    'reset_token_expiry': None
                }}
            )
            print("Password updated successfully in MongoDB")
        else:
            # Fallback to JSON storage
            user.set_password(new_password)
            user.reset_token = None
            user.reset_token_expiry = None
            save_users()
        
        return jsonify({
            'success': True,
            'message': 'Password has been reset successfully',
            'redirectTo': '/login'  # Redirect to login page after successful reset
        })
        
    except Exception as e:
        print(f"Password reset error: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/quotation_preview')
@login_required
@company_required
def quotation_preview():
    app.logger.info("[DEBUG] quotation_preview() called")
    
    # Get current date and time
    current_datetime = get_india_time()
    quote_date = current_datetime.strftime('%d-%m-%Y')
    quote_time = current_datetime.strftime('%H:%M:%S')
    
    cart = get_user_cart()
    app.logger.info(f"[DEBUG] Cart contains {len(cart.get('products', []))} products")
    
    if not cart.get('products'):
        app.logger.warning("[DEBUG] Empty cart, redirecting to cart page")
        flash('Your cart is empty', 'warning')
        return redirect(url_for('cart'))

    # Get company info from selected_company dict first, then fallback to direct session values
    selected_company = session.get('selected_company', {})
    app.logger.info(f"[DEBUG] Selected company from session: {selected_company}")

    payment_terms_default = ''
    company_id = session.get('company_id') or (selected_company.get('id') if isinstance(selected_company, dict) else None)
    payment_terms_company_id = session.get('payment_terms_company_id')

    # If the company changed, don't leak previous company's payment terms.
    if str(payment_terms_company_id or '') != str(company_id or ''):
        session.pop('payment_terms', None)
        session['payment_terms_company_id'] = str(company_id or '')
        session.modified = True

    payment_terms = (session.get('payment_terms') or '').strip()
    if not payment_terms and company_id and MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and ObjectId.is_valid(str(company_id)):
        try:
            company_doc = mongo_db.companies.find_one({'_id': ObjectId(str(company_id))}, {'last_payment_terms': 1})
            payment_terms = ((company_doc or {}).get('last_payment_terms') or '').strip()
            if payment_terms:
                session['payment_terms'] = payment_terms
                session['payment_terms_company_id'] = str(company_id)
                session.modified = True
        except Exception as e:
            app.logger.warning(f"Failed to load last payment terms for company {company_id}: {e}")

    # If still empty, keep it as "Non selected" (UI will show placeholder).
    if not payment_terms:
        payment_terms = payment_terms_default
        session['payment_terms'] = payment_terms
        session['payment_terms_company_id'] = str(company_id or '')
        session.modified = True
    
    customer_name = selected_company.get('name') or session.get('company_name', '')
    customer_email = selected_company.get('email') or session.get('company_email', '')
    app.logger.info(f"[DEBUG] Resolved customer: {customer_name} <{customer_email}>")
    
    # If we have company ID but no name/email, try to look it up
    if not customer_name and 'company_id' in session:
        try:
            company_id = session['company_id']
            file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
            with open(file_path, 'r') as f:
                companies = json.load(f)
            
            company = next((c for c in companies if str(c.get('id')) == str(company_id)), None)
            if company:
                customer_name = company.get('Company Name', customer_name)
                customer_email = company.get('EmailID', customer_email)
        except Exception as e:
            app.logger.error(f"Error looking up company info: {str(e)}")
    
    # Ensure values are stored in both places for consistency
    if customer_name or customer_email:
        if not isinstance(selected_company, dict):
            selected_company = {}
        
        if customer_name:
            selected_company['name'] = customer_name
            session['company_name'] = customer_name
        if customer_email:
            selected_company['email'] = customer_email
            session['company_email'] = customer_email
        
        session['selected_company'] = selected_company

    # Ensure all items have required fields and calculate subtotal
    subtotal = 0
    for item in cart.get('products', []):
        # Ensure all required fields exist with defaults
        item.setdefault('type', '')
        item.setdefault('quantity', 1)
        item.setdefault('discount_percent', 0)
        item.setdefault('gst_percent', 18)  # Default GST for mpack
        item.setdefault('unit_price', 0)
        item.setdefault('base_price', 0)
        item.setdefault('bar_price', 0)

        raw_quantity = item.get('quantity', 1)
        if item['type'] == 'chemical':
            try:
                quantity_value = float(item.get('quantity_litre') or item.get('total_litre') or raw_quantity or 0)
            except (TypeError, ValueError):
                quantity_value = 0.0
            quantity_value = max(quantity_value, 0.0)
        else:
            try:
                quantity_value = float(raw_quantity or 0)
            except (TypeError, ValueError):
                quantity_value = 0.0
            if quantity_value <= 0:
                quantity_value = 1.0

        if item['type'] == 'mpack':
            # Calculate mpack total matching cart template's approach
            price = float(item['unit_price'])
            quantity = int(quantity_value) if quantity_value else 1
            discount_percent = float(item['discount_percent'])
            gst_percent = float(item['gst_percent'])

            subtotal = price * quantity
            discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
            price_after_discount = subtotal - discount_amount
            gst_amount = (price_after_discount * gst_percent / 100) if gst_percent else 0
            final_total = price_after_discount + gst_amount
            
            # Store calculations in the item
            item['calculations'] = {
                'unit_price': round(price, 2),
                'quantity': quantity,
                'subtotal': round(subtotal, 2),
                'discount_percent': discount_percent,
                'discount_amount': round(discount_amount, 2),
                'price_after_discount': round(price_after_discount, 2),
                'gst_percent': gst_percent,
                'gst_amount': round(gst_amount, 2),
                'final_total': round(final_total, 2)
            }
            item_subtotal = final_total
            
        elif item['type'] == 'blanket':
            # Calculate blanket total matching cart template's approach
            base_price = float(item.get('base_price', 0))
            bar_price = float(item.get('bar_price', 0))
            quantity = int(quantity_value) if quantity_value else 1
            discount_percent = float(item.get('discount_percent', 0))
            gst_percent = float(item.get('gst_percent', 18))

            # Calculate unit price as base + bar price
            unit_price = base_price + bar_price
            
            # Calculate subtotal (unit price * quantity)
            subtotal = unit_price * quantity
            
            # Calculate discount amount
            discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
            
            # Apply discount to get discounted subtotal
            discounted_subtotal = subtotal - discount_amount
            
            # Calculate GST on discounted amount
            gst_amount = (discounted_subtotal * gst_percent / 100)
            
            # Final total after discount and GST
            final_total = discounted_subtotal + gst_amount
            
            # Store calculations in the item
            item['calculations'] = {
                'base_price': round(base_price, 2),
                'bar_price': round(bar_price, 2),
                'unit_price': round(unit_price, 2),
                'quantity': quantity,
                'subtotal': round(subtotal, 2),
                'discount_percent': discount_percent,
                'discount_amount': round(discount_amount, 2),
                'discounted_subtotal': round(subtotal - discount_amount, 2),
                'gst_percent': gst_percent,
                'gst_amount': round(gst_amount, 2),
                'final_total': round(final_total, 2)
            }
            item_subtotal = final_total
            
        else:
            # Handle other product types
            price = float(item.get('unit_price', 0))
            quantity = quantity_value
            discount_percent = float(item.get('discount_percent', 0))
            gst_percent = float(item.get('gst_percent', 18))

            subtotal = price * quantity
            discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
            discounted_subtotal = subtotal - discount_amount
            gst_amount = (discounted_subtotal * gst_percent / 100) if gst_percent else 0
            final_total = discounted_subtotal + gst_amount
            
            item['calculations'] = {
                'unit_price': round(price, 2),
                'quantity': quantity,
                'subtotal': round(subtotal, 2),
                'discount_percent': discount_percent,
                'discount_amount': round(discount_amount, 2),
                'gst_percent': gst_percent,
                'gst_amount': round(gst_amount, 2),
                'final_total': round(final_total, 2)
            }
            item_subtotal = final_total
        
        subtotal += item_subtotal
    
    # Calculate final totals with appropriate tax rates and discounts
    subtotal_blankets = 0
    subtotal_mpacks = 0
    discount_blankets = 0
    discount_mpacks = 0
    
    for item in cart.get('products', []):
        item_calc = item.get('calculations', {})
        item_subtotal = item_calc.get('subtotal', 0)
        item_discount = item_calc.get('discount_amount', 0)
        
        if item.get('type') == 'blanket':
            subtotal_blankets += item_subtotal
            discount_blankets += item_discount
        else:  # Assume mpacks for other types
            subtotal_mpacks += item_subtotal
            discount_mpacks += item_discount
    
    # Calculate amounts after discount
    subtotal_after_discount_blankets = max(0, subtotal_blankets - discount_blankets)
    subtotal_after_discount_mpacks = max(0, subtotal_mpacks - discount_mpacks)
    
    # Calculate GST for each category (on discounted amount)
    gst_blankets = subtotal_after_discount_blankets * 0.18  # 18% GST for blankets
    gst_mpacks = subtotal_after_discount_mpacks * 0.18      # 18% GST for mpacks
    
    # Calculate final totals
    subtotal_before_discount = subtotal_blankets + subtotal_mpacks
    total_discount = discount_blankets + discount_mpacks
    subtotal_after_discount = subtotal_after_discount_blankets + subtotal_after_discount_mpacks
    total_gst = gst_blankets + gst_mpacks
    total = subtotal_after_discount + total_gst
    
    # Round to 2 decimal places for display
    subtotal_before_discount = round(subtotal_before_discount, 2)
    total_discount = round(total_discount, 2)
    subtotal_after_discount = round(subtotal_after_discount, 2)
    total = round(total, 2)
    total_gst = round(total_gst, 2)

    # Ensure session is saved before rendering the template
    session.modified = True
    
    # Calculate cart_total as the subtotal after discount but before taxes
    cart_total = subtotal_after_discount
    
    context = {
        'cart': cart,
        'quote_date': quote_date,
        'quote_time': quote_time,
        'company_name': customer_name,
        'company_email': customer_email,
        'payment_terms': payment_terms,
        'company_details': build_quotation_company_details(
            selected_company,
            session.get('company_id'),
            session.get('company_email'),
            fallback={
                'name': customer_name,
                'email': customer_email
            }
        ),
        'now': current_datetime,  # Add current datetime object for the template
        'calculations': {
            'subtotal_before_discount': subtotal_before_discount,
            'total_discount': total_discount,
            'subtotal_after_discount': subtotal_after_discount,
            'total': total,
            'gst_breakdown': {
                'blankets': {
                    'subtotal': round(subtotal_blankets, 2),
                    'discount': round(discount_blankets, 2),
                    'subtotal_after_discount': round(subtotal_after_discount_blankets, 2),
                    'gst': round(gst_blankets, 2),
                    'rate': 18
                },
                'mpacks': {
                    'subtotal': round(subtotal_mpacks, 2),
                    'discount': round(discount_mpacks, 2),
                    'subtotal_after_discount': round(subtotal_after_discount_mpacks, 2),
                    'gst': round(gst_mpacks, 2),
                    'rate': 18
                },
                'total_gst': round(total_gst, 2)
            }
        },
        'cart_total': subtotal_after_discount  # cart_total is the subtotal after discount but before taxes
    }
    
    return render_template('quotation.html', **context)


@app.route('/quotation_pdf')
@login_required
@company_required
def quotation_pdf():
    if HTML is None:
        flash('PDF generation is not available on this server.', 'danger')
        return redirect(url_for('quotation_preview'))

    current_datetime = get_india_time()
    quote_date = current_datetime.strftime('%d-%m-%Y')
    quote_time = current_datetime.strftime('%H:%M:%S')

    cart = get_user_cart()
    if not cart.get('products'):
        flash('Your cart is empty', 'warning')
        return redirect(url_for('cart'))

    selected_company = session.get('selected_company', {})
    if not isinstance(selected_company, dict):
        selected_company = {}

    payment_terms = session.get('payment_terms', '')

    customer_name = selected_company.get('name') or session.get('company_name', '')
    customer_email = selected_company.get('email') or session.get('company_email', '')

    # Reuse the quotation_preview calculation logic so PDF always has complete totals.
    subtotal_blankets = 0
    subtotal_mpacks = 0
    discount_blankets = 0
    discount_mpacks = 0

    for item in cart.get('products', []):
        item.setdefault('type', '')
        item.setdefault('quantity', 1)
        item.setdefault('discount_percent', 0)
        item.setdefault('gst_percent', 18)
        item.setdefault('unit_price', 0)
        item.setdefault('base_price', 0)
        item.setdefault('bar_price', 0)

        raw_quantity = item.get('quantity', 1)
        if item['type'] == 'chemical':
            try:
                quantity_value = float(item.get('quantity_litre') or item.get('total_litre') or raw_quantity or 0)
            except (TypeError, ValueError):
                quantity_value = 0.0
            quantity_value = max(quantity_value, 0.0)
        else:
            try:
                quantity_value = float(raw_quantity or 0)
            except (TypeError, ValueError):
                quantity_value = 0.0
            if quantity_value <= 0:
                quantity_value = 1.0

        price = float(item.get('unit_price', 0) or 0)
        discount_percent = float(item.get('discount_percent', 0) or 0)
        gst_percent = float(item.get('gst_percent', 18) or 18)

        subtotal = price * quantity_value
        discount_amount = (subtotal * discount_percent / 100) if discount_percent else 0
        discounted_subtotal = subtotal - discount_amount
        gst_amount = (discounted_subtotal * gst_percent / 100) if gst_percent else 0
        final_total = discounted_subtotal + gst_amount

        item['calculations'] = item.get('calculations') or {}
        item['calculations'].update({
            'unit_price': round(price, 2),
            'quantity': quantity_value,
            'subtotal': round(subtotal, 2),
            'discount_percent': discount_percent,
            'discount_amount': round(discount_amount, 2),
            'discounted_subtotal': round(discounted_subtotal, 2),
            'gst_percent': gst_percent,
            'gst_amount': round(gst_amount, 2),
            'final_total': round(final_total, 2)
        })

        if item.get('type') == 'blanket':
            subtotal_blankets += subtotal
            discount_blankets += discount_amount
        else:
            subtotal_mpacks += subtotal
            discount_mpacks += discount_amount

    subtotal_before_discount = subtotal_blankets + subtotal_mpacks
    total_discount = discount_blankets + discount_mpacks
    subtotal_after_discount_blankets = max(0, subtotal_blankets - discount_blankets)
    subtotal_after_discount_mpacks = max(0, subtotal_mpacks - discount_mpacks)
    subtotal_after_discount = subtotal_after_discount_blankets + subtotal_after_discount_mpacks
    gst_blankets = subtotal_after_discount_blankets * 0.18
    gst_mpacks = subtotal_after_discount_mpacks * 0.18
    total_gst = gst_blankets + gst_mpacks
    total = subtotal_after_discount + total_gst

    calculations = {
        'subtotal_before_discount': round(subtotal_before_discount, 2),
        'total_discount': round(total_discount, 2),
        'subtotal_after_discount': round(subtotal_after_discount, 2),
        'total': round(total, 2),
        'gst_breakdown': {
            'total_gst': round(total_gst, 2)
        }
    }

    context = {
        'cart': cart,
        'quote_date': quote_date,
        'quote_time': quote_time,
        'company_name': customer_name,
        'company_email': customer_email,
        'payment_terms': payment_terms,
        'company_details': build_quotation_company_details(
            selected_company,
            session.get('company_id'),
            session.get('company_email'),
            fallback={
                'name': customer_name,
                'email': customer_email
            }
        ),
        'calculations': calculations,
        'now': current_datetime
    }

    # Render HTML first (absolute URLs needed for remote assets like logo)
    html = render_template('quotation_pdf.html', **context)
    pdf_bytes = HTML(string=html, base_url=request.url_root).write_pdf()

    filename = f"quotation_{quote_date.replace('-', '')}_{current_user.username}.pdf"
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/api/payment_terms', methods=['POST'])
@login_required
@company_required
def set_payment_terms():
    data = request.get_json() or {}
    payment_terms = (data.get('payment_terms') or '').strip()
    session['payment_terms'] = payment_terms
    session['payment_terms_company_id'] = str(session.get('company_id') or (session.get('selected_company', {}) or {}).get('id') or '')
    session.modified = True

    company_id = session.get('company_id') or (session.get('selected_company', {}) or {}).get('id')
    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and payment_terms and company_id and ObjectId.is_valid(str(company_id)):
        try:
            mongo_db.companies.update_one(
                {'_id': ObjectId(str(company_id))},
                {'$set': {'last_payment_terms': payment_terms, 'updated_at': datetime.utcnow()}},
                upsert=False
            )
        except Exception as e:
            app.logger.warning(f"Failed to persist payment terms for company {company_id}: {e}")

    return jsonify({'success': True, 'payment_terms': payment_terms})


@app.route('/api/quotation_phones', methods=['POST'])
@login_required
@company_required
def set_quotation_phones():
    data = request.get_json() or {}
    customer_phone = (data.get('customer_phone') or '').strip()
    prepared_by_phone = (data.get('prepared_by_phone') or '').strip()

    session['quotation_customer_phone'] = customer_phone
    session['quotation_prepared_by_phone'] = prepared_by_phone
    session.modified = True

    # Persist customer phone to the selected company record in MongoDB
    if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and customer_phone:
        company_id = session.get('company_id') or (session.get('selected_company', {}).get('id') if isinstance(session.get('selected_company'), dict) else None)
        if company_id and ObjectId.is_valid(str(company_id)):
            try:
                mongo_db.companies.update_one(
                    {'_id': ObjectId(str(company_id))},
                    {'$set': {'Phone': customer_phone, 'updated_at': datetime.utcnow()}}
                )
                app.logger.info(f"Updated company {company_id} phone to {customer_phone}")
            except Exception as e:
                app.logger.error(f"Failed to update company phone: {e}")

    return jsonify({'success': True, 'customer_phone': customer_phone, 'prepared_by_phone': prepared_by_phone})

# ---------------------------------------------------------------------------
# Send Quotation Route
# ---------------------------------------------------------------------------
@app.route('/send_quotation', methods=['POST'])
@login_required
@company_required
def send_quotation():
    """Generate quotation from current cart and email it to customer and CGI."""
    try:
        # Parse optional notes from request body
        data = request.get_json() or {}
        notes = (data.get('notes') or '').strip()
        payment_terms = (data.get('payment_terms') or session.get('payment_terms') or '').strip()
        customer_phone = (data.get('customer_phone') or session.get('quotation_customer_phone') or '').strip()
        prepared_by_phone = (data.get('prepared_by_phone') or session.get('quotation_prepared_by_phone') or '').strip()

        # Fetch cart
        cart = get_user_cart()
        products = cart.get('products', [])
    except Exception as e:
        app.logger.error(f"Error fetching cart or parsing data: {str(e)}")
        return jsonify({
            'error': f'Failed to fetch cart or parse data: {str(e)}',
            'details': str(e)
        }), 500

    try:
        if not products:
            return jsonify({'error': 'Cart is empty'}), 400

        if not customer_phone:
            return jsonify({'error': 'Customer phone is required'}), 400

        if not prepared_by_phone:
            return jsonify({'error': 'Prepared-by phone is required'}), 400

        selected_company = session.get('selected_company', {})
        if not isinstance(selected_company, dict):
            selected_company = {}
        # Get company info with proper fallbacks - prioritize database over session
        customer_name = 'Not specified'
        customer_email = ''
        # First try to get from user's company_id if available
        if hasattr(current_user, 'company_id') and current_user.company_id:
            customer_name = get_company_name_by_id(current_user.company_id)
            customer_email = get_company_email_by_id(current_user.company_id)

        # If not found in user's company_id, try session
        if customer_name == 'Not specified' or not customer_email:
            # Get from session if available
            if not customer_email:
                customer_email = (
                    selected_company.get('email') or 
                    session.get('company_email') or 
                    (hasattr(current_user, 'email') and current_user.email) or 
                    ''
                )
            
            if customer_name == 'Not specified':
                customer_name = (
                    selected_company.get('name') or 
                    session.get('company_name') or 
                    (hasattr(current_user, 'company_name') and current_user.company_name) or 
                    'Not specified'
                )
        
        if (customer_name == 'Not specified' or not customer_name) and customer_email:
            # Try Mongo lookup by email
            if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
                try:
                    company_doc = mongo_db.companies.find_one({'EmailID': customer_email})
                    if company_doc:
                        customer_name = (
                            company_doc.get('Company Name') or
                            company_doc.get('name') or
                            customer_name
                        )
                except Exception as lookup_error:
                    app.logger.error(f"Error looking up company by email: {lookup_error}")

            # JSON fallback lookup
            if (customer_name == 'Not specified' or not customer_name):
                try:
                    file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
                    if os.path.exists(file_path):
                        with open(file_path, 'r', encoding='utf-8') as f:
                            companies = json.load(f)
                        if isinstance(companies, dict):
                            companies = companies.get('companies', [])
                        match = next(
                            (
                                c for c in companies
                                if (c.get('EmailID') or c.get('email') or '').lower() == customer_email.lower()
                            ),
                            None
                        )
                        if match:
                            customer_name = (
                                match.get('Company Name') or
                                match.get('name') or
                                customer_name
                            )
                except Exception as file_error:
                    app.logger.error(f"Error loading company email mappings: {file_error}")

        # Final fallback to user's email if still no email
        if not customer_email and hasattr(current_user, 'email'):
            customer_email = current_user.email

        if not customer_email:
            return jsonify({'error': 'Customer email is required'}), 400

        if not customer_name or customer_name == 'Not specified':
            customer_name = customer_email or 'Customer'

        # Update session with the latest values
        if customer_name and customer_name != 'Not specified':
            # Update user's company info in database if using MongoDB
            if MONGO_AVAILABLE and USE_MONGO and hasattr(current_user, 'id'):
                try:
                    users_col.update_one(
                        {'_id': current_user.id},
                        {'$set': {
                            'company_name': customer_name,
                            'company_email': customer_email,
                            'company_id': current_user.company_id if hasattr(current_user, 'company_id') else None
                        }}
                    )
                except Exception as e:
                    app.logger.error(f"Error updating user's company info: {str(e)}")
            
            # Persist customer phone to the selected company record in MongoDB
            company_id = session.get('company_id') or (session.get('selected_company', {}).get('id') if isinstance(session.get('selected_company'), dict) else None)
            if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and customer_phone and company_id and ObjectId.is_valid(str(company_id)):
                try:
                    mongo_db.companies.update_one(
                        {'_id': ObjectId(str(company_id))},
                        {'$set': {'Phone': customer_phone, 'updated_at': datetime.utcnow()}}
                    )
                    app.logger.info(f"Updated company {company_id} phone to {customer_phone} on send_quotation")
                except Exception as e:
                    app.logger.error(f"Failed to update company phone on send_quotation: {e}")
            
            # Do not hydrate session company selection here. Company selection must be explicit.

        customer_details = build_quotation_company_details(
            selected_company,
            session.get('company_id'),
            session.get('company_email'),
            fallback={'name': customer_name, 'email': customer_email}
        )

        customer_name_html = escape(customer_details.get('name') or 'Not specified')
        customer_email_value = customer_details.get('email') or '--'
        if customer_email_value and customer_email_value not in ('--', 'Not specified'):
            customer_email_html = (
                f"<a href='mailto:{escape(customer_email_value)}' style='color: #0d6efd; text-decoration: none; word-break: break-word;'>{escape(customer_email_value)}</a>"
            )
        else:
            customer_email_html = "<span style='color: #6c757d;'>--</span>"

        customer_address_value = customer_details.get('address') or '--'
        customer_address_html = escape(customer_address_value).replace('\n', '<br>')
        customer_gst_display = customer_details.get('gst_display') or 'URP'
        customer_gst_html = escape(customer_gst_display)

        # Send to customer, operations email, and current user (remove duplicates)
        user_email = current_user.email if hasattr(current_user, 'email') else None
        recipients = list({email for email in [customer_email, 'operations@chemo.in', user_email] if email})

        quote_generated_at_utc = datetime.utcnow().replace(tzinfo=timezone.utc)
        quote_generated_at = quote_generated_at_utc.astimezone(IST)

        # Generate a unique quote ID early so it is always available for attachments/logging.
        quote_id = get_next_quote_id()

        if payment_terms:
            session['payment_terms'] = payment_terms
            session.modified = True

            company_id = session.get('company_id') or (session.get('selected_company', {}) or {}).get('id')
            if company_id and MONGO_AVAILABLE and USE_MONGO and mongo_db is not None and ObjectId.is_valid(str(company_id)):
                try:
                    mongo_db.companies.update_one(
                        {'_id': ObjectId(str(company_id))},
                        {'$set': {'last_payment_terms': payment_terms, 'updated_at': quote_generated_at}},
                        upsert=False
                    )
                except Exception as e:
                    app.logger.warning(f"Failed to update company payment terms for company {company_id}: {e}")
        quote_date_display = quote_generated_at.strftime('%d/%m/%Y')
        quote_time_display = quote_generated_at.strftime('%I:%M %p')

        payment_terms_display = payment_terms or '---'

        # Table rows with header
        rows_html = """
        <table style='width: 100%; border-collapse: collapse; margin: 20px 0;'>
            <thead>
                <tr style='background-color: #1a5276; color: white;'>
                    <th style='padding: 10px; text-align: left;'>Item</th>
                    <th style='padding: 10px; text-align: left;'>Machine</th>
                    <th style='padding: 10px; text-align: left;'>Product Type</th>
                    <th style='padding: 10px; text-align: left;'>Type</th>
                    <th style='padding: 10px; text-align: left;'>Thickness</th>
                    <th style='padding: 10px; text-align: left;'>Size</th>
                    <th style='padding: 10px; text-align: left;'>Barring Type</th>
                    <th style='padding: 10px; text-align: right;'>Qty</th>
                    <th style='padding: 10px; text-align: right;'>Price</th>
                    <th style='padding: 10px; text-align: right;'>Discount</th>
                    <th style='padding: 10px; text-align: right;'>Net Amount</th>
                </tr>
            </thead>
            <tbody>
        """

        subtotal = 0
        subtotal_before_discount = 0.0
        subtotal_after_discount = 0.0
        total_discount = 0.0
        total_gst = 0.0
        for idx, p in enumerate(products, start=1):
            machine = p.get('machine', '')
            prod_type = p.get('type', '')

            # Quantity handling (use litres for chemical items)
            raw_qty = p.get('quantity', 1)
            if prod_type == 'chemical':
                qty_source = p.get('quantity_litre') or p.get('total_litre') or raw_qty
            else:
                qty_source = raw_qty

            try:
                qty = float(qty_source if qty_source not in (None, '') else 0)
            except (TypeError, ValueError):
                qty = 0.0

            if prod_type == 'chemical':
                if qty <= 0:
                    display_qty = '0L'
                elif qty.is_integer():
                    display_qty = f"{int(qty)}L"
                else:
                    display_qty = f"{qty:.2f}L"
            else:
                if qty <= 0:
                    qty = 1.0
                display_qty = f"{qty:.0f}" if qty.is_integer() else f"{qty:.2f}"

            # Dimensions
            if p.get('size'):
                dimensions = p['size']
            else:
                length = p.get('length') or ''
                width = p.get('width') or ''
                unit = p.get('unit', '')
                dimensions = f"{length} x {width} {unit}" if length and width else '----'
            
            # Calculate total based on product type
            if prod_type == 'mpack':
                # Always recalculate MPack totals to ensure fresh values after quantity changes
                unit_price = float(p.get('unit_price', 0))
                discount_percent = float(p.get('discount_percent', 0))
                gst_percent = float(p.get('gst_percent', 18))  # 18% GST for MPack

                subtotal_val = unit_price * qty
                discount_amount = (subtotal_val * discount_percent / 100) if discount_percent else 0
                taxable_amount = subtotal_val - discount_amount
                gst_amount = taxable_amount * gst_percent / 100
                total_val = taxable_amount + gst_amount
                
                # Store discount percent for email template
                p['discount_percent_display'] = discount_percent
                
                # Add discount percent to calculations for display
                p['calculations'] = p.get('calculations', {})
                p['calculations']['discount_percent'] = discount_percent

                # Update (or create) calculations dict so subsequent routes remain consistent
                p['calculations'] = {
                    'unit_price': round(unit_price, 2),
                    'quantity': qty,
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'taxable_amount': round(taxable_amount, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(total_val, 2)
                }
                
            elif prod_type == 'blanket':
                # Always recalculate Blanket totals as well
                base_price = float(p.get('base_price', 0))
                bar_price = float(p.get('bar_price', 0))
                unit_price = base_price + bar_price
                discount_percent = float(p.get('discount_percent', 0))
                gst_percent = float(p.get('gst_percent', 18))

                subtotal_val = unit_price * qty
                discount_amount = subtotal_val * discount_percent / 100 if discount_percent else 0
                taxable_amount = subtotal_val - discount_amount
                gst_amount = taxable_amount * gst_percent / 100
                total_val = taxable_amount + gst_amount

                # Sync calculations back to product
                p['calculations'] = {
                    'unit_price': round(unit_price, 2),
                    'quantity': qty,
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'taxable_amount': round(taxable_amount, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(total_val, 2)
                }
            else:
                unit_price = float(p.get('unit_price', 0))
                discount_percent = float(p.get('discount_percent', 0))
                gst_percent = float(p.get('gst_percent', 18))

                subtotal_val = unit_price * qty
                discount_amount = (subtotal_val * discount_percent / 100) if discount_percent else 0
                taxable_amount = subtotal_val - discount_amount
                gst_amount = taxable_amount * gst_percent / 100
                total_val = taxable_amount + gst_amount

                p['calculations'] = {
                    'unit_price': round(unit_price, 2),
                    'quantity': qty,
                    'subtotal': round(subtotal_val, 2),
                    'discount_percent': discount_percent,
                    'discount_amount': round(discount_amount, 2),
                    'taxable_amount': round(taxable_amount, 2),
                    'gst_percent': gst_percent,
                    'gst_amount': round(gst_amount, 2),
                    'final_total': round(total_val, 2)
                }

            subtotal += total_val

            calc = p.get('calculations', {})
            line_unit_price = float(calc.get('unit_price', p.get('unit_price', p.get('base_price', 0) or 0)))
            line_quantity = float(calc.get('quantity', p.get('quantity', 1) or 0))
            line_pre_discount = line_unit_price * line_quantity
            line_discount = float(calc.get('discount_amount', 0) or 0)
            line_taxable = float(calc.get('taxable_amount', calc.get('subtotal', line_pre_discount - line_discount) or 0))
            line_gst = float(calc.get('gst_amount', 0) or 0)

            subtotal_before_discount += line_pre_discount
            total_discount += line_discount
            subtotal_after_discount += line_taxable
            total_gst += line_gst
            
            if prod_type == 'chemical':
                thickness_display = '&nbsp;'
            elif p.get('thickness'):
                if prod_type == 'blanket':
                    thickness_display = f"{str(p.get('thickness')).replace('.0', '')} mm"
                else:
                    thickness_value = p.get('thickness')
                    if thickness_value and not str(thickness_value).endswith(('mm', 'micron', 'in', 'cm')):
                        try:
                            thickness_float = float(thickness_value)
                        except (TypeError, ValueError):
                            thickness_float = None
                        thickness_suffix = ' mm' if thickness_float and thickness_float >= 1 else ''
                    else:
                        thickness_suffix = ''
                    thickness_display = f"{thickness_value}{thickness_suffix}" if thickness_value else '----'
            else:
                thickness_display = '----'

            net_amount = line_taxable

            rows_html += f"""
                <tr>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{idx}</td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{machine}</td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{'Underpacking' if prod_type == 'mpack' else prod_type if prod_type else '----'}</td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>
                        {p.get('blanket_type', p.get('name', '----')) if prod_type == 'blanket' 
                        else p.get('underpacking_type', '----').replace('_', ' ').title() if prod_type == 'mpack' 
                        else p.get('name', '----')}
                    </td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{thickness_display}</td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{dimensions}</td>
                    <td style='padding: 8px; border: 1px solid #ddd;'>{p.get('bar_type', '----') if prod_type == 'blanket' else '----'}</td>
                    <td style='padding: 8px; text-align: right; border: 1px solid #ddd;'>{display_qty}</td>
                    <td style='padding: 8px; text-align: right; border: 1px solid #ddd;'>₹{p.get('unit_price', p.get('base_price', 0)):,.2f}</td>
                    <td style='padding: 8px; text-align: right; border: 1px solid #ddd;'>{p.get('discount_percent', 0):.1f}%</td>
                    <td style='padding: 8px; text-align: right; border: 1px solid #ddd;'>₹{net_amount:,.2f}</td>
                </tr>
            """
        
        # Close the table
        rows_html += """
            </tbody>
        </table>
        <p>For more information, please contact: <a href='mailto:info@chemo.in'>info@chemo.in</a></p>
        """

        # Calculate discount information from products
        blanket_discounts = [p.get('discount_percent', 0) for p in products if p.get('type') == 'blanket' and p.get('discount_percent', 0) > 0]
        mpack_discounts = [p.get('discount_percent', 0) for p in products if p.get('type') == 'mpack' and p.get('discount_percent', 0) > 0]
        
        # Generate discount text for email
        discount_text = []
        if blanket_discounts:
            discount_text.append(f"{max(blanket_discounts):.1f}% Blanket")
        if mpack_discounts:
            discount_text.append(f"{max(mpack_discounts):.1f}% Underpacking")
        discount_text = ", ".join(discount_text)
        
        # Determine if we should show the discount row
        show_discount = bool(blanket_discounts or mpack_discounts)

        total = subtotal_after_discount + total_gst

        subtotal_before_discount = round(subtotal_before_discount, 2)
        total_discount = round(total_discount, 2)
        subtotal_after_discount = round(subtotal_after_discount, 2)
        total_gst = round(total_gst, 2)
        total = round(total, 2)

        pdf_attachments = None
        try:
            if HTML is not None:
                pdf_context = {
                    'cart': cart,
                    'quote_date': quote_generated_at.strftime('%d-%m-%Y'),
                    'quote_time': quote_generated_at.strftime('%H:%M:%S'),
                    'company_name': customer_name,
                    'company_email': customer_email,
                    'payment_terms': payment_terms,
                    'company_details': build_quotation_company_details(
                        session.get('selected_company', {}) if isinstance(session.get('selected_company', {}), dict) else {},
                        session.get('company_id'),
                        session.get('company_email'),
                        fallback={'name': customer_name, 'email': customer_email}
                    ),
                    'calculations': {
                        'subtotal_before_discount': subtotal_before_discount,
                        'total_discount': total_discount,
                        'subtotal_after_discount': subtotal_after_discount,
                        'total': total,
                        'gst_breakdown': {
                            'total_gst': total_gst
                        }
                    },
                    'now': quote_generated_at
                }
                pdf_html = render_template('quotation_pdf.html', **pdf_context)
                pdf_bytes = HTML(string=pdf_html, base_url=request.url_root).write_pdf()
                safe_customer_name = secure_filename((customer_name or '').strip()) or 'customer'
                pdf_filename = f"{safe_customer_name}_{quote_id}.pdf"
                pdf_attachments = [
                    {
                        'filename': pdf_filename,
                        'content': base64.b64encode(pdf_bytes).decode('utf-8'),
                        'type': 'application/pdf'
                    }
                ]
        except Exception as pdf_err:
            app.logger.warning(f"Failed to generate PDF attachment for email: {pdf_err}")

        logo_src = "https://cgi-logo.tiiny.site/CGI_LOGO.svg"

        company_name_display = "Chemo Graphic International (CGI)"
        company_email_html = "<a href='mailto:info@chemo.in' style='color: #0d6efd; text-decoration: none; word-break: break-word;'>info@chemo.in</a>"
        company_phone_html = "<a href='tel:+919930070755' style='color: #0d6efd; text-decoration: none;'>9930070755</a>"
        company_address_html = "113, High Tech Industrial Centre,<br>Caves Road, Jogeshwari (East),<br>Mumbai - 400060"
        company_gst_html = "27AAAPB9020H1Z6"

        email_content = f"""
        <div style='font-family: Arial, sans-serif; color: #333; max-width: 900px; margin: 0 auto; line-height: 1.6; background-color: #e0caa9; padding: 12px;'>
          <div style='background-color: white; border-radius: 0.5rem; box-shadow: 0 0.125rem 0.25rem rgba(0, 0, 0, 0.075); padding: 1.5rem; margin-bottom: 1rem;'>
            <div>
              <div style='text-align: center; margin-bottom: 1.25rem;'>
                <img src='{logo_src}' alt='CGI Logo' style='max-width: 180px; margin-bottom: 0.5rem;'>
                <h2 style='margin: 0 0 0.5rem 0; color: #2c3e50;'>QUOTATION</h2>
                <p style='color: #6c757d; margin: 0; font-size: 0.9rem;'>{quote_date_display}</p>
              </div>
              
              <div style='margin-bottom: 2rem;'>
                <table role='presentation' cellpadding='0' cellspacing='0' width='100%' style='border-collapse: separate; border-spacing: 24px 0;'>
                  <tr>
                    <td style='vertical-align: top; width: 50%; padding: 0;'>
                      <table role='presentation' cellpadding='0' cellspacing='0' width='100%' style='border-radius: 12px; border: 1px solid #dee2e6; overflow: hidden;'>
                      <tr>
                        <td style='background-color: #f8f9fa; padding: 16px 20px;'>
                          <h5 style='margin: 0; font-size: 16px; font-weight: 600;'>Company Information</h5>
                        </td>
                      </tr>
                      <tr>
                        <td style='padding: 20px; background-color: #ffffff;'>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-building' style='color: #0d6efd; margin-right: 10px; font-size: 18px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Company Name</div>
                              <div style='font-weight: 600;'>{escape(company_name_display)}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-envelope' style='color: #6c757d; margin-right: 10px; font-size: 16px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Email</div>
                              <div>{company_email_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex;'>
                            <i class='fas fa-map-marker-alt' style='color: #6c757d; margin-right: 10px; font-size: 16px; margin-top: 4px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Address</div>
                              <div>{company_address_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-phone' style='color: #6c757d; margin-right: 10px; font-size: 16px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Phone</div>
                              <div>{company_phone_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-id-card' style='color: #6c757d; margin-right: 10px; font-size: 16px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>GSTIN</div>
                              <div>{company_gst_html}</div>
                            </div>
                          </div>
                          <div style='margin-top: 16px; padding-top: 16px; border-top: 1px solid #dee2e6;'>
                            <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Prepared by</div>
                            <div style='font-weight: 600;'>{current_user.username}</div>
                            <div><a href='mailto:{current_user.email}' style='color: #0d6efd; text-decoration: none;'>{current_user.email}</a></div>
                          </div>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <td style='vertical-align: top; width: 50%; padding: 0;'>
                    <table role='presentation' cellpadding='0' cellspacing='0' width='100%' style='border-radius: 12px; border: 1px solid #dee2e6; overflow: hidden;'>
                      <tr>
                        <td style='background-color: #f8f9fa; padding: 16px 20px;'>
                          <h5 style='margin: 0; font-size: 16px; font-weight: 600;'>Customer Information</h5>
                        </td>
                      </tr>
                      <tr>
                        <td style='padding: 20px; background-color: #ffffff;'>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-building' style='color: #0d6efd; margin-right: 10px; font-size: 18px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Company Name</div>
                              <div style='font-weight: 600;'>{customer_name_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-envelope' style='color: #6c757d; margin-right: 10px; font-size: 16px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Email</div>
                              <div>{customer_email_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex;'>
                            <i class='fas fa-map-marker-alt' style='color: #6c757d; margin-right: 10px; font-size: 16px; margin-top: 4px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Address</div>
                              <div>{customer_address_html}</div>
                            </div>
                          </div>
                          <div style='margin-bottom: 16px; display: flex; align-items: center;'>
                            <i class='fas fa-id-card' style='color: #6c757d; margin-right: 10px; font-size: 16px;'></i>
                            <div>
                              <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>GSTIN</div>
                              <div>{customer_gst_html}</div>
                            </div>
                          </div>
                          <div style='margin-top: 16px; padding-top: 16px; border-top: 1px solid #dee2e6;'>
                            <div style='color: #6c757d; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em;'>Quotation Details</div>
                            <div style='font-weight: 600;'>Quotation #{quote_id}</div>
                            <div>{quote_date_display} | {quote_time_display}</div>
                          </div>
                        </td>
                      </tr>
                    </table>
                  </td>
                </tr>
              </table>
            </div>
            
            <div style='margin: 1.5rem 0; border: 1px solid #dee2e6; border-radius: 0.25rem; overflow: hidden;'>
              <div style='background-color: #f8f9fa; padding: 0.75rem 1.25rem; border-bottom: 1px solid rgba(0,0,0,0.125);'>
                <h5 style='margin: 0; font-size: 1rem;'>Quotation Details</h5>
              </div>
              <div style='padding: 1.5rem; background-color: white;'>
                <p style='margin-bottom: 1.5rem; font-size: 1rem;'>Hello <strong>{customer_name}</strong>,</p>
                <p style='margin-bottom: 1.5rem;'>This is <strong>{current_user.username}</strong> from CGI.</p>
                <p style='margin-bottom: 1.5rem;'>Here is the proposed quotation for the required products:</p>

                <p style='margin: 0 0 1rem 0;'><strong>Payment Terms:</strong> {payment_terms_display}</p>
                {'<p style="margin-bottom: 1.5rem;"><strong>Notes:</strong><br>' + notes + '</p>' if notes else ''}
                
                <div style='overflow-x: auto; margin: 1.5rem 0;'>
{rows_html}
                </div>
                
                <!-- Tax and Total Breakdown -->
                <div style='margin: 2rem 0;'>
                    <div style='display: flex; justify-content: flex-end;'>
                        <div style='width: 100%; max-width: 360px; margin-left: auto;'>
                            <div style='overflow-x: auto; margin: 1.5rem 0;'>
                                <table style='width: 100%; border-collapse: collapse;'>
                                    <tbody>
                                        <tr>
                                            <td style='padding: 8px; text-align: right; width: 70%;'>Subtotal (Pre-Discount):</td>
                                            <td style='padding: 8px; text-align: right; width: 30%;'>₹{sum((p.get('unit_price', p.get('base_price', 0))) * p.get('quantity', 1) for p in products):,.2f}</td>
                                        </tr>
                                        {f'''
                                        <tr style="display: {'table-row' if show_discount else 'none'};">
                                            <td style="padding: 8px; text-align: right;">Discount :</td>
                                            <td style="padding: 8px; text-align: right; color: #dc3545;">-₹{total_discount:,.2f}</td>
                                        </tr>
                                        ''' if True else ''}
                                        <tr style='border-top: 1px solid #dee2e6;'>
                                            <td style='padding: 8px; text-align: right; font-weight: bold;'>Total (Pre-GST):</td>
                                            <td style='padding: 8px; text-align: right; font-weight: bold;'>₹{sum(p.get("calculations", {}).get("taxable_amount", p.get("calculations", {}).get("subtotal", 0)) for p in products):,.2f}</td>
                                        </tr>
                                    
                                        <tr>
                                            <td style='padding: 8px; text-align: right;'>GST (18%):</td>
                                            <td style='padding: 8px; text-align: right;'>₹{total_gst:,.2f}</td>
                                        </tr>
                                        
                                        <tr style='border-top: 1px solid #dee2e6;'>
                                            <td style='padding: 8px; text-align: right; font-weight: bold;'>Total (After GST):</td>
                                            <td style='padding: 8px; text-align: right; font-weight: bold;'>₹{sum(p.get("calculations", {}).get("final_total", 0) for p in products):,.2f}</td>
                                        </tr>
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                </div>
                
                <p style='margin: 2rem 0 1rem 0;'>Thank you for your business!<br>— Team CGI</p>
              </div>
            </div>
            
            <div style='margin-top: 1.5rem; padding: 1rem; background-color: #f8f9fa; border-radius: 0.25rem; text-align: center;'>
              <p style='color: #6c757d; font-size: 0.8rem; margin: 0;'>
                This quotation is not a contract or invoice. It is our best estimate.
              </p>
            </div>
          </div>
        </div>
        """

        # Persist quotation in DB
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                mongo_db.quotations.insert_one({
                    'quote_id': quote_id,
                    'created_at': quote_generated_at,
                    'created_at_iso': quote_generated_at.isoformat(),
                    'created_at_ist': quote_generated_at.isoformat(),
                    'created_at_utc': quote_generated_at_utc.isoformat(),
                    'from_company': 'Chemo Graphic International (CGI)',
                    'from_email': 'info@chemo.in',
                    'prepared_by_name': current_user.username,
                    'prepared_by_email': current_user.email,
                    'prepared_by_phone': prepared_by_phone,
                    'user_id': str(current_user.id),
                    'username': current_user.username,
                    'user_email': current_user.email,
                    'company_name': customer_name,
                    'company_email': customer_email,
                    'company_phone': customer_phone,
                    'payment_terms': payment_terms,
                    'products': products,
                    'products_count': len(products),
                    'subtotal_before_discount': subtotal_before_discount,
                    'total_discount': total_discount,
                    'subtotal_after_discount': subtotal_after_discount,
                    'total_amount_pre_gst': subtotal_after_discount,
                    'gst_amount': total_gst,
                    'total_amount_post_gst': total,
                    'total_gst': total_gst,
                    'discount_text': discount_text,
                    'notes': notes,
                    'generated_at_time_display': quote_time_display,
                    'generated_at_date_display': quote_date_display
                })
            except Exception as db_err:
                app.logger.error(f"Failed to save quotation: {db_err}")

        email_sent = send_email_resend(
            to=recipients,
            subject=f"Quotation from Chemo INTERNATIONAL - {quote_date_display}",
            html=email_content,
            attachments=pdf_attachments
        )

        if email_sent:
            app.logger.info("Quotation email sent successfully via Resend")
        else:
            app.logger.error("Quotation email failed to send via Resend")

        # Send WhatsApp message to both customer and prepared-by/user with PDF attachment
        wa_body = f"Quotation #{quote_id} from Chemo Graphic International\n\nDate: {quote_date_display}\nTotal: ₹{total:,.2f}\n\nPayment Terms: {payment_terms or 'N/A'}\n\nPlease find the quotation PDF attached."
        wa_attachment = None
        if pdf_attachments and len(pdf_attachments) > 0:
            wa_attachment = {
                'filename': pdf_attachments[0]['filename'],
                'content': pdf_attachments[0]['content'],
                'type': pdf_attachments[0]['type']
            }

        whatsapp_results = {}
        targets = []
        if customer_phone:
            targets.append(('customer', customer_phone))
        user_phone_target = prepared_by_phone or getattr(current_user, 'phone', None)
        if user_phone_target:
            targets.append(('user', user_phone_target))

        seen_numbers = set()
        for label, phone in targets:
            phone_key = (phone or '').strip()
            if not phone_key or phone_key in seen_numbers:
                continue
            seen_numbers.add(phone_key)
            sent_ok = send_whatsapp_message(phone, wa_body, wa_attachment)
            whatsapp_results[label] = {
                'phone': phone,
                'sent': bool(sent_ok)
            }
            if sent_ok:
                app.logger.info(f"Quotation WhatsApp sent to {label}: {phone}")
            else:
                app.logger.warning(f"Quotation WhatsApp failed to send to {label}: {phone}")

        whatsapp_sent = any(v.get('sent') for v in whatsapp_results.values())

        # Clear cart after attempting to send email
        clear_cart()
        
        # Instead of removing selected_company, keep it in the session
        # This ensures the company selection persists after sending a quotation
        
        return jsonify({
            'success': True,
            'message': 'Quotation processed successfully',
            'email_sent': email_sent,
            'whatsapp_sent': whatsapp_sent,
            'whatsapp_results': whatsapp_results,
            'quote_id': quote_id,
            'generated_at': quote_generated_at.isoformat(),
            'generated_at_date': quote_date_display,
            'generated_at_time': quote_time_display,
            'company': {
                'id': session.get('selected_company', {}).get('id'),
                'name': session.get('selected_company', {}).get('name'),
                'email': session.get('selected_company', {}).get('email'),
                'phone': customer_phone
            }
        })
    except Exception as e:
        app.logger.error(f"Error sending quotation: {str(e)}")
        return jsonify({
            'error': 'Failed to send quotation',
            'details': str(e)
        }), 500

@app.route('/api/request-otp', methods=['POST'])
def api_request_otp():
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
            
        # Generate OTP
        otp = str(random.randint(100000, 999999))
        
        # Store OTP in user session with proper expiry format
        session['otp'] = otp
        session['otp_expiry'] = (datetime.now() + timedelta(minutes=5)).isoformat()
        
        # Send OTP to email
        if email_config_valid:
            if not send_email_resend(
                to=email,
                subject='Your OTP for Registration',
                text=f"Your OTP is: {otp}\nThis OTP will expire in 5 minutes."
            ):
                app.logger.warning("Error sending OTP email via Resend")
                return jsonify({'error': 'Failed to send OTP. Please try again later.'}), 500
        else:
            app.logger.info("Email configuration invalid; skipping OTP email send")
                
        return jsonify({
            'success': True,
            'message': 'OTP has been sent to your email'
        })
        
    except Exception as e:
        print(f"OTP request error: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/verify-otp', methods=['POST'])
def api_verify_otp():
    try:
        data = request.get_json()
        otp = data.get('otp')
        
        if not otp:
            return jsonify({'error': 'OTP is required'}), 400
            
        # Get stored OTP from session
        stored_otp = session.get('otp')
        otp_expiry = session.get('otp_expiry')
        
        if not stored_otp:
            return jsonify({'error': 'No OTP requested. Please request OTP first.'}), 400
            
        if otp_expiry and datetime.now() > datetime.fromisoformat(str(otp_expiry)):
            return jsonify({'error': 'OTP has expired. Please request a new OTP.'}), 400
            
        if otp != stored_otp:
            return jsonify({'error': 'Invalid OTP'}), 401
            
        # Clear OTP from session after successful verification
        session.pop('otp', None)
        session.pop('otp_expiry', None)
        
        return jsonify({
            'success': True,
            'message': 'OTP verified successfully'
        })
        
    except Exception as e:
        print(f"OTP verification error: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/auth/register/complete', methods=['POST'])
def api_register_complete():
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        phone = (data.get('phone') or '').strip()
        new_user = None
        
        print(f"\n🔍 Registration attempt for: {email} ({username})")
        print(f"MongoDB Status - Available: {MONGO_AVAILABLE}, Using: {USE_MONGO}")
        print(f"Users Collection: {users_col}")
        if users_col is not None:
            print(f"Collection name: {users_col.name}")
            print(f"Database name: {users_col.database.name}")
        
        otp = data.get('otp', '').strip()  # Get OTP if provided

        # Input validation
        if not all([email, username, password]):
            return jsonify({'error': 'Email, username, and password are required'}), 400
            
        print(f'Registration attempt - Email: {email}, Username: {username}')

        if MONGO_AVAILABLE and USE_MONGO:
            try:
                # Check for existing user
                existing_user = mu_find_user_by_email_or_username(email) or mu_find_user_by_email_or_username(username)
                if existing_user:
                    print(f'Registration failed: User already exists with email/username: {email}/{username}')
                    return jsonify({'error': 'Email or username already exists'}), 400
                
                # Create user in MongoDB
                print('Creating new user in MongoDB...')
                print(f'User details - Email: {email}, Username: {username}')
                
                # Debug: Check if we can access the users collection
                print(f'Users collection exists: {users_col is not None}')
                if users_col is not None:
                    print(f'Current users in DB: {users_col.count_documents({})}')
                
                # Create user and retrieve document
                user_id = mu_create_user(email, username, password, phone)
                print(f'User created with ID: {user_id}')
                
                # Try to retrieve user document using both UUID and ObjectId formats
                doc = mu_find_user_by_id(user_id)
                if not doc:
                    try:
                        from bson import ObjectId
                        doc = mu_find_user_by_id(ObjectId(user_id))
                    except Exception as e:
                        print(f"Error retrieving user document: {str(e)}")
                        traceback.print_exc()
                        return jsonify({'error': 'Failed to create user in database'}), 500
                
                print(f'Retrieved user document: {doc is not None}')
                
                if not doc:
                    traceback.print_exc()
                    return jsonify({'error': 'Failed to create user in database'}), 500

                new_user = User(
                    id=str(doc['_id']),
                    email=doc['email'],
                    username=doc['username'],
                    password_hash=doc['password_hash'],
                    is_verified=doc.get('is_verified', False),
                    otp_verified=doc.get('otp_verified', False),
                    company_id=doc.get('company_id')
                )
                
            except Exception as e:
                error_msg = f"❌ MongoDB Error: {str(e)}"
                print(error_msg)
                traceback.print_exc()
                
                # Try to clean up any partially created user
                try:
                    if 'user_id' in locals() and user_id:
                        users_col.delete_one({'_id': user_id})
                except Exception as cleanup_error:
                    print(f"Error during cleanup: {str(cleanup_error)}")
                
                return jsonify({'error': 'Failed to create user in database'}), 500
                
        else:
            # Fallback to JSON storage
            log_time("Using JSON storage fallback")
            try:
                users = load_users()
                
                # Check for existing user in the loaded data
                if any(u.email.lower() == email.lower() or u.username.lower() == username.lower() 
                      for u in users.values() if hasattr(u, 'email') and hasattr(u, 'username')):
                    return jsonify({'error': 'Email or username already exists'}), 409
                    
                # Create new user
                user_id = str(uuid.uuid4())
                new_user = User(user_id, email, username, password)
                new_user.set_password(password)
                users[user_id] = new_user
                
                # Save with retry logic
                max_retries = 3
                for attempt in range(max_retries):
                    if save_users(users):
                        break
                    if attempt == max_retries - 1:
                        return jsonify({'error': 'Failed to save user data after multiple attempts'}), 500
                    time.sleep(0.5)  # Short delay before retry
                    
                log_time("User created in JSON storage")
                
            except Exception as e:
                error_msg = f"❌ JSON storage error: {str(e)}"
                print(error_msg)
                traceback.print_exc()
                return jsonify({'error': 'Failed to create user in JSON storage'}), 500

        # 4. Finalize registration
        try:
            # Auto-login the newly registered user
            login_user(new_user)
            
            # Log successful registration
            log_time(f"User {username} registered and logged in successfully")
            
            # Prepare success response
            response_data = {
                'success': True,
                'message': 'Registration successful',
                'redirectTo': '/index',  # Redirect to index after successful registration
                'user': {
                    'id': new_user.id,
                    'email': new_user.email,
                    'username': new_user.username
                },
                'timestamp': time.time(),
                'executionTime': f"{time.time() - start_time:.2f}s"
            }
            
            # Add additional debug info in development
            if app.debug or app.config.get('ENV') == 'development':
                response_data['debug'] = {
                    'mongoAvailable': MONGO_AVAILABLE,
                    'useMongo': USE_MONGO,
                    'userStorage': 'MongoDB' if MONGO_AVAILABLE and USE_MONGO else 'JSON',
                    'userCount': users_col.count_documents({}) if MONGO_AVAILABLE and USE_MONGO else len(users)
                }
            
            return jsonify(response_data)
            
        except Exception as e:
            error_msg = f"❌ Error during login after registration: {str(e)}"
            print(error_msg)
            traceback.print_exc()
            
            # Even if login fails, registration was successful
            return jsonify({
                'success': True,
                'message': 'Registration successful. Please log in.',
                'redirectTo': '/login',
                'error': 'Auto-login failed'
            }), 200
        
    except Exception as e:
        error_msg = f"❌ Registration error: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        
        # Log the full error for debugging
        app.logger.error("Registration failed", exc_info=True)
        
        # Return a more specific error message if possible
        error_detail = str(e)
        if "duplicate key error" in error_detail.lower():
            if "email" in error_detail.lower():
                return jsonify({'error': 'A user with this email already exists'}), 409
            elif "username" in error_detail.lower():
                return jsonify({'error': 'This username is already taken'}), 409
                
        return jsonify({
            'error': 'An error occurred during registration',
            'details': str(e) if app.debug else None,
            'timestamp': time.time(),
            'executionTime': f"{time.time() - start_time:.2f}s"
        }), 500

@app.route('/api/auth/login', methods=['GET', 'POST', 'OPTIONS'])
def api_login():
    if request.method == 'OPTIONS':
        # Handle preflight request
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        return response
        
    try:
        # Handle both form data and JSON
        if request.is_json:
            data = request.get_json()
            if not data:
                return jsonify({'error': 'Invalid JSON data'}), 400, {'Content-Type': 'application/json'}
        else:
            data = request.form
            if not data:
                return jsonify({'error': 'Invalid form data'}), 400, {'Content-Type': 'application/json'}

        identifier = (data.get('identifier') or data.get('email') or data.get('username', '')).strip()
        password = (data.get('password') or '').strip()
        
        print(f'Login attempt - Identifier: {identifier}')
        
        if not identifier or not password:
            print('Login failed: Missing identifier or password')
            return jsonify({'error': 'Email/username and password are required'}), 400
            
        if MONGO_AVAILABLE and USE_MONGO:
            try:
                print('\n=== Login Debug ===')
                print(f'MONGO_AVAILABLE: {MONGO_AVAILABLE}, USE_MONGO: {USE_MONGO}')
                print(f'Users collection: {users_col}')
                if users_col is not None:
                    print(f'Collection name: {users_col.name}, DB: {users_col.database.name}')
                
                print('Attempting to find user in MongoDB...')
                # Find user by email or username in Mongo (case-insensitive)
                doc = mu_find_user_by_email_or_username(identifier)
                
                if not doc:
                    print(f'❌ User not found for identifier: {identifier}')
                    # Check if the identifier is an email or username
                    is_email = '@' in identifier
                    if is_email:
                        return jsonify({
                            'error': 'Email not found',
                            'message': 'No account found with this email address. Please check and try again.'
                        }), 401
                    else:
                        return jsonify({
                            'error': 'Username not found',
                            'message': 'No account found with this username. Please check and try again.'
                        }), 401
                    
                print(f'✅ User found in MongoDB:')
                print(f'   Email: {doc.get("email")} (stored in DB)')
                print(f'   Username: {doc.get("username")} (stored in DB)')
                print(f'   ID: {doc.get("_id")}')
                print(f'   Has password_hash: {"password_hash" in doc}')
                
                # Ensure we have the correct case for the username from the DB
                # This ensures we return the exact case that was used during registration
                identifier = doc.get('email', identifier) if '@' in identifier else doc.get('username', identifier)
                
                # Verify password
                print('\nVerifying password...')
                is_password_correct = mu_verify_password(doc, password)
                print(f'Password verification result: {is_password_correct}')
                
                if not is_password_correct:
                    print('❌ Password verification failed')
                    return jsonify({
                        'error': 'Incorrect password',
                        'message': 'The password you entered is incorrect. Please try again.'
                    }), 401
                
                # Create user object
                user = User(
                    id=str(doc['_id']),
                    email=doc['email'],
                    username=doc['username'],
                    password_hash=doc['password_hash'],
                    is_verified=doc.get('is_verified', False),
                    otp_verified=doc.get('otp_verified', False)
                )
                
                print(f'Successfully created user object for login: {user.email} (ID: {user.id})')

                # Clear any previously selected company from earlier sessions
                reset_company_selection_session()

                login_user(user)
                clear_login_prompt_flash()
                pricing_mode = 'gm' if has_gm_pricing_access(user) else 'standard'
                set_pricing_mode(pricing_mode)
                redirect_path = '/gm-page' if pricing_mode == 'gm' else '/index'
                print(f'User {user.username} logged in successfully (mode={pricing_mode})')
                
                if request.is_json:
                    response = jsonify({
                        'success': True,
                        'message': 'Login successful',
                        'redirectTo': redirect_path,
                        'user': {
                            'id': str(user.id),
                            'email': user.email,
                            'username': user.username
                        }
                    })
                else:
                    # For form submission, redirect directly
                    return redirect(url_for('gm_page' if pricing_mode == 'gm' else 'index'))
                
                # Set session
                session['user_id'] = str(user.id)
                session['user_email'] = user.email
                session['username'] = user.username
                
                return response
                
            except Exception as e:
                print(f'MongoDB login error: {str(e)}')
                import traceback
                traceback.print_exc()
                return jsonify({'error': 'Authentication service unavailable'}), 500

        # ---------------- JSON fallback path -----------------
        print('Falling back to JSON user storage')
        global users
        users = load_users()
        
        # Check if user exists in our loaded users
        user = None
        for user_id, u in users.items():
            if u.email == identifier or u.username == identifier:
                user = u
                break
                
        if not user:
            # If user not found in loaded users, try to load from file directly
            try:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    all_users = json.load(f)
                
                for user_id, user_data in all_users.items():
                    if user_data.get('email') == identifier or user_data.get('username') == identifier:
                        # Create User object from file data
                        user = User(
                            id=user_id,
                            email=user_data['email'],
                            username=user_data['username'],
                            password_hash=user_data['password_hash'],
                            is_verified=user_data.get('is_verified', False),
                            otp_verified=user_data.get('otp_verified', False)
                        )
                        # Add to our users dictionary
                        users[user_id] = user
                        break
            except Exception as e:
                print(f"Error loading user from file: {str(e)}")
                return jsonify({'error': 'Internal server error'}), 500
                
        if not user:
            print(f'User not found in JSON storage for identifier: {identifier}')
            return jsonify({'error': 'Invalid email/username or password'}), 401
            
        if not user.check_password(password):
            print('Password verification failed for JSON user')
            return jsonify({'error': 'Invalid email/username or password'}), 401

        reset_company_selection_session()
        login_user(user)
        clear_login_prompt_flash()
        pricing_mode = 'gm' if has_gm_pricing_access(user) else 'standard'
        set_pricing_mode(pricing_mode)
        redirect_path = '/gm-page' if pricing_mode == 'gm' else '/index'
        print(f'User {user.username} logged in successfully (JSON storage, mode={pricing_mode})')
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'redirectTo': redirect_path,
            'user': {
                'id': user.id,
                'email': user.email,
                'username': user.username
            }
        })
        
    except Exception as e:
        print(f"Unexpected login error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'An unexpected error occurred during login'}), 500

@app.route('/api/auth/logout', methods=['GET', 'POST'])
@app.route('/logout', methods=['GET', 'POST'])  # Add this line to support both /api/auth/logout and /logout
@login_required
def api_logout():
    try:
        logout_user()
        session.clear()  # Clear the session data
        if request.method == 'POST' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Logged out successfully'})
        return redirect(url_for('login'))
    except Exception as e:
        print(f"Logout error: {str(e)}")
        if request.method == 'POST' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Internal server error'}), 500
        flash('An error occurred during logout', 'error')
        return redirect(url_for('login'))  # Always redirect to login on error

@app.route('/api/auth/user', methods=['GET'])
def api_user():
    try:
        if current_user.is_authenticated:
            user_data = {
                'success': True,
                'user': {
                    'id': current_user.id,
                    'email': current_user.email,
                    'username': current_user.username,
                    'company_id': getattr(current_user, 'company_id', None)
                }
            }
            return jsonify(user_data)
        else:
            return jsonify({'error': 'Not logged in'}), 401
    except Exception as e:
        print(f"User error: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/profile/account')
@login_required
def api_profile_account():
    user = current_user
    return jsonify({
        'username': user.username,
        'email': user.email,
        'created_at': user.created_at.strftime('%Y-%m-%d'),
        'company_id': user.company_id,
        'role': user.role if hasattr(user, 'role') else 'user'
    })

@app.route('/api/profile/update', methods=['POST'])
@login_required
def api_profile_update():
    data = request.get_json()
    user_id = current_user.get_id()
    
    if not user_id:
        return jsonify({'error': 'User not found'}), 404
    
    try:
        if MONGO_AVAILABLE and USE_MONGO:
            user = mu_find_user_by_id(user_id)
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            # Update user fields
            if 'username' in data:
                user['username'] = data['username']
            if 'email' in data:
                user['email'] = data['email']
            
            # Save to MongoDB
            mu_update_user(user_id, user)
            
            return jsonify({'success': True, 'message': 'Profile updated successfully'})
        else:
            return jsonify({'error': 'Database not available'}), 500
            
    except Exception as e:
        app.logger.error(f"Error updating profile: {str(e)}")
        return jsonify({'error': 'Failed to update profile'}), 500

# Product pages
@app.route('/mpacks')
@login_required
@company_required
def mpacks():
    # Get company_id from query parameters
    company_id = request.args.get('company_id')
    
    # Initialize company info
    company_name = ''
    company_email = ''
    
    # If company_id is provided in the URL
    if company_id:
        # Try to get company info by ID
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        # Fall back to session data if no company_id in URL
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')
    
    # Update session with final values
    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id
            
    # Log the company info being sent to template
    app.logger.info(f"Rendering mpacks with company: {company_name}, email: {company_email}")
    
    response = render_template('products/chemicals/mpack.html', 
                           current_company={
                               'id': company_id,
                               'name': company_name,
                               'email': company_email
                           })
    
    app.logger.info("Template rendered successfully")
    return response


@app.route('/spray-powder')
@login_required
@company_required
def spray_powder():
    company_id = request.args.get('company_id')
    company_name = ''
    company_email = ''

    if company_id:
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')

    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id

    app.logger.info(f"Rendering spray powder with company: {company_name}, email: {company_email}")

    return render_template(
        'products/spray_powder/spray_powder.html',
        current_company={
            'id': company_id,
            'name': company_name,
            'email': company_email
        }
    )


@app.route('/litho-perforation-rules')
@login_required
@company_required
def litho_perforation_rules():
    return render_company_product_page('products/litho_perforation/litho_perforation.html')


@app.route('/ejection-rubbers')
@login_required
@company_required
def ejection_rubbers():
    return render_company_product_page('products/ejection_rubbers/ejection_rubbers.html')


@app.route('/autowash-cloth')
@login_required
@company_required
def autowash_cloth():
    return render_company_product_page('products/autowash_cloth/autowash_cloth.html')


@app.route('/presspahn')
@login_required
@company_required
def presspahn():
    return render_company_product_page('products/presspahn/presspahn.html')


@app.route('/plotters')
@login_required
@company_required
def plotters():
    return render_company_product_page('products/plotters/plotters.html')


@app.route('/misc-products')
@login_required
@company_required
def misc_products():
    return render_company_product_page('products/misc_products/misc_products.html')


@app.route('/creasing-rule')
@login_required
@company_required
def creasing_rule():
    """Legacy route retained for compatibility – redirect to unified configurator."""
    company_id = request.args.get('company_id')

    if not company_id:
        selected_company = session.get('selected_company') or {}
        company_id = (
            selected_company.get('id')
            or session.get('company_id')
            or getattr(current_user, 'company_id', None)
        )

    redirect_args = {}
    if company_id:
        redirect_args['company_id'] = company_id

    app.logger.info(
        "Redirecting /creasing-rule to /cutting-rule with company_id=%s",
        company_id
    )

    return redirect(url_for('cutting_rule', **redirect_args))


@app.route('/rules')
@app.route('/cutting-rule')
@login_required
@company_required
def cutting_rule():
    company_id = request.args.get('company_id')
    company_name = ''
    company_email = ''

    if company_id:
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')

    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id

    app.logger.info(f"Rendering cutting rule with company: {company_name}, email: {company_email}")

    return render_template(
        'products/cutting_rule/cutting_rule.html',
        current_company={
            'id': company_id,
            'name': company_name,
            'email': company_email
        }
    )


@app.route('/creasing-matrix')
@login_required
@company_required
def creasing_matrix():
    company_id = request.args.get('company_id')
    company_name = ''
    company_email = ''

    if company_id:
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')

    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id

    app.logger.info(f"Rendering creasing matrix with company: {company_name}, email: {company_email}")

    return render_template(
        'products/creasing_matrix/creasing_matrix.html',
        current_company={
            'id': company_id,
            'name': company_name,
            'email': company_email
        }
    )

@app.route('/chemicals')
@app.route('/chemicals-maintenance')
@login_required
@company_required
def chemicals_maintenance():
    # Get company_id from query parameters
    company_id = request.args.get('company_id')

    # Initialize company info
    company_name = ''
    company_email = ''

    # If company_id is provided in the URL
    if company_id:
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')

    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id

    app.logger.info(f"Rendering chemicals-maintenance with company: {company_name}, email: {company_email}")

    return render_template('products/chemicals/chemicals.html',
                           current_company={
                               'id': company_id,
                               'name': company_name,
                               'email': company_email
                           })


@app.route('/blankets')
@login_required
@company_required
def blankets():
    # Get company_id from query parameters
    company_id = request.args.get('company_id')
    
    # Debug log current session
    app.logger.debug(f"Session data: {dict(session)}")
    app.logger.debug(f"Current user: {current_user}")
    
    # Initialize company info
    company_name = ''
    company_email = ''
    
    # If company_id is provided in the URL
    if company_id:
        # Try to get company info by ID
        company_name = get_company_name_by_id(company_id)
        company_email = get_company_email_by_id(company_id)
    else:
        # Fall back to session data if no company_id in URL
        selected_company = session.get('selected_company', {})
        company_name = selected_company.get('name') or session.get('company_name')
        company_email = selected_company.get('email') or session.get('company_email')
        company_id = selected_company.get('id') or session.get('company_id')
    
    # Update session with final values
    session['company_name'] = company_name
    session['company_email'] = company_email
    session['company_id'] = company_id
            
    # Log the company info being sent to template
    app.logger.info(f"Rendering blankets with company: {company_name}, email: {company_email}")
    
    # Create response and set company data in the session cookie
    extended_discount_allowed = has_extended_discount_access(current_user)

    response = make_response(render_template('products/blankets/blankets.html',
                         current_discount_cap=get_restricted_discount_cap(current_user),
                         company_name=company_name,
                         company_email=company_email,
                         company_id=company_id,
                         extended_discount_allowed=extended_discount_allowed,
                         current_company={
                             'id': company_id,
                             'name': company_name,
                             'email': company_email
                         }))
    
    # Set company info in cookies for client-side access
    response.set_cookie('company_name', company_name or '', httponly=True, samesite='Lax')
    response.set_cookie('company_email', company_email or '', httponly=True, samesite='Lax')
    response.set_cookie('company_id', str(company_id) if company_id else '', httponly=True, samesite='Lax')
    
    return response

# Reset password page
@app.route('/reset-password')
def reset_password_page():
    return render_template('reset_password.html')

# Helper functions to get company name and email by ID
def get_company_name_by_id(company_id):
    """Get company name by ID.
    Priority: MongoDB -> JSON fallback."""
    try:
        # Try MongoDB first
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                doc = mongo_db.companies.find_one({'_id': ObjectId(company_id)})
            except Exception:
                doc = mongo_db.companies.find_one({'_id': company_id})
            if doc:
                normalized = {k.lower().replace(' ', ''): v for k, v in doc.items()}
                for key in ['name', 'companyname', 'company_name']:
                    if key in normalized and normalized[key]:
                        return normalized[key]
        # Skip JSON fallback when MongoDB is enabled
        if not (MONGO_AVAILABLE and USE_MONGO and mongo_db is not None):
            # Fallback to JSON file lookup
            file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
            with open(file_path, 'r') as f:
                companies = json.load(f)
                # Convert company_id to int if it's a string
                try:
                    idx = int(company_id) - 1
                    if 0 <= idx < len(companies):
                        return companies[idx].get('Company Name', '')
                except (ValueError, TypeError):
                    # If company_id is not a number, try to find by exact match in ID field
                    for company in companies:
                        if str(company.get('id', '')).lower() == str(company_id).lower():
                            return company.get('Company Name', '')

    except Exception as e:
        app.logger.error(f"Error getting company name: {e}")
    return ''

def get_company_email_by_id(company_id):
    """Get company email by ID.
    Priority: MongoDB -> JSON fallback."""
    try:
        if MONGO_AVAILABLE and USE_MONGO and mongo_db is not None:
            try:
                doc = mongo_db.companies.find_one({'_id': ObjectId(company_id)})
            except Exception:
                doc = mongo_db.companies.find_one({'_id': company_id})
            if doc:
                normalized = {k.lower().replace(' ', ''): v for k, v in doc.items()}
                for key in ['email', 'emailid', 'email_id', 'emailid']:
                    if key in normalized and normalized[key]:
                        return normalized[key]
        # Skip JSON fallback when MongoDB is enabled
        if not (MONGO_AVAILABLE and USE_MONGO and mongo_db is not None):
            file_path = os.path.join(app.root_path, 'static', 'data', 'company_emails.json')
            with open(file_path, 'r') as f:
                companies = json.load(f)
                # Convert company_id to int if it's a string
                try:
                    idx = int(company_id) - 1
                    if 0 <= idx < len(companies):
                        return companies[idx].get('EmailID', '')
                except (ValueError, TypeError):
                    # If company_id is not a number, try to find by exact match in ID field
                    for company in companies:
                        if str(company.get('id', '')).lower() == str(company_id).lower():
                            return company.get('EmailID', '')
    except Exception as e:
        app.logger.error(f"Error getting company email: {e}")
        return ''

# Error handling
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

# Start app
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 3000))
    if os.environ.get('FLASK_ENV') == 'production':
        serve(app, host="0.0.0.0", port=port)
    else:
        app.run(host='0.0.0.0', port=port, debug=True)

# Initialize users dictionary after all function definitions
if USE_MONGO:
    users = {}
else:
    users = load_users()
    print(f"Loaded {len(users)} users from file")
